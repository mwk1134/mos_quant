"""
미국 상장 2x/3x 레버리지 ETF 전체에 대해 백테스팅을 실행하고
결과를 엑셀 보고서로 출력하는 스크립트
"""
import time
import sys

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from backtester_any_ticker import AnyTickerQuantTrader
from backtester_soxl_excel import load_parameters_from_excel, calculate_mdd


ETF_LIST = [
    # === 3X Bull ===
    # 주요 지수
    {"ticker": "TQQQ", "name": "ProShares UltraPro QQQ", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "UPRO", "name": "ProShares UltraPro S&P 500", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "SPXL", "name": "Direxion Daily S&P 500 Bull 3X", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "UDOW", "name": "ProShares UltraPro Dow30", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "TNA", "name": "Direxion Daily Small Cap Bull 3X", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "URTY", "name": "ProShares UltraPro Russell 2000", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "MIDU", "name": "Direxion Daily Mid Cap Bull 3X", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "UMDD", "name": "ProShares UltraPro MidCap400", "leverage": "3x", "direction": "Bull", "category": "지수"},
    {"ticker": "SAA", "name": "ProShares UltraPro SmallCap600", "leverage": "3x", "direction": "Bull", "category": "지수"},
    # 섹터
    {"ticker": "SOXL", "name": "Direxion Daily Semiconductor Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "TECL", "name": "Direxion Daily Technology Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "FAS", "name": "Direxion Daily Financial Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "DPST", "name": "Direxion Daily Regional Banks Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "NAIL", "name": "Direxion Daily Homebuilders Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "DFEN", "name": "Direxion Daily Aerospace & Defense Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "LABU", "name": "Direxion Daily S&P Biotech Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "CURE", "name": "Direxion Daily Healthcare Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "DUSL", "name": "Direxion Daily Industrials Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "WANT", "name": "Direxion Daily Consumer Discretionary Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "RETL", "name": "Direxion Daily Retail Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "WEBL", "name": "Direxion Daily Dow Jones Internet Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "ERX", "name": "Direxion Daily Energy Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "GUSH", "name": "Direxion Daily S&P Oil & Gas E&P Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    {"ticker": "DRN", "name": "Direxion Daily Real Estate Bull 3X", "leverage": "3x", "direction": "Bull", "category": "섹터"},
    # 원자재
    {"ticker": "NUGT", "name": "Direxion Daily Gold Miners Bull 3X", "leverage": "3x", "direction": "Bull", "category": "원자재"},
    {"ticker": "JNUG", "name": "Direxion Daily Jr Gold Miners Bull 3X", "leverage": "3x", "direction": "Bull", "category": "원자재"},
    # 채권
    {"ticker": "TMF", "name": "Direxion Daily 20+ Year Treasury Bull 3X", "leverage": "3x", "direction": "Bull", "category": "채권"},
    # 해외
    {"ticker": "YINN", "name": "Direxion Daily FTSE China Bull 3X", "leverage": "3x", "direction": "Bull", "category": "해외"},
    {"ticker": "BRZU", "name": "Direxion Daily Brazil Bull 3X", "leverage": "3x", "direction": "Bull", "category": "해외"},
    {"ticker": "INDL", "name": "Direxion Daily MSCI India Bull 3X", "leverage": "3x", "direction": "Bull", "category": "해외"},
    {"ticker": "EURL", "name": "Direxion Daily FTSE Europe Bull 3X", "leverage": "3x", "direction": "Bull", "category": "해외"},
    {"ticker": "EDC", "name": "Direxion Daily Emerging Markets Bull 3X", "leverage": "3x", "direction": "Bull", "category": "해외"},
    # ETN
    {"ticker": "FNGU", "name": "MicroSectors FANG+ 3X Leveraged ETN", "leverage": "3x", "direction": "Bull", "category": "ETN"},
    {"ticker": "BULZ", "name": "MicroSectors FANG & Innovation 3X ETN", "leverage": "3x", "direction": "Bull", "category": "ETN"},
    {"ticker": "SHNY", "name": "MicroSectors Gold 3X Leveraged ETN", "leverage": "3x", "direction": "Bull", "category": "ETN"},

    # === 3X Bear ===
    {"ticker": "SQQQ", "name": "ProShares UltraPro Short QQQ", "leverage": "3x", "direction": "Bear", "category": "지수"},
    {"ticker": "SPXU", "name": "ProShares UltraPro Short S&P 500", "leverage": "3x", "direction": "Bear", "category": "지수"},
    {"ticker": "SPXS", "name": "Direxion Daily S&P 500 Bear 3X", "leverage": "3x", "direction": "Bear", "category": "지수"},
    {"ticker": "SDOW", "name": "ProShares UltraPro Short Dow30", "leverage": "3x", "direction": "Bear", "category": "지수"},
    {"ticker": "TZA", "name": "Direxion Daily Small Cap Bear 3X", "leverage": "3x", "direction": "Bear", "category": "지수"},
    {"ticker": "SRTY", "name": "ProShares UltraPro Short Russell 2000", "leverage": "3x", "direction": "Bear", "category": "지수"},
    {"ticker": "SOXS", "name": "Direxion Daily Semiconductor Bear 3X", "leverage": "3x", "direction": "Bear", "category": "섹터"},
    {"ticker": "TECS", "name": "Direxion Daily Technology Bear 3X", "leverage": "3x", "direction": "Bear", "category": "섹터"},
    {"ticker": "FAZ", "name": "Direxion Daily Financial Bear 3X", "leverage": "3x", "direction": "Bear", "category": "섹터"},
    {"ticker": "LABD", "name": "Direxion Daily S&P Biotech Bear 3X", "leverage": "3x", "direction": "Bear", "category": "섹터"},
    {"ticker": "ERY", "name": "Direxion Daily Energy Bear 3X", "leverage": "3x", "direction": "Bear", "category": "섹터"},
    {"ticker": "DRIP", "name": "Direxion Daily S&P Oil & Gas E&P Bear 3X", "leverage": "3x", "direction": "Bear", "category": "섹터"},
    {"ticker": "DUST", "name": "Direxion Daily Gold Miners Bear 3X", "leverage": "3x", "direction": "Bear", "category": "원자재"},
    {"ticker": "JDST", "name": "Direxion Daily Jr Gold Miners Bear 3X", "leverage": "3x", "direction": "Bear", "category": "원자재"},
    {"ticker": "YANG", "name": "Direxion Daily FTSE China Bear 3X", "leverage": "3x", "direction": "Bear", "category": "해외"},
    {"ticker": "EDZ", "name": "Direxion Daily Emerging Markets Bear 3X", "leverage": "3x", "direction": "Bear", "category": "해외"},
    {"ticker": "TMV", "name": "Direxion Daily 20+ Year Treasury Bear 3X", "leverage": "3x", "direction": "Bear", "category": "채권"},
    {"ticker": "TTT", "name": "ProShares UltraPro Short 20+ Year Treasury", "leverage": "3x", "direction": "Bear", "category": "채권"},

    # === 2X Bull ===
    {"ticker": "QLD", "name": "ProShares Ultra QQQ", "leverage": "2x", "direction": "Bull", "category": "지수"},
    {"ticker": "SSO", "name": "ProShares Ultra S&P 500", "leverage": "2x", "direction": "Bull", "category": "지수"},
    {"ticker": "SPUU", "name": "Direxion Daily S&P 500 Bull 2X", "leverage": "2x", "direction": "Bull", "category": "지수"},
    {"ticker": "DDM", "name": "ProShares Ultra Dow30", "leverage": "2x", "direction": "Bull", "category": "지수"},
    {"ticker": "UWM", "name": "ProShares Ultra Russell 2000", "leverage": "2x", "direction": "Bull", "category": "지수"},
    {"ticker": "MVV", "name": "ProShares Ultra MidCap400", "leverage": "2x", "direction": "Bull", "category": "지수"},
    {"ticker": "USD", "name": "ProShares Ultra Semiconductors", "leverage": "2x", "direction": "Bull", "category": "섹터"},
    {"ticker": "ROM", "name": "ProShares Ultra Technology", "leverage": "2x", "direction": "Bull", "category": "섹터"},
    {"ticker": "UYG", "name": "ProShares Ultra Financials", "leverage": "2x", "direction": "Bull", "category": "섹터"},
    {"ticker": "BIB", "name": "ProShares Ultra NASDAQ Biotechnology", "leverage": "2x", "direction": "Bull", "category": "섹터"},
    {"ticker": "DIG", "name": "ProShares Ultra Energy", "leverage": "2x", "direction": "Bull", "category": "섹터"},
    {"ticker": "AGQ", "name": "ProShares Ultra Silver", "leverage": "2x", "direction": "Bull", "category": "원자재"},
    {"ticker": "UGL", "name": "ProShares Ultra Gold", "leverage": "2x", "direction": "Bull", "category": "원자재"},
    {"ticker": "BOIL", "name": "ProShares Ultra Bloomberg Natural Gas", "leverage": "2x", "direction": "Bull", "category": "원자재"},
    {"ticker": "UCO", "name": "ProShares Ultra Bloomberg Crude Oil", "leverage": "2x", "direction": "Bull", "category": "원자재"},
    {"ticker": "CWEB", "name": "Direxion Daily CSI China Internet Bull 2X", "leverage": "2x", "direction": "Bull", "category": "해외"},
    {"ticker": "CHAU", "name": "Direxion Daily CSI 300 China A Share Bull 2X", "leverage": "2x", "direction": "Bull", "category": "해외"},
    {"ticker": "BITU", "name": "ProShares Ultra Bitcoin ETF", "leverage": "2x", "direction": "Bull", "category": "암호화폐"},
    {"ticker": "BITX", "name": "Volatility Shares 2X Bitcoin Strategy ETF", "leverage": "2x", "direction": "Bull", "category": "암호화폐"},
    {"ticker": "FNGO", "name": "MicroSectors FANG+ 2X Leveraged ETN", "leverage": "2x", "direction": "Bull", "category": "ETN"},
    {"ticker": "DGP", "name": "DB Gold Double Long ETN", "leverage": "2x", "direction": "Bull", "category": "ETN"},

    # === 2X Bear ===
    {"ticker": "QID", "name": "ProShares UltraShort QQQ", "leverage": "2x", "direction": "Bear", "category": "지수"},
    {"ticker": "SDS", "name": "ProShares UltraShort S&P 500", "leverage": "2x", "direction": "Bear", "category": "지수"},
    {"ticker": "TBT", "name": "ProShares UltraShort 20+ Year Treasury", "leverage": "2x", "direction": "Bear", "category": "채권"},
    {"ticker": "KOLD", "name": "ProShares UltraShort Bloomberg Natural Gas", "leverage": "2x", "direction": "Bear", "category": "원자재"},
    {"ticker": "SCO", "name": "ProShares UltraShort Bloomberg Crude Oil", "leverage": "2x", "direction": "Bear", "category": "원자재"},
    {"ticker": "ZSL", "name": "ProShares UltraShort Silver", "leverage": "2x", "direction": "Bear", "category": "원자재"},
]


HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

RESULT_COLUMNS = [
    ("티커", 10),
    ("종목명", 42),
    ("레버리지", 8),
    ("방향", 7),
    ("카테고리", 10),
    ("시작일", 12),
    ("종료일", 12),
    ("거래일수", 10),
    ("초기자본($)", 14),
    ("최종자산($)", 14),
    ("총수익률(%)", 12),
    ("연평균수익률(%)", 14),
    ("최대낙폭MDD(%)", 14),
    ("MDD 발생일", 12),
]

FAIL_COLUMNS = [
    ("티커", 10),
    ("종목명", 42),
    ("레버리지", 8),
    ("방향", 7),
    ("카테고리", 10),
    ("실패 사유", 60),
]


def write_header(ws, columns):
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[cell.column_letter].width = col_width


def write_data_row(ws, row_idx, values):
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if isinstance(value, float):
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal="right")
        elif isinstance(value, int):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="center")


def run_all_etf_backtest():
    print("=" * 70)
    print("  ALL ETF 백테스팅 (2x/3x 레버리지 ETF 전체)")
    print("=" * 70)

    try:
        ag_config, sf_config = load_parameters_from_excel("파라미터.xlsx")
    except Exception as e:
        print(f"\n파라미터 로드 실패: {e}")
        return

    initial_capital = 40000   
    start_date = "2025-01-01"
    end_date = datetime.now().strftime("%Y-%m-%d")
    total = len(ETF_LIST)

    print(f"\n투자원금: ${initial_capital:,.0f}")
    print(f"백테스팅 기간: {start_date} ~ {end_date}")
    print(f"대상 ETF: {total}종")
    print(f"예상 소요시간: 약 {total * 20 // 60}~{total * 40 // 60}분")
    print("=" * 70)

    success_results = []
    fail_results = []

    for idx, etf in enumerate(ETF_LIST, 1):
        ticker = etf["ticker"]
        name = etf["name"]
        print(f"\n[{idx}/{total}] {ticker} ({name}) 백테스팅 중...", end=" ", flush=True)

        try:
            trader = AnyTickerQuantTrader(
                ticker=ticker,
                initial_capital=initial_capital,
                sf_config=sf_config,
                ag_config=ag_config,
            )

            # stdout 출력 억제 (백테스팅 내부 로그가 너무 많음)
            import io, contextlib
            with contextlib.redirect_stdout(io.StringIO()):
                result = trader.run_backtest(start_date, end_date)

            if "error" in result:
                reason = result["error"]
                print(f"SKIP - {reason[:60]}")
                fail_results.append({**etf, "reason": reason})
                time.sleep(1)
                continue

            daily_records = result.get("daily_records", [])
            mdd = calculate_mdd(daily_records)
            trading_days = result.get("trading_days", 0)
            final_value = result.get("final_value", 0)
            total_return = result.get("total_return", 0)

            cagr = 0.0
            if trading_days > 0:
                years = trading_days / 252
                if years > 0 and final_value > 0:
                    cagr = ((final_value / initial_capital) ** (1 / years) - 1) * 100

            actual_start = daily_records[0]["date"] if daily_records else start_date
            actual_end = daily_records[-1]["date"] if daily_records else end_date

            success_results.append({
                "ticker": ticker,
                "name": name,
                "leverage": etf["leverage"],
                "direction": etf["direction"],
                "category": etf["category"],
                "start": actual_start,
                "end": actual_end,
                "trading_days": trading_days,
                "initial_capital": initial_capital,
                "final_value": round(final_value, 2),
                "total_return": round(total_return, 2),
                "cagr": round(cagr, 2),
                "mdd": round(mdd.get("mdd_percent", 0), 2),
                "mdd_date": mdd.get("mdd_date", ""),
            })

            print(f"OK  수익률: {total_return:+.2f}%  CAGR: {cagr:+.2f}%  MDD: {mdd.get('mdd_percent',0):.2f}%")

        except Exception as e:
            reason = str(e)[:120]
            print(f"ERROR - {reason}")
            fail_results.append({**etf, "reason": reason})

        time.sleep(2)

    # 총수익률 기준 내림차순 정렬
    success_results.sort(key=lambda x: x["total_return"], reverse=True)

    # 엑셀 보고서 생성
    output_file = f"ETF_백테스팅_결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()

    # 시트 1: 종합 결과
    ws1 = wb.active
    ws1.title = "종합 결과"
    write_header(ws1, RESULT_COLUMNS)
    ws1.freeze_panes = "A2"

    for row_idx, r in enumerate(success_results, 2):
        write_data_row(ws1, row_idx, [
            r["ticker"], r["name"], r["leverage"], r["direction"], r["category"],
            r["start"], r["end"], r["trading_days"],
            r["initial_capital"], r["final_value"],
            r["total_return"], r["cagr"], r["mdd"], r["mdd_date"],
        ])

    # 시트 2: 실패 목록
    ws2 = wb.create_sheet("실패 목록")
    write_header(ws2, FAIL_COLUMNS)
    ws2.freeze_panes = "A2"

    for row_idx, f in enumerate(fail_results, 2):
        write_data_row(ws2, row_idx, [
            f["ticker"], f["name"], f["leverage"], f["direction"],
            f["category"], f["reason"],
        ])

    wb.save(output_file)
    print("\n" + "=" * 70)
    print(f"  보고서 저장 완료: {output_file}")
    print(f"  성공: {len(success_results)}종 / 실패: {len(fail_results)}종")
    print("=" * 70)


if __name__ == "__main__":
    run_all_etf_backtest()
