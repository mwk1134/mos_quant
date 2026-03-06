"""
파라미터.xlsx 파일에서 파라미터를 읽어서 백테스팅을 실행하는 스크립트 (임의 티커 지원)
초기자산 40000달러, 백테스팅 기간은 실행 시 수동으로 설정 가능
"""
import openpyxl
from datetime import datetime, timedelta
from soxl_quant_system import SOXLQuantTrader


def safe_float(value, default=None):
    """안전하게 float로 변환 (None, 빈 문자열, '-' 등 처리)"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip()
    if value_str == '' or value_str == '-' or value_str.lower() == 'none':
        return default
    try:
        return float(value_str)
    except (ValueError, TypeError):
        return default

def safe_int(value, default=None):
    """안전하게 int로 변환 (None, 빈 문자열, '-' 등 처리)"""
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    value_str = str(value).strip()
    if value_str == '' or value_str == '-' or value_str.lower() == 'none':
        return default
    try:
        return int(float(value_str))
    except (ValueError, TypeError):
        return default


def load_parameters_from_excel(excel_file: str = "파라미터.xlsx"):
    """
    엑셀 파일에서 파라미터 읽기
    Args:
        excel_file: 엑셀 파일 경로
    Returns:
        tuple: (ag_config, sf_config)
    """
    import os
    try:
        # 파일 존재 확인
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_file}")
        
        print(f"📂 엑셀 파일 읽기: {excel_file}")
        file_size = os.path.getsize(excel_file)
        file_mtime = datetime.fromtimestamp(os.path.getmtime(excel_file))
        print(f"   파일 크기: {file_size} bytes")
        print(f"   파일 수정 시간: {file_mtime.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 엑셀 파일 열기
        # data_only=False: 파일이 열려있어도 최신 값을 읽을 수 있음
        # read_only=True: 읽기 전용 모드 (더 빠르고 메모리 효율적)
        # 파일이 열려있으면 read_only=False로 자동 전환
        try:
            # 먼저 read_only=True로 시도
            wb = openpyxl.load_workbook(excel_file, data_only=False, read_only=True)
            print("   ✅ 읽기 전용 모드로 열기 성공")
        except (PermissionError, Exception) as e:
            # 파일이 열려있거나 잠겨있는 경우 read_only=False로 재시도
            print(f"   ⚠️ 읽기 전용 모드 실패 ({str(e)[:50]}), 일반 모드로 재시도...")
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=False, read_only=False)
                print("   ✅ 일반 모드로 열기 성공 (파일이 열려있을 수 있음)")
            except Exception as e2:
                raise PermissionError(f"엑셀 파일을 읽을 수 없습니다. 파일을 닫고 다시 시도해주세요: {excel_file}\n오류: {e2}")
        
        ws = wb.active
        print(f"📋 활성 시트: {ws.title}")
        
        # 공세모드(AG) 파라미터 읽기
        # 엑셀에 소수점(0.035)으로 저장되어 있으면 퍼센트(3.5)로 변환
        ag_buy_threshold_raw = safe_float(ws['B11'].value)  # 공세모드 매수임계값
        ag_sell_threshold_raw = safe_float(ws['B12'].value)  # 공세모드 매도임계값
        print(f"🔍 B11 원본값: {ws['B11'].value} → 변환: {ag_buy_threshold_raw}")
        print(f"🔍 B12 원본값: {ws['B12'].value} → 변환: {ag_sell_threshold_raw}")
        
        # 값이 1보다 작으면 소수점 값이므로 100을 곱해서 퍼센트로 변환
        ag_buy_threshold = ag_buy_threshold_raw * 100 if ag_buy_threshold_raw and ag_buy_threshold_raw < 1 else ag_buy_threshold_raw
        ag_sell_threshold = ag_sell_threshold_raw * 100 if ag_sell_threshold_raw and ag_sell_threshold_raw < 1 else ag_sell_threshold_raw
        ag_max_hold_days = safe_int(ws['B13'].value)  # 공세모드 최대보유일
        ag_split_count = safe_int(ws['B14'].value)  # 공세모드 분할횟수
        
        # 안전모드(SF) 파라미터 읽기
        # 엑셀에 소수점(0.035)으로 저장되어 있으면 퍼센트(3.5)로 변환
        sf_buy_threshold_raw = safe_float(ws['B15'].value)  # 안전모드 매수임계값
        sf_sell_threshold_raw = safe_float(ws['B16'].value)  # 안전모드 매도임계값
        print(f"🔍 B15 원본값: {ws['B15'].value} → 변환: {sf_buy_threshold_raw}")
        print(f"🔍 B16 원본값: {ws['B16'].value} → 변환: {sf_sell_threshold_raw}")
        
        # 값이 1보다 작으면 소수점 값이므로 100을 곱해서 퍼센트로 변환
        sf_buy_threshold = sf_buy_threshold_raw * 100 if sf_buy_threshold_raw and sf_buy_threshold_raw < 1 else sf_buy_threshold_raw
        sf_sell_threshold = sf_sell_threshold_raw * 100 if sf_sell_threshold_raw and sf_sell_threshold_raw < 1 else sf_sell_threshold_raw
        sf_max_hold_days = safe_int(ws['B17'].value)  # 안전모드 최대보유일
        sf_split_count = safe_int(ws['B18'].value)  # 안전모드 분할횟수
        
        # 필수 파라미터 검증
        if ag_buy_threshold is None or ag_sell_threshold is None:
            raise ValueError("공세모드 매수/매도 임계값이 필요합니다 (B11, B12)")
        if sf_buy_threshold is None or sf_sell_threshold is None:
            raise ValueError("안전모드 매수/매도 임계값이 필요합니다 (B15, B16)")
        
        # 공세모드 회차별 비중 읽기 (B21~B28)
        ag_split_ratios = []
        print(f"📊 공세모드 회차별 비중 읽기 (B21~B28):")
        for row in range(21, 29):  # B21 ~ B28
            cell_value = ws[f'B{row}'].value
            ratio = safe_float(cell_value)
            if ratio is not None:
                ag_split_ratios.append(ratio)
                print(f"   B{row}: {cell_value} → {ratio}")
            else:
                print(f"   B{row}: 빈칸")
        
        # 안전모드 회차별 비중 읽기 (B29~B36, 빈칸이 있으면 매수하지 않음)
        sf_split_ratios = []
        print(f"📊 안전모드 회차별 비중 읽기 (B29~B36):")
        for row in range(29, 37):  # B29 ~ B36
            cell_value = ws[f'B{row}'].value
            ratio = safe_float(cell_value)
            if ratio is not None:
                sf_split_ratios.append(ratio)
                print(f"   B{row}: {cell_value} → {ratio}")
            else:
                print(f"   B{row}: 빈칸 (이후 무시)")
                # 빈칸이 있으면 그 이후는 무시 (매수하지 않음)
                break
        
        # 공세모드 설정
        ag_config = {
            "buy_threshold": ag_buy_threshold,
            "sell_threshold": ag_sell_threshold,
            "max_hold_days": ag_max_hold_days,
            "split_count": len(ag_split_ratios) if ag_split_count is None else ag_split_count,
            "split_ratios": ag_split_ratios
        }
        
        # 안전모드 설정
        sf_config = {
            "buy_threshold": sf_buy_threshold,
            "sell_threshold": sf_sell_threshold,
            "max_hold_days": sf_max_hold_days,
            "split_count": len(sf_split_ratios) if sf_split_count is None else sf_split_count,
            "split_ratios": sf_split_ratios
        }
        
        print("\n✅ 파라미터 로드 완료")
        print(f"   공세모드: 매수 {ag_config['buy_threshold']}%, 매도 {ag_config['sell_threshold']}%, 보유일 {ag_config['max_hold_days']}일, 분할 {ag_config['split_count']}회")
        print(f"   안전모드: 매수 {sf_config['buy_threshold']}%, 매도 {sf_config['sell_threshold']}%, 보유일 {sf_config['max_hold_days']}일, 분할 {sf_config['split_count']}회")
        print(f"   공세모드 비중 합: {sum(ag_split_ratios):.4f}")
        print(f"   안전모드 비중 합: {sum(sf_split_ratios):.4f}")
        
        wb.close()  # 파일 닫기
        
        return ag_config, sf_config
        
    except Exception as e:
        print(f"❌ 엑셀 파일 읽기 오류: {e}")
        import traceback
        traceback.print_exc()
        print("\n💡 팁:")
        print("   1. 엑셀 파일이 열려있으면 닫고 다시 시도해주세요.")
        print("   2. 엑셀 파일을 저장했는지 확인해주세요.")
        print("   3. 파일 경로가 올바른지 확인해주세요.")
        raise


def validate_date(date_str: str) -> bool:
    """
    날짜 형식 검증 (YYYY-MM-DD)
    Args:
        date_str: 검증할 날짜 문자열
    Returns:
        bool: 유효한 날짜 형식이면 True
    """
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def calculate_mdd(daily_records):
    """
    최대 낙폭(MDD) 계산
    Args:
        daily_records: 일별 기록 리스트
    Returns:
        dict: MDD 정보
    """
    if not daily_records:
        return {"mdd_percent": 0.0, "mdd_date": "", "mdd_value": 0.0}
    
    peak = daily_records[0]['total_assets']
    max_drawdown = 0.0
    mdd_date = ""
    mdd_value = 0.0
    
    for record in daily_records:
        total_assets = record['total_assets']
        if total_assets > peak:
            peak = total_assets
        
        drawdown = (peak - total_assets) / peak * 100
        if drawdown > max_drawdown:
            max_drawdown = drawdown
            mdd_date = record['date']
            mdd_value = total_assets
    
    return {
        "mdd_percent": max_drawdown,
        "mdd_date": mdd_date,
        "mdd_value": mdd_value
    }


class AnyTickerQuantTrader(SOXLQuantTrader):
    """임의 티커를 지원하는 퀀트 트레이더"""
    
    def __init__(self, ticker: str, initial_capital: float = 40000, sf_config=None, ag_config=None):
        """
        초기화
        Args:
            ticker: 거래할 티커 심볼 (예: "SOXL", "SHNY", "TQQQ" 등)
            initial_capital: 투자원금
            sf_config: SF 모드 설정
            ag_config: AG 모드 설정
        """
        self.ticker = ticker.upper()  # 티커를 대문자로 변환
        # super().__init__ 내부에서 RSI 업데이트 시 get_stock_data가 호출되므로,
        # 부모 __init__ 호출 전에 원본 메서드를 미리 바인딩해둠
        self._original_get_stock_data = SOXLQuantTrader.get_stock_data.__get__(self, type(self))
        super().__init__(initial_capital, sf_config, ag_config)
    
    def get_stock_data(self, symbol: str, period: str = "1mo"):
        """
        주식 데이터 가져오기 (SOXL 요청을 입력받은 티커로 리다이렉트)
        """
        # SOXL 요청을 입력받은 티커로 변경
        if symbol == "SOXL":
            symbol = self.ticker
        
        # 원본 메서드 호출
        return self._original_get_stock_data(symbol, period)


def main():
    """메인 실행 함수"""
    print("=" * 60)
    print("🚀 파라미터 기반 백테스팅 (임의 티커 지원)")
    print("=" * 60)
    
    # 티커 입력 받기
    ticker = input("\n📈 거래할 티커를 입력하세요 (예: SOXL, SHNY, TQQQ 등): ").strip().upper()
    if not ticker:
        print("❌ 티커를 입력해주세요.")
        return
    
    print(f"📈 선택된 티커: {ticker}")
    
    # 백테스팅 기간 입력 받기
    default_start_date = "2011-01-01"
    default_end_date = "2025-12-07"
    
    print(f"\n📅 백테스팅 기간 설정")
    print(f"   기본값: {default_start_date} ~ {default_end_date}")
    print(f"   (엔터를 누르면 기본값 사용)")
    
    # 시작일 입력
    start_date_input = input(f"\n   시작일을 입력하세요 (YYYY-MM-DD, 기본값: {default_start_date}): ").strip()
    if not start_date_input:
        start_date = default_start_date
    else:
        if not validate_date(start_date_input):
            print(f"❌ 잘못된 날짜 형식입니다. YYYY-MM-DD 형식으로 입력해주세요.")
            return
        start_date = start_date_input
    
    # 종료일 입력
    end_date_input = input(f"   종료일을 입력하세요 (YYYY-MM-DD, 기본값: {default_end_date}): ").strip()
    if not end_date_input:
        end_date = default_end_date
    else:
        if not validate_date(end_date_input):
            print(f"❌ 잘못된 날짜 형식입니다. YYYY-MM-DD 형식으로 입력해주세요.")
            return
        end_date = end_date_input
    
    # 시작일이 종료일보다 늦으면 오류
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    if start_dt > end_dt:
        print(f"❌ 시작일({start_date})이 종료일({end_date})보다 늦습니다.")
        return
    
    print(f"\n✅ 설정된 기간: {start_date} ~ {end_date}")
    
    # 엑셀 파일에서 파라미터 읽기
    try:
        ag_config, sf_config = load_parameters_from_excel("파라미터.xlsx")
    except Exception as e:
        print(f"❌ 파라미터 로드 실패: {e}")
        return
    
    # 기본 설정값
    initial_capital = 40000  # 투자원금 4만 달러
    
    print(f"\n💰 투자원금: ${initial_capital:,.0f}")
    print(f"📅 투자기간: {start_date} ~ {end_date}")
    
    # 트레이더 초기화 (파라미터 적용)
    print("\n🔄 트레이더 초기화 중...")
    print(f"   적용될 공세모드 설정: {ag_config}")
    print(f"   적용될 안전모드 설정: {sf_config}")
    trader = AnyTickerQuantTrader(
        ticker=ticker,
        initial_capital=initial_capital,
        sf_config=sf_config,
        ag_config=ag_config
    )
    
    # 트레이더에 실제로 적용된 설정 확인
    print(f"\n✅ 트레이더 초기화 완료")
    print(f"   티커: {trader.ticker}")
    print(f"   실제 공세모드 설정: {trader.ag_config}")
    print(f"   실제 안전모드 설정: {trader.sf_config}")
    
    # RSI 참조 데이터 확인
    print("\n🔍 주간 RSI 데이터 확인 중...")
    try:
        rsi_ref_data = trader.load_rsi_reference_data()
        if not rsi_ref_data or len(rsi_ref_data) == 0:
            print("❌ 주간 RSI 데이터가 없습니다!")
            print("   백테스팅을 실행하기 전에 주간 RSI 데이터를 업데이트해주세요.")
            print("   웹앱을 실행하거나 update_rsi_data.py를 실행하여 RSI 데이터를 생성하세요.")
            return
        
        # 시작 날짜의 주차 RSI 확인 (start_dt는 이미 위에서 정의됨)
        # 시작일이 속한 주의 금요일 계산 (soxl_quant_system.py와 동일한 로직)
        days_until_friday = (4 - start_dt.weekday()) % 7  # 금요일(4)까지의 일수
        if days_until_friday == 0 and start_dt.weekday() != 4:  # 금요일이 아닌데 계산이 0이면 다음 주 금요일
            days_until_friday = 7
        start_week_friday = start_dt + timedelta(days=days_until_friday)
        
        # 시작 주차의 RSI 확인
        start_rsi = trader.get_rsi_from_reference(start_week_friday, rsi_ref_data)
        if start_rsi is None:
            print(f"❌ 백테스팅 시작일({start_date})의 주간 RSI 데이터가 없습니다!")
            print(f"   필요한 주차: {start_week_friday.strftime('%Y-%m-%d')} (금요일)")
            print("   백테스팅을 실행하기 전에 주간 RSI 데이터를 업데이트해주세요.")
            print("   웹앱을 실행하거나 update_rsi_data.py를 실행하여 RSI 데이터를 생성하세요.")
            return
        
        # 1주전, 2주전 RSI도 확인 (모드 결정에 필요)
        prev_week_friday = start_week_friday - timedelta(days=7)
        two_weeks_ago_friday = start_week_friday - timedelta(days=14)
        
        prev_rsi = trader.get_rsi_from_reference(prev_week_friday, rsi_ref_data)
        two_weeks_ago_rsi = trader.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
        
        if prev_rsi is None or two_weeks_ago_rsi is None:
            print(f"❌ 백테스팅 시작일({start_date})의 1주전 또는 2주전 RSI 데이터가 없습니다!")
            if prev_rsi is None:
                print(f"   누락된 주차: {prev_week_friday.strftime('%Y-%m-%d')} (1주전 금요일)")
            if two_weeks_ago_rsi is None:
                print(f"   누락된 주차: {two_weeks_ago_friday.strftime('%Y-%m-%d')} (2주전 금요일)")
            print("   백테스팅을 실행하기 전에 주간 RSI 데이터를 업데이트해주세요.")
            print("   웹앱을 실행하거나 update_rsi_data.py를 실행하여 RSI 데이터를 생성하세요.")
            return
        
        print(f"✅ 주간 RSI 데이터 확인 완료")
        print(f"   시작 주차 ({start_week_friday.strftime('%Y-%m-%d')}): RSI {start_rsi:.2f}")
        print(f"   1주전 ({prev_week_friday.strftime('%Y-%m-%d')}): RSI {prev_rsi:.2f}")
        print(f"   2주전 ({two_weeks_ago_friday.strftime('%Y-%m-%d')}): RSI {two_weeks_ago_rsi:.2f}")
        
    except Exception as e:
        print(f"❌ RSI 데이터 확인 중 오류 발생: {e}")
        print("   백테스팅을 실행하기 전에 주간 RSI 데이터를 업데이트해주세요.")
        print("   웹앱을 실행하거나 update_rsi_data.py를 실행하여 RSI 데이터를 생성하세요.")
        return
    
    # 백테스팅 실행
    print("\n📊 백테스팅 실행 중...")
    backtest_result = trader.run_backtest(start_date, end_date)
    
    if "error" in backtest_result:
        print(f"\n❌ 백테스팅 실패: {backtest_result['error']}")
        print("\n💡 해결 방법:")
        print("   1. 웹앱을 실행하여 RSI 데이터를 자동으로 업데이트하세요.")
        print("   2. 또는 update_rsi_data.py를 실행하여 RSI 데이터를 수동으로 업데이트하세요.")
        return
    
    # MDD 계산
    mdd_info = calculate_mdd(backtest_result['daily_records'])
    
    # 실제 데이터가 존재한 기간 계산
    daily_records = backtest_result['daily_records']
    if daily_records and len(daily_records) > 0:
        actual_start_date = daily_records[0]['date']
        actual_end_date = daily_records[-1]['date']
    else:
        actual_start_date = backtest_result['start_date']
        actual_end_date = backtest_result['end_date']
    
    # 결과 출력
    print("\n" + "=" * 60)
    print(f"📊 백테스팅 결과 ({ticker})")
    print("=" * 60)
    print(f"기간: {actual_start_date} ~ {actual_end_date}")
    print(f"거래일수: {backtest_result['trading_days']}일")
    print(f"초기자본: ${backtest_result['initial_capital']:,.0f}")
    print(f"최종자산: ${backtest_result['final_value']:,.0f}")
    print(f"총수익률: {backtest_result['total_return']:+.2f}%")
    print(f"최대 MDD: {mdd_info.get('mdd_percent', 0.0):.2f}%")
    print(f"MDD 발생일: {mdd_info.get('mdd_date', 'N/A')}")
    print(f"최종보유포지션: {backtest_result['final_positions']}개")
    print(f"총 거래일수: {len(backtest_result['daily_records'])}일")
    
    # 연평균 수익률 계산
    if backtest_result['trading_days'] > 0:
        years = backtest_result['trading_days'] / 252  # 연간 거래일 약 252일
        if years > 0:
            annual_return = ((backtest_result['final_value'] / backtest_result['initial_capital']) ** (1 / years) - 1) * 100
            print(f"연평균 수익률: {annual_return:+.2f}%")
    
    print("=" * 60)


if __name__ == "__main__":
    main()
