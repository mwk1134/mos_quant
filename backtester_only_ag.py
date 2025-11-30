"""
공세(AG) 모드만으로 백테스트를 수행하는 전용 스크립트.

`soxl_quant_system.SOXLQuantTrader`는 내부적으로 SOXL 심볼을 사용하도록
구현되어 있으므로, 이 모듈에서는 원하는 티커로 매핑하는 어댑터를 제공하고
동시에 안전(SF) 모드를 사실상 비활성화하여 공세(AG) 모드만 동작하도록 구성한다.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from soxl_quant_system import SOXLQuantTrader


class AggressiveOnlyTrader(SOXLQuantTrader):
    """
    SF(안전) 모드를 사실상 비활성화하여 AG(공세) 모드만 동작하도록 하는 트레이더.
    - 매우 비현실적인 SF 매수/매도 임계값으로 설정하여 SF 진입 자체가 발생하지 않도록 함.
    """

    def __init__(self, initial_capital: float = 40_000, ag_config: Optional[Dict] = None):
        # 사용자가 ag_config를 넘기면 반영, 그렇지 않으면 기본값은 SOXLQuantTrader의 기본을 사용
        aggressive_config = ag_config.copy() if ag_config is not None else None

        # SF 모드는 진입이 불가능하도록 극단적으로 높게 설정
        disabled_sf_config = {
            "buy_threshold": 999.0,
            "sell_threshold": 999.0,
            "max_hold_days": 1,
            "split_count": 1,
            "split_ratios": [1.0],
        }
        super().__init__(initial_capital=initial_capital, sf_config=disabled_sf_config, ag_config=aggressive_config)


class TickerMappedAggressiveTrader(AggressiveOnlyTrader):
    """
    내부적으로 SOXL을 참조하는 호출을 지정한 티커로 매핑하고,
    동시에 공세(AG) 모드만 동작하도록 하는 래퍼 클래스.
    """

    def __init__(self, target_symbol: str, *args, **kwargs):
        self._target_symbol = target_symbol.upper()
        super().__init__(*args, **kwargs)

    def get_stock_data(self, symbol: str, period: str = "1mo"):
        if symbol.upper() == "SOXL":
            symbol = self._target_symbol
        return super().get_stock_data(symbol, period)


def run_backtest(
    start_date: str,
    end_date: Optional[str] = None,
    initial_capital: float = 40_000,
    *,
    symbol: str = "SOXL",
    quiet: bool = False,
    ag_config: Optional[Dict] = None,
) -> Dict:
    """
    공세(AG) 모드만으로 백테스트를 수행하는 헬퍼 함수.
    """
    symbol = symbol.upper()
    trader_class = AggressiveOnlyTrader if symbol == "SOXL" else TickerMappedAggressiveTrader

    if trader_class is TickerMappedAggressiveTrader:
        trader = trader_class(symbol, initial_capital=initial_capital, ag_config=ag_config)
    else:
        trader = trader_class(initial_capital=initial_capital, ag_config=ag_config)

    if quiet:
        trader.session_start_date = start_date
        result = trader.simulate_from_start_to_today(start_date, quiet=True)
        if result.get("skipped"):
            result.update({"initial_capital": initial_capital})
        result.setdefault("symbol", symbol)
        return result

    result = trader.run_backtest(start_date, end_date)
    result.setdefault("symbol", symbol)
    return result


def _calculate_mdd(daily_records: List[Dict]) -> Dict:
    """
    일별 기록으로부터 MDD(Maximum Drawdown) 정보를 계산한다.
    soxl_quant_system.SOXLQuantTrader.calculate_mdd와 동일한 로직/구조를 유지한다.
    """
    if not daily_records:
        return {
            "mdd_percent": 0.0,
            "mdd_date": "",
            "mdd_value": 0.0,
            "mdd_peak_date": "",
            "overall_peak_date": "",
            "overall_peak_value": 0.0,
        }

    max_assets = 0.0
    max_drawdown = 0.0
    mdd_peak_date = ""
    mdd_date = ""
    mdd_value = 0.0

    overall_max_assets = 0.0
    overall_peak_date = ""

    current_peak_assets = 0.0
    current_peak_date = ""

    for record in daily_records:
        current_assets = float(record.get("total_assets", 0.0) or 0.0)
        record_date = record.get("date", "")

        if current_assets > overall_max_assets:
            overall_max_assets = current_assets
            overall_peak_date = record_date

        if current_assets > current_peak_assets:
            current_peak_assets = current_assets
            current_peak_date = record_date

        if current_peak_assets > 0:
            drawdown = (current_peak_assets - current_assets) / current_peak_assets * 100
            if drawdown > max_drawdown:
                max_drawdown = drawdown
                mdd_date = record_date
                mdd_value = current_assets
                mdd_peak_date = current_peak_date

    return {
        "mdd_percent": max_drawdown,
        "mdd_date": mdd_date,
        "mdd_value": mdd_value,
        "mdd_peak_date": mdd_peak_date,
        "overall_peak_date": overall_peak_date,
        "overall_peak_value": overall_max_assets,
    }


def print_summary(result: Dict) -> None:
    """
    `run_backtest` 결과를 간단하게 요약 출력합니다.
    """
    if "error" in result:
        print(f"❌ 백테스트 실패: {result['error']}")
        return

    symbol = result.get("symbol", "SOXL")
    start = result.get("start_date", "?")
    end = result.get("end_date", datetime.now().strftime("%Y-%m-%d"))
    final_value = result.get("final_value")
    total_return = result.get("total_return")
    trading_days = result.get("trading_days")
    daily_records = result.get("daily_records", [])
    mdd_info = _calculate_mdd(daily_records)

    print("\n📊 백테스트 결과 요약 (공세모드 전용)")
    print("-" * 40)
    print(f"티커        : {symbol}")
    print(f"기간        : {start} ~ {end}")
    if trading_days is not None:
        print(f"거래일 수   : {trading_days}일")
    if final_value is not None:
        print(f"최종 자산   : ${final_value:,.0f}")
    if total_return is not None:
        print(f"총 수익률   : {total_return:+.2f}%")
    if mdd_info.get("mdd_percent", 0.0):
        print(f"MDD         : {mdd_info['mdd_percent']:.2f}%")
        if mdd_info.get("mdd_date"):
            print(f"   발생일   : {mdd_info['mdd_date']}")
            print(f"   기준 최고일: {mdd_info['mdd_peak_date']}")
            print(f"   최저 자산: ${mdd_info['mdd_value']:,.0f}")
    print("-" * 40)


def export_backtest_to_excel(result: Dict, filename: Optional[str] = None) -> Optional[Path]:
    """
    백테스트 결과를 엑셀 파일로 저장한다. (공세모드 전용)
    soxl_quant_system.SOXLQuantTrader.export_backtest_to_excel와 동일한 포맷을 사용한다.
    """
    if "error" in result:
        print(f"❌ 엑셀 내보내기 실패: {result['error']}")
        return None

    daily_records = result.get("daily_records", [])
    if not daily_records:
        print("⚠️ 저장할 일별 기록이 없습니다.")
        return None

    symbol = result.get("symbol", "SOXL")
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{symbol}_백테스팅_AG전용_{result.get('start_date', 'unknown')}_{timestamp}.xlsx"

    output_path = Path(filename).resolve()

    wb = openpyxl.Workbook()

    center_alignment = Alignment(horizontal="center", vertical="center")

    ws_summary = wb.active
    ws_summary.title = "백테스팅 요약"
    ws_summary.freeze_panes = "A2"

    mdd_info = _calculate_mdd(daily_records)

    summary_data = [
        [f"{symbol} 백테스팅 결과 (공세모드 전용)", ""],
        ["", ""],
        ["시작일", result.get("start_date", "")],
        ["종료일", result.get("end_date", "")],
        ["거래일수", f"{result.get('trading_days', 0)}일"],
        ["", ""],
        ["초기자본", f"${result.get('initial_capital', 0):,.0f}"],
        ["최종자산", f"${result.get('final_value', 0):,.0f}"],
        ["총수익률", f"{result.get('total_return', 0.0):+.2f}%"],
        ["최종보유포지션", f"{result.get('final_positions', 0)}개"],
        ["", ""],
        ["=== 리스크 지표 ===", ""],
        ["MDD (최대낙폭)", f"{mdd_info.get('mdd_percent', 0.0):.2f}%"],
        ["MDD 발생일", mdd_info.get("mdd_date", "")],
        ["최저자산", f"${mdd_info.get('mdd_value', 0.0):,.0f}"],
        ["MDD 발생 최고자산일", mdd_info.get("mdd_peak_date", "")],
        ["최고자산일", mdd_info.get("overall_peak_date", "")],
        ["최고자산", f"${mdd_info.get('overall_peak_value', 0.0):,.0f}"],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell1 = ws_summary.cell(row=row_idx, column=1, value=label)
        cell2 = ws_summary.cell(row=row_idx, column=2, value=value)
        cell1.alignment = center_alignment
        cell2.alignment = center_alignment

    title_font = Font(size=16, bold=True)
    title_cell = ws_summary.cell(row=1, column=1)
    title_cell.font = title_font
    title_cell.alignment = center_alignment

    ws_detail = wb.create_sheet("매매 상세내역")
    ws_detail.freeze_panes = "A2"

    headers = [
        "날짜", "주차", "RSI", "모드", "현재회차", "1회시드",
        "매수주문가", "종가", "매도목표가", "손절예정일", "거래일수",
        "매수체결", "수량", "매수대금", "매도일", "매도체결", "보유기간",
        "보유", "실현손익", "누적실현", "당일실현", "예수금", "총자산",
    ]

    header_font = Font(size=11, bold=True)
    header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    prev_close_price = None

    for row_idx, record in enumerate(daily_records, 2):
        cell_date = ws_detail.cell(row=row_idx, column=1, value=record.get("date", ""))
        cell_date.alignment = center_alignment
        if row_idx == 2 or "(월" in str(record.get("date", "")):
            cell_date.font = Font(bold=True)

        ws_detail.cell(row=row_idx, column=2, value=record.get("week", 0)).alignment = center_alignment

        rsi_value = record.get("rsi", 0.0) or 0.0
        cell_rsi = ws_detail.cell(row=row_idx, column=3, value=f"{rsi_value:.2f}")
        cell_rsi.alignment = center_alignment

        mode_value = record.get("mode", "")
        cell_mode = ws_detail.cell(row=row_idx, column=4, value=mode_value)
        cell_mode.alignment = center_alignment
        if mode_value == "SF":
            cell_mode.font = Font(color="008000")
        elif mode_value == "AG":
            cell_mode.font = Font(color="FF8C00")

        ws_detail.cell(row=row_idx, column=5, value=record.get("current_round", 0)).alignment = center_alignment

        seed_amount = record.get("seed_amount", 0.0) or 0.0
        cell_seed = ws_detail.cell(row=row_idx, column=6, value=f"${seed_amount:,.0f}" if seed_amount else "")
        cell_seed.alignment = center_alignment

        buy_order_price = record.get("buy_order_price", 0.0) or 0.0
        cell_buy_order = ws_detail.cell(row=row_idx, column=7, value=f"${buy_order_price:.2f}")
        cell_buy_order.alignment = center_alignment

        close_price = record.get("close_price", 0.0) or 0.0
        cell_close = ws_detail.cell(row=row_idx, column=8, value=f"{close_price:.2f}")
        cell_close.alignment = center_alignment
        if prev_close_price is not None:
            if close_price > prev_close_price:
                cell_close.font = Font(color="FF0000")
            elif close_price < prev_close_price:
                cell_close.font = Font(color="0000FF")
        prev_close_price = close_price

        sell_target = record.get("sell_target_price", 0.0) or 0.0
        cell_sell_target = ws_detail.cell(row=row_idx, column=9, value=f"${sell_target:.2f}")
        cell_sell_target.alignment = center_alignment

        ws_detail.cell(row=row_idx, column=10, value=record.get("stop_loss_date", "")).alignment = center_alignment
        ws_detail.cell(row=row_idx, column=11, value=record.get("trading_days", 0)).alignment = center_alignment

        buy_executed_price = record.get("buy_executed_price", 0.0) or 0.0
        cell_buy_exec = ws_detail.cell(
            row=row_idx,
            column=12,
            value=f"${buy_executed_price:.2f}" if buy_executed_price else "",
        )
        cell_buy_exec.alignment = center_alignment
        if buy_executed_price:
            cell_buy_exec.font = Font(color="FF0000")

        buy_quantity = record.get("buy_quantity", 0) or 0
        cell_buy_qty = ws_detail.cell(row=row_idx, column=13, value=buy_quantity if buy_quantity else "")
        cell_buy_qty.alignment = center_alignment
        if buy_quantity:
            cell_buy_qty.font = Font(color="FF0000")

        buy_amount = record.get("buy_amount", 0.0) or 0.0
        cell_buy_amount = ws_detail.cell(
            row=row_idx,
            column=14,
            value=f"${buy_amount:,.0f}" if buy_amount else "",
        )
        cell_buy_amount.alignment = center_alignment
        if buy_amount:
            cell_buy_amount.font = Font(color="FF0000")

        cell_sell_date = ws_detail.cell(row=row_idx, column=15, value=record.get("sell_date", ""))
        cell_sell_date.alignment = center_alignment
        if record.get("sell_date"):
            cell_sell_date.font = Font(color="0000FF")

        sell_executed_price = record.get("sell_executed_price", 0.0) or 0.0
        cell_sell_exec = ws_detail.cell(
            row=row_idx,
            column=16,
            value=f"${sell_executed_price:.2f}" if sell_executed_price else "",
        )
        cell_sell_exec.alignment = center_alignment
        if sell_executed_price:
            cell_sell_exec.font = Font(color="0000FF")

        holding_days = record.get("holding_days", 0) or 0
        ws_detail.cell(
            row=row_idx,
            column=17,
            value=f"{holding_days}일" if holding_days else "",
        ).alignment = center_alignment

        ws_detail.cell(row=row_idx, column=18, value=record.get("holdings", 0)).alignment = center_alignment

        realized_pnl = record.get("realized_pnl", 0.0) or 0.0
        cell_realized = ws_detail.cell(
            row=row_idx,
            column=19,
            value=f"${realized_pnl:,.0f}" if realized_pnl else "",
        )
        cell_realized.alignment = center_alignment

        cumulative_realized = record.get("cumulative_realized", 0.0) or 0.0
        cell_cum_realized = ws_detail.cell(
            row=row_idx,
            column=20,
            value=f"${cumulative_realized:,.0f}",
        )
        cell_cum_realized.alignment = center_alignment
        cell_cum_realized.font = Font(color="FF0000")

        daily_realized = record.get("daily_realized", 0.0) or 0.0
        cell_daily_realized = ws_detail.cell(
            row=row_idx,
            column=21,
            value=f"${daily_realized:,.0f}" if daily_realized else "",
        )
        cell_daily_realized.alignment = center_alignment

        cash_balance = record.get("cash_balance", 0.0) or 0.0
        cell_cash = ws_detail.cell(row=row_idx, column=22, value=f"${cash_balance:,.0f}")
        cell_cash.alignment = center_alignment

        total_assets = record.get("total_assets", 0.0) or 0.0
        cell_total_assets = ws_detail.cell(row=row_idx, column=23, value=total_assets)
        cell_total_assets.alignment = center_alignment
        cell_total_assets.number_format = "#,##0"

    for ws in [ws_summary, ws_detail]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value = str(cell.value) if cell.value is not None else ""
                if len(value) > max_length:
                    max_length = len(value)
            ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

    wb.save(output_path)
    print(f"✅ 백테스트 결과가 저장되었습니다: {output_path}")
    return output_path


if __name__ == "__main__":
    print("⚔️ 레버리지 ETF 백테스터 (공세모드 전용)")
    print("===============================")

    symbol_input = input("백테스트할 티커를 입력하세요 (기본 SOXL): ").strip().upper()
    symbol = symbol_input or "SOXL"

    start_input = input("시작 날짜를 입력하세요 (YYYY-MM-DD, 기본: 1년 전): ").strip()
    if start_input:
        start_date = start_input
    else:
        start_date = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")

    end_input = input("종료 날짜를 입력하세요 (YYYY-MM-DD, 기본: 오늘): ").strip()
    end_date = end_input or None

    try:
        capital_input = input("초기 투자금을 입력하세요 (미입력 시 40000): ").strip()
        initial_capital = float(capital_input) if capital_input else 40_000
    except ValueError:
        print("유효하지 않은 값이 입력되어 기본 투자금 40,000달러를 사용합니다.")
        initial_capital = 40_000

    result = run_backtest(
        start_date=start_date,
        end_date=end_date,
        initial_capital=initial_capital,
        symbol=symbol,
        )
    print_summary(result)

    export_choice = input("엑셀 파일로 저장할까요? (y/N): ").strip().lower()
    if export_choice == "y":
        export_backtest_to_excel(result)



