import requests
import json
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

import sys
import io
from contextlib import redirect_stdout
from pathlib import Path


class SOXLQuantTrader:
    """SOXL 퀀트투자 시스템"""

    
    def _resolve_data_path(self, filename: str) -> Path:
        base_dir = Path(__file__).resolve().parent
        data_dir = base_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        return data_dir / filename

    def load_rsi_reference_data(self, filename: str = "weekly_rsi_reference.json") -> dict:
        """
        RSI 참조 데이터 로드 (JSON 형식)
        Args:
            filename: RSI 참조 파일명
        Returns:
            dict: RSI 참조 데이터
        """
        try:
            # PyInstaller 실행파일에서 파일 경로 처리
            if getattr(sys, 'frozen', False):
                # 실행파일로 실행된 경우
                if hasattr(sys, '_MEIPASS'):
                    # PyInstaller의 임시 폴더
                    application_path = sys._MEIPASS
                else:
                    # 일반 실행파일
                    application_path = os.path.dirname(sys.executable)
                file_path = os.path.join(application_path, filename)
            else:
                # 스크립트로 실행된 경우
                file_path = str(self._resolve_data_path(filename))
            
            # data 폴더가 없으면 생성
            data_dir = os.path.dirname(file_path)
            if data_dir and not os.path.exists(data_dir):
                os.makedirs(data_dir, exist_ok=True)
                print(f"📁 {data_dir} 폴더 생성 완료")
            
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    rsi_data = json.load(f)
                
                # 메타데이터 출력
                metadata = rsi_data.get('metadata', {})
                total_weeks = metadata.get('total_weeks', 0)
                last_updated = metadata.get('last_updated', 'Unknown')
                
                print(f"[INFO] RSI 참조 데이터 로드 완료")
                print(f"   - 파일 경로: {file_path}")
                print(f"   - 총 {len(rsi_data)-1}개 연도 데이터 ({total_weeks}주차)")
                print(f"   - 마지막 업데이트: {last_updated}")
                
                return rsi_data
            else:
                print(f"⚠️ RSI 참조 파일이 없습니다: {file_path}")
                return {}
        except Exception as e:
            print(f"[ERROR] RSI 참조 데이터 로드 오류: {e}")
            return {}
    
    def get_rsi_from_reference(self, date: datetime, rsi_data: dict) -> float:
        """
        특정 날짜의 RSI 값을 참조 데이터에서 가져오기 (JSON 형식)
        JSON 파일 전체에서 해당 날짜를 찾는 강력한 검색 로직
        Args:
            date: 확인할 날짜
            rsi_data: RSI 참조 데이터 (JSON)
        Returns:
            float: RSI 값 (없으면 None)
        """
        try:
            if not rsi_data:
                return None
            
            date_str = date.strftime('%Y-%m-%d')
            
            # 1단계: 모든 연도에서 해당 날짜가 포함되는 주차 찾기
            available_years = [y for y in rsi_data.keys() if y != 'metadata']
            available_years.sort(reverse=True)  # 최신 연도부터 검색
            
            for year in available_years:
                if 'weeks' not in rsi_data[year]:
                    continue
                    
                weeks = rsi_data[year]['weeks']
                
                # 해당 날짜가 포함되는 주차 찾기
                for week_data in weeks:
                    start_date = week_data['start']
                    end_date = week_data['end']
                    if start_date <= date_str <= end_date:
                        return float(week_data['rsi'])
            
            # 2단계: 정확한 주차가 없으면 가장 가까운 이전 주차의 RSI 사용
            # 단, 7일 이상 차이나면 다른 주차이므로 None 반환 (실시간 계산 유도)
            all_weeks = []
            for year in available_years:
                if 'weeks' not in rsi_data[year]:
                    continue
                for week_data in rsi_data[year]['weeks']:
                    week_data_copy = week_data.copy()
                    week_data_copy['year'] = year
                    all_weeks.append(week_data_copy)
            
            # 종료일 기준으로 정렬
            all_weeks.sort(key=lambda x: x['end'])
            
            # 해당 날짜보다 이전의 가장 가까운 주차 찾기 (7일 이내만)
            for week_data in reversed(all_weeks):
                if week_data['end'] <= date_str:
                    gap_days = (datetime.strptime(date_str, '%Y-%m-%d') - datetime.strptime(week_data['end'], '%Y-%m-%d')).days
                    if gap_days < 7:
                        return float(week_data['rsi'])
                    else:
                        return None
            
            return None
        except Exception as e:
            print(f"[ERROR] RSI 참조 데이터 조회 오류: {e}")
            return None
    
    def check_and_update_rsi_data(self, filename: str = "weekly_rsi_reference.json") -> bool:
        """
        RSI 참조 데이터가 최신인지 확인하고 필요시 업데이트 (JSON 형식)
        최신 주간 RSI 값이 비어있는지도 확인하여 자동 업데이트
        Args:
            filename: RSI 참조 파일명
        Returns:
            bool: 업데이트 성공 여부 (True: 최신 상태, False: 업데이트 필요)
        """
        try:
            today = datetime.now()
            current_year = today.strftime('%Y')
            
            # PyInstaller 실행파일에서 파일 경로 처리
            if getattr(sys, 'frozen', False):
                # 실행파일로 실행된 경우
                application_path = os.path.dirname(sys.executable)
                file_path = os.path.join(application_path, filename)
            else:
                # 스크립트로 실행된 경우
                file_path = str(self._resolve_data_path(filename))
            
            # data 폴더가 없으면 생성
            data_dir = os.path.dirname(file_path)
            if data_dir and not os.path.exists(data_dir):
                os.makedirs(data_dir, exist_ok=True)
                print(f"📁 {data_dir} 폴더 생성 완료")
            
            # 기존 RSI 데이터 로드
            if os.path.exists(file_path):
                #print(f"🔍 JSON 파일 로드 시도: {file_path}")
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
                
                # 디버깅: 로드된 데이터 구조 확인
                print(f"[SUCCESS] JSON 파일 로드 성공!")
                #print(f"   - 파일 크기: {os.path.getsize(file_path)} bytes")
                #print(f"   - 로드된 키들: {list(existing_data.keys())}")
                #print(f"   - 총 연도 수: {len([k for k in existing_data.keys() if k != 'metadata'])}")
                
                # 2024년, 2025년 데이터 확인
                if '2024' in existing_data:
                    print(f"   - 2024년 데이터: {len(existing_data['2024']['weeks'])}주차")
                if '2025' in existing_data:
                    print(f"   - 2025년 데이터: {len(existing_data['2025']['weeks'])}주차")
                
                metadata = existing_data.get('metadata', {})
                last_updated = metadata.get('last_updated', '')
                
                # 최신 주간 RSI 값 확인 (현재 연도의 가장 최근 주차)
                latest_rsi_missing = False
                if current_year in existing_data and existing_data[current_year].get('weeks'):
                    # 현재 연도의 가장 최근 주차 찾기
                    current_year_weeks = existing_data[current_year]['weeks']
                    if current_year_weeks:
                        # 가장 최근 주차의 종료일 확인
                        latest_week = max(current_year_weeks, key=lambda x: x.get('end', ''))
                        latest_week_end = datetime.strptime(latest_week['end'], '%Y-%m-%d')
                        
                        # 오늘 날짜가 가장 최근 주차 종료일보다 7일 이상 지났으면 업데이트 필요
                        days_since_latest = (today - latest_week_end).days
                        if days_since_latest > 7:
                            print(f"⚠️ 최신 주간 RSI가 {days_since_latest}일 전 데이터입니다. 업데이트가 필요합니다.")
                            latest_rsi_missing = True
                        else:
                            print(f"✅ 최신 주간 RSI 확인: {latest_week['end']} ({days_since_latest}일 전)")
                    else:
                        print("⚠️ 현재 연도 데이터가 비어있습니다. 업데이트가 필요합니다.")
                        latest_rsi_missing = True
                else:
                    print("⚠️ 현재 연도 데이터가 없습니다. 업데이트가 필요합니다.")
                    latest_rsi_missing = True
                
                if last_updated:
                    last_update_date = datetime.strptime(last_updated, '%Y-%m-%d')
                    print(f"📅 RSI 참조 데이터 마지막 업데이트: {last_updated}")
                    
                    # 마지막 업데이트 이후 새로운 금요일(완료된 주차)이 지났는지 확인
                    # 새로운 금요일이 지났으면 해당 주차의 RSI를 계산해야 하므로 업데이트 필요
                    days_since_friday = (today.weekday() - 4) % 7
                    if days_since_friday == 0 and today.weekday() != 4:
                        days_since_friday = 7
                    latest_passed_friday = today - timedelta(days=days_since_friday)
                    
                    if last_update_date >= latest_passed_friday and not latest_rsi_missing:
                        print(f"[SUCCESS] RSI 참조 데이터가 최신 상태입니다. (마지막 완료 금요일: {latest_passed_friday.strftime('%Y-%m-%d')})")
                        return True
                    
                    if latest_rsi_missing:
                        print(f"⚠️ 최신 주간 RSI 값이 비어있어 업데이트가 필요합니다.")
                    else:
                        print(f"⚠️ 마지막 업데이트({last_updated}) 이후 새로운 완료 주차({latest_passed_friday.strftime('%Y-%m-%d')})가 있어 업데이트가 필요합니다.")
                else:
                    print("⚠️ RSI 참조 데이터 메타데이터가 없습니다.")
                    latest_rsi_missing = True
            else:
                print("⚠️ RSI 참조 파일이 없습니다. 전체 데이터 생성이 필요합니다.")
                latest_rsi_missing = True
            
            # 최신 RSI가 비어있거나 업데이트가 필요한 경우 False 반환 (자동 업데이트 트리거)
            if latest_rsi_missing:
                print("\n[INFO] 최신 주간 RSI 값이 비어있어 자동 업데이트를 진행합니다.")
            else:
                print("\n[INFO] RSI 참조 데이터 업데이트가 필요합니다.")
            
            return False
            
        except Exception as e:
            print(f"[ERROR] RSI 데이터 확인 오류: {e}")
            return False
    
    def update_rsi_reference_file(self, filename: str = "weekly_rsi_reference.json") -> bool:
        """
        RSI 참조 파일을 최신 데이터로 업데이트 (JSON 형식)
        오늘 날짜까지의 주간 RSI를 자동으로 계산하여 업데이트
        Args:
            filename: RSI 참조 파일명
        Returns:
            bool: 업데이트 성공 여부
        """
        try:
            print("[INFO] RSI 참조 데이터 업데이트 중...")
            print("[INFO] 오늘 날짜까지의 주간 RSI를 자동 계산하여 업데이트합니다.")
            
            # PyInstaller 실행파일에서 파일 경로 처리
            if getattr(sys, 'frozen', False):
                # 실행파일로 실행된 경우
                application_path = os.path.dirname(sys.executable)
                file_path = os.path.join(application_path, filename)
            else:
                # 스크립트로 실행된 경우
                file_path = str(self._resolve_data_path(filename))
            
            # data 폴더가 없으면 생성
            data_dir = os.path.dirname(file_path)
            if data_dir and not os.path.exists(data_dir):
                os.makedirs(data_dir, exist_ok=True)
                print(f"📁 {data_dir} 폴더 생성 완료")
            
            # 기존 JSON 데이터 로드
            existing_data = {}
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            
            # 현재 연도와 주차 계산
            today = datetime.now()
            current_year = today.strftime('%Y')
            
            # QQQ 데이터 가져오기 (5년 - 정확한 RSI 계산을 위해 충분한 기간)
            print("[INFO] QQQ 데이터 가져오는 중 (5y)...")
            qqq_data = self.get_stock_data("QQQ", "5y")
            if qqq_data is None:
                print("[ERROR] QQQ 데이터를 가져올 수 없습니다.")
                return False
            
            # 주간 데이터로 변환
            weekly_data = qqq_data.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            
            print(f"[INFO] 주간 데이터 {len(weekly_data)}주 계산 완료")
            
            # 현재 연도 데이터 초기화
            if current_year not in existing_data:
                existing_data[current_year] = {
                    "description": f"{current_year}년 주간 RSI 데이터",
                    "weeks": []
                }
            
            # 최근 12주 RSI 계산 및 업데이트
            recent_weeks = weekly_data.tail(12)  # 최근 12주
            
            # 완료된 주차만 필터링: 금요일(주 마감일)이 오늘 이전인 주차만 계산
            # 불완전한 주(금요일이 아직 지나지 않은 주)는 제외하여 모든 거래일 데이터 기반의 정확한 RSI만 저장
            today_ts = pd.Timestamp(today.date())
            recent_weeks = recent_weeks[recent_weeks.index <= today_ts]
            print(f"[INFO] 완료된 주차 필터링: {len(recent_weeks)}주 (오늘 {today.strftime('%Y-%m-%d')} 이전 금요일까지)")
            
            for i, (week_end, week_row) in enumerate(recent_weeks.iterrows()):
                # 해당 주의 시작일 계산 (월요일)
                week_start = week_end - timedelta(days=4)  # 금요일에서 4일 전 = 월요일
                
                # 주차 번호 계산 (해당 연도의 몇 번째 주인지)
                week_num = week_start.isocalendar()[1]
                
                # RSI 계산
                data_until_week = qqq_data[qqq_data.index <= week_end]
                if len(data_until_week) >= 20:  # 충분한 데이터가 있을 때
                    rsi_value = self.calculate_weekly_rsi(data_until_week)
                    if rsi_value is not None:
                        # 기존 데이터에서 해당 주차 찾기
                        week_exists = False
                        for j, existing_week in enumerate(existing_data[current_year]['weeks']):
                            if existing_week['week'] == week_num:
                                # 기존 데이터가 있고 rsi 값이 이미 설정되어 있으면 업데이트 건너뜀 (수동 입력 데이터 보호)
                                if 'rsi' in existing_week and existing_week['rsi'] is not None:
                                    week_exists = True
                                    break
                                
                                # 기존 데이터 업데이트 (rsi가 없는 경우만)
                                existing_data[current_year]['weeks'][j] = {
                                    "start": week_start.strftime('%Y-%m-%d'),
                                    "end": week_end.strftime('%Y-%m-%d'),
                                    "week": week_num,
                                    "rsi": round(rsi_value, 2)
                                }
                                week_exists = True
                                break
                        
                        if not week_exists:
                            # 새로운 주차 데이터 추가
                            existing_data[current_year]['weeks'].append({
                                "start": week_start.strftime('%Y-%m-%d'),
                                "end": week_end.strftime('%Y-%m-%d'),
                                "week": week_num,
                                "rsi": round(rsi_value, 2)
                            })
                        
                        print(f"   주차 {week_num}: {week_start.strftime('%m-%d')} ~ {week_end.strftime('%m-%d')} | RSI: {rsi_value:.2f}")
            
            # 주차별로 정렬
            existing_data[current_year]['weeks'].sort(key=lambda x: x['week'])
            
            # 메타데이터 업데이트
            total_weeks = sum(len(year_data['weeks']) for year, year_data in existing_data.items() if year != 'metadata')
            existing_data['metadata'] = {
                "last_updated": today.strftime('%Y-%m-%d'),
                "total_years": len([k for k in existing_data.keys() if k != 'metadata']),
                "total_weeks": total_weeks,
                "description": "QQQ 주간 RSI 참조 데이터 (14주 Wilder's RSI)"
            }
            
            # JSON 파일로 저장
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=2)
            
            print("[SUCCESS] RSI 참조 데이터 업데이트 완료!")
            print(f"   - {current_year}년 데이터 업데이트")
            print(f"   - 총 {total_weeks}개 주차 데이터")
            print(f"   - 마지막 업데이트: {today.strftime('%Y-%m-%d')}")
            
            return True
            
        except Exception as e:
            print(f"[ERROR] RSI 참조 파일 업데이트 오류: {e}")
            return False
    
    def __init__(self, initial_capital: float = 40000, sf_config: Optional[Dict] = None, ag_config: Optional[Dict] = None):
        """
        초기화
        Args:
            initial_capital: 투자원금 (기본값: 40000달러)
            sf_config: SF 모드 설정 (None이면 기본값 사용)
            ag_config: AG 모드 설정 (None이면 기본값 사용)
        """
        self.initial_capital = initial_capital
        
        # 성능 최적화를 위한 캐시
        self._stock_data_cache = {}  # 주식 데이터 캐시
        self._simulation_cache = {}  # 시뮬레이션 결과 캐시
        
        # 데이터 경고 저장 (Close가 None인 날짜들)
        self._data_warnings = []
        
        # 세션 상태: 사용자 입력 시작일 (파일 저장 없음)
        self.session_start_date: Optional[str] = None
        
        # 테스트용 오늘 날짜 오버라이드 (YYYY-MM-DD)
        self.test_today_override: Optional[str] = None

        self.current_mode = None  # RSI 기준에 따라 동적으로 결정
        
        # 미국 주식 시장 휴장일 목록 (2024-2025)
        self.us_holidays = [
            # 2024년 휴장일
            "2024-01-01",  # New Year's Day
            "2024-01-15",  # Martin Luther King Jr. Day
            "2024-02-19",  # Washington's Birthday
            "2024-03-29",  # Good Friday
            "2024-05-27",  # Memorial Day
            "2024-06-19",  # Juneteenth
            "2024-07-04",  # Independence Day
            "2024-09-02",  # Labor Day
            "2024-11-28",  # Thanksgiving Day
            "2024-12-25",  # Christmas Day
            
            # 2025년 휴장일
            "2025-01-01",  # New Year's Day
            "2025-01-20",  # Martin Luther King Jr. Day
            "2025-02-17",  # Washington's Birthday
            "2025-04-18",  # Good Friday
            "2025-05-26",  # Memorial Day
            "2025-06-19",  # Juneteenth
            "2025-07-04",  # Independence Day
            "2025-09-01",  # Labor Day
            "2025-11-27",  # Thanksgiving Day
            "2025-12-25",  # Christmas Day
            
            # 특별 휴장일
            "2025-01-09",  # Jimmy Carter National Day of Mourning
        ]
        
        # RSI 참조 데이터 확인 및 업데이트 (오류 발생 시에도 계속 진행)
        try:
            if not self.check_and_update_rsi_data():
                print("[INFO] RSI 참조 데이터 업데이트 중...")
                if self.update_rsi_reference_file():
                    print("[SUCCESS] RSI 참조 데이터 업데이트 완료")
                else:
                    print("[ERROR] RSI 참조 데이터 업데이트 실패")
        except Exception as e:
            # RSI 업데이트 실패해도 백테스트는 계속 진행
            print(f"[WARNING] RSI 업데이트 중 오류 발생 (무시하고 계속 진행): {str(e)[:100]}")
        
        # SF모드 설정 (사용자 지정 또는 기본값)
        if sf_config is not None:
            self.sf_config = sf_config.copy()
        else:
            self.sf_config = {
                "buy_threshold": 3.5,   # 전일 종가 대비 +3.5%에 매수 (매수가)
                "sell_threshold": 1.4,  # 전일 종가 대비 +1.4%에 매도 (매도가)
                "max_hold_days": 35,    # 최대 보유기간 35일
                "split_count": 7,       # 7회 분할매수
                "split_ratios": [0.049, 0.127, 0.230, 0.257, 0.028, 0.169, 0.140]
            }
        
        # AG모드 설정 (사용자 지정 또는 기본값)
        if ag_config is not None:
            self.ag_config = ag_config.copy()
        else:
            self.ag_config = {
                "buy_threshold": 3.6,   # 전일 종가 대비 +3.6%에 매수 (매수가)
                "sell_threshold": 3.5,  # 전일 종가 대비 +3.5%에 매도 (매도가)
                "max_hold_days": 7,     # 최대 보유기간 7일
                "split_count": 8,       # 8회 분할매수
                "split_ratios": [0.062, 0.134, 0.118, 0.148, 0.150, 0.182, 0.186, 0.020]
            }
        
        # 포지션 관리 (회차별)
        self.positions = []  # [{"round": 1, "buy_date": date, "buy_price": price, "shares": shares, "amount": amount}]
        self.current_round = 1
        self.available_cash = initial_capital
        

        # 투자원금 관리 (10거래일마다 업데이트)
        self.current_investment_capital = initial_capital
        self.trading_days_count = 0  # 거래일 카운터
        
        # 시드증액 관리
        self.seed_increases = []  # [{"date": "2025-10-21", "amount": 31000, "description": "시드증액"}]
        self.processed_seed_dates = set()  # 이미 처리된 시드증액 날짜 추적
        
        # 세션 상태: 사용자 입력 시작일 (파일 저장 없음)
        self.session_start_date: Optional[str] = None
        
        # 테스트용 오늘 날짜 오버라이드 (YYYY-MM-DD)
        self.test_today_override: Optional[str] = None
        
        # 주차 추적 (모드 전환 제어용)
        self.current_week_friday: Optional[datetime] = None  # 현재 주차의 금요일

    def set_test_today(self, date_str: Optional[str]):
        """테스트용 오늘 날짜 설정/해제. None 또는 빈문자면 해제."""
        if not date_str:
            self.test_today_override = None
            print("🧪 테스트 날짜 해제됨 (실제 오늘 날짜 사용)")
            return
        try:
            # 형식 검증
            _ = datetime.strptime(date_str, "%Y-%m-%d")
            self.test_today_override = date_str
            print(f"🧪 테스트 날짜 설정: {date_str}")
        except ValueError:
            print("❌ 날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력해주세요.")

    def get_today_date(self) -> datetime:
        """오버라이드된 오늘 날짜(자정)를 datetime으로 반환."""
        if self.test_today_override:
            d = datetime.strptime(self.test_today_override, "%Y-%m-%d").date()
            return datetime(d.year, d.month, d.day)
        return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    def add_seed_increase(self, date: str, amount: float, description: str = ""):
        """시드증액 추가"""
        seed_increase = {
            "date": date,
            "amount": amount,
            "description": description
        }
        self.seed_increases.append(seed_increase)
        # 날짜순으로 정렬
        self.seed_increases.sort(key=lambda x: x["date"])
        print(f"시드증액 추가: {date} - ${amount:,.0f} ({description})")
    
    def get_seed_increases_for_date(self, date: str) -> List[Dict]:
        """특정 날짜의 시드증액 목록 반환"""
        return [si for si in self.seed_increases if si["date"] == date]

    def simulate_from_start_to_today(self, start_date: str, quiet: bool = True) -> Dict:
        """
        시작일부터 최근 거래일까지 시뮬레이션 수행하여 현재 포지션 상태를 맞춘다.
        Args:
            start_date: YYYY-MM-DD 형식 시작일
            quiet: 출력 억제 여부 (기본 True)
        Returns:
            Dict: run_backtest 요약 결과
        """
        # 캐시 키 생성 (시작일 + 초기투자금 + 테스트날짜 + 시드증액 정보)
        # 시드증액 정보를 문자열로 변환하여 캐시 키에 포함
        seed_increases_str = ""
        if self.seed_increases:
            # 시드증액을 날짜와 금액 기준으로 정렬하여 문자열로 변환
            sorted_seeds = sorted(self.seed_increases, key=lambda x: x["date"])
            seed_increases_str = "_".join([f"{s['date']}_{s['amount']}" for s in sorted_seeds])
        
        cache_key = f"{start_date}_{self.initial_capital}_{self.test_today_override or 'real'}_{seed_increases_str}"
        
        # 캐시된 결과가 있고 2분 이내면 재사용
        if cache_key in self._simulation_cache:
            cached_result, cache_time = self._simulation_cache[cache_key]
            if (datetime.now() - cache_time).seconds < 30:  # 30초 캐시
                print(f"⚡ 시뮬레이션 결과 캐시에서 로드 ({start_date})")
                return cached_result
        
        latest_trading_day = self.get_latest_trading_day()
        
        # 시드증액 날짜 중 가장 늦은 날짜 확인 (시드증액이 시뮬레이션에 반영되도록)
        latest_seed_date = None
        if self.seed_increases:
            seed_dates = [datetime.strptime(si["date"], "%Y-%m-%d").date() for si in self.seed_increases]
            latest_seed_date = max(seed_dates)
        
        # 시뮬레이션 종료일 결정: 최신 거래일과 시드증액 날짜 중 더 늦은 날짜 사용
        if latest_seed_date and latest_seed_date > latest_trading_day.date():
            # 시드증액 날짜가 더 늦으면, 해당 날짜까지 시뮬레이션 실행
            # (시장이 열려있어도 시드증액 날짜까지는 실행해야 함)
            end_date_str = latest_seed_date.strftime('%Y-%m-%d')
            effective_end_date = latest_seed_date
        else:
            end_date_str = latest_trading_day.strftime('%Y-%m-%d')
            effective_end_date = latest_trading_day.date()

        # 시작일이 최신 거래일보다 늦는 경우(예: 시작일=오늘, 장 미마감으로 최신 거래일=어제)
        # 백테스트를 건너뛰고 포트폴리오만 초기화하여 당일 추천이 가능하도록 처리
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = effective_end_date
            if start_dt > end_dt:
                if not quiet:
                    print(f"⚠️ 백테스트 스킵: 시작일({start_dt})이 종료일({end_dt})보다 늦음")
                self.reset_portfolio()
                minimal_result = {"skipped": True, "start_date": start_date, "end_date": end_date_str}
                self._simulation_cache[cache_key] = (minimal_result, datetime.now())
                return minimal_result
        except Exception:
            pass
        
        if quiet:
            buf = io.StringIO()
            with redirect_stdout(buf):
                result = self.run_backtest(start_date, end_date_str)
        else:
            result = self.run_backtest(start_date, end_date_str)
        
        # 캐시에 저장
        self._simulation_cache[cache_key] = (result, datetime.now())
        
        return result
    
    def is_market_closed(self, date: datetime) -> bool:
        """
        주식 시장 휴장일 확인
        Args:
            date: 확인할 날짜
        Returns:
            bool: 휴장일이면 True, 거래일이면 False
        """
        # 주말 확인 (토요일=5, 일요일=6)
        if date.weekday() >= 5:
            return True
        
        # 휴장일 확인
        date_str = date.strftime("%Y-%m-%d")
        if date_str in self.us_holidays:
            return True
        
        return False

    def _is_dst_approx(self, dt_utc: datetime) -> bool:
        """미국 서머타임 대략 판별 (3~10월은 DST라고 가정)."""
        return 3 <= dt_utc.month <= 10

    def get_us_eastern_now(self) -> datetime:
        """미국 동부시간(ET) 현재시각 (DST 단순 가정)."""
        if self.test_today_override:
            # 테스트 날짜 기준 정오(12:00) ET로 간주
            d = datetime.strptime(self.test_today_override, "%Y-%m-%d")
            return datetime(d.year, d.month, d.day, 12, 0, 0)
        now_utc = datetime.utcnow()
        offset_hours = 4 if self._is_dst_approx(now_utc) else 5
        return now_utc - timedelta(hours=offset_hours)

    def is_regular_session_closed_now(self) -> bool:
        """정규장(09:30~16:00 ET) 기준으로 오늘 장이 마감됐는지 여부."""
        et_now = self.get_us_eastern_now()
        # 거래일이 아니면(주말/휴장일) '이미 마감'으로 간주
        if not self.is_trading_day(et_now):
            return True
        # 16:00 ET 이후면 마감
        return et_now.hour > 16 or (et_now.hour == 16 and et_now.minute >= 0)
    
    def get_latest_trading_day(self) -> datetime:
        """
        가장 최근 거래일 찾기 (미국 시장 마감 기준)
        Returns:
            datetime: 가장 최근 거래일
        """
        # 미국 시장이 마감되었는지 확인
        if self.is_regular_session_closed_now():
            # 시장이 마감되었으면 오늘을 최신 거래일로 사용
            today = self.get_today_date()
            print(f"📊 미국 시장 마감됨 - 오늘을 최신 거래일로 사용: {today.strftime('%Y-%m-%d')}")
            return today
        else:
            # 시장이 아직 열려있으면 어제를 최신 거래일로 사용
            yesterday = self.get_today_date() - timedelta(days=1)
            print(f"📊 미국 시장 개장 중 - 어제를 최신 거래일로 사용: {yesterday.strftime('%Y-%m-%d')}")
            return yesterday
        
        # 데이터를 가져올 수 없는 경우 기존 로직 사용
        today = self.get_today_date()
        while self.is_market_closed(today):
            today -= timedelta(days=1)
        print(f"⚠️ 데이터 없음, 계산된 최신 거래일: {today.strftime('%Y-%m-%d')}")
        return today
        
    def get_stock_data(self, symbol: str, period: str = "1mo") -> Optional[pd.DataFrame]:
        """
        Yahoo Finance API를 통해 주식 데이터 가져오기 (캐싱 적용)
        Args:
            symbol: 주식 심볼 (예: "SOXL", "QQQ")
            period: 기간 (1d, 5d, 1mo, 3mo, 6mo, 1y, 2y, 5y, 10y, ytd, max)
        Returns:
            DataFrame: 주식 데이터 (Date, Open, High, Low, Close, Volume)
        """
        # 캐시 키 생성
        cache_key = f"{symbol}_{period}"
        current_time = datetime.now()
        
        # 캐시된 데이터가 있고 1분 이내면 재사용 (더 자주 업데이트)
        if cache_key in self._stock_data_cache:
            cached_data, cache_time = self._stock_data_cache[cache_key]
            if (current_time - cache_time).seconds < 60:  # 1분 캐시
                print(f"📊 {symbol} 데이터 캐시에서 로드 (기간: {period})")
                return cached_data
        
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            # 15y가 지원되지 않으면 10y로 대체
            if period == "15y":
                # 먼저 15y 시도, 실패하면 10y로 대체
                params_list = [{'range': '15y', 'interval': '1d'}, {'range': '10y', 'interval': '1d'}]
            else:
                params_list = [{'range': period, 'interval': '1d'}]
            
            print(f"[INFO] {symbol} 데이터 가져오는 중...")
            
            # 여러 파라미터 시도
            for i, params in enumerate(params_list):
                try:
                    print(f"   시도 {i+1}/{len(params_list)}: range={params['range']}")
                    response = requests.get(url, headers=headers, params=params, timeout=15)
                    
                    if response.status_code == 200:
                        data = response.json()
                        
                        if 'chart' in data and 'result' in data['chart'] and data['chart']['result']:
                            result = data['chart']['result'][0]
                            
                            if 'timestamp' in result and 'indicators' in result:
                                timestamps = result['timestamp']
                                quote_data = result['indicators']['quote'][0]
                                
                                # 디버깅: 12월 12일 timestamp 확인
                                target_debug_ts = int(datetime(2025, 12, 12, 0, 0, 0).timestamp())
                                target_debug_date = datetime(2025, 12, 12).date()
                                
                                # 12월 12일 timestamp가 있는지 확인
                                dec12_timestamp_found = False
                                dec12_index_in_array = None
                                for i, ts in enumerate(timestamps):
                                    ts_date = datetime.fromtimestamp(ts).date()
                                    if ts_date == target_debug_date:
                                        dec12_timestamp_found = True
                                        dec12_index_in_array = i
                                        print(f"🔍 [DEBUG] 12월 12일 timestamp 발견! 인덱스: {i}, timestamp: {ts}")
                                        break
                                
                                if not dec12_timestamp_found:
                                    print(f"⚠️ [DEBUG] 12월 12일 timestamp가 Yahoo Finance API 응답에 없습니다!")
                                    # 12월 12일 전후 날짜 확인
                                    nearby_dates = []
                                    for ts in timestamps:
                                        ts_date = datetime.fromtimestamp(ts).date()
                                        if abs((ts_date - target_debug_date).days) <= 3:
                                            nearby_dates.append((ts_date, ts))
                                    if nearby_dates:
                                        print(f"   주변 날짜 timestamps:")
                                        for date, ts in sorted(nearby_dates):
                                            print(f"     {date.strftime('%Y-%m-%d')}: {ts}")
                                
                                # DataFrame 생성
                                df_data = {
                                    'Date': [datetime.fromtimestamp(ts) for ts in timestamps],
                                    'Open': quote_data.get('open', [None] * len(timestamps)),
                                    'High': quote_data.get('high', [None] * len(timestamps)),
                                    'Low': quote_data.get('low', [None] * len(timestamps)),
                                    'Close': quote_data.get('close', [None] * len(timestamps)),
                                    'Volume': quote_data.get('volume', [None] * len(timestamps))
                                }
                                
                                df = pd.DataFrame(df_data)
                                
                                # 수동 데이터 보정 적용 (Yahoo Finance API가 제공하지 않는 날짜)
                                # SOXL만 보정 적용 (symbol별로 분리)
                                manual_corrections = {}
                                if symbol == "SOXL":
                                    manual_corrections = {
                                        "2025-12-12": {
                                            "Open": 46.92,
                                            "High": 47.38,
                                            "Low": 41.06,
                                            "Close": 41.71,
                                            "Volume": 138088200
                                        }
                                    }
                                
                                # 원본 API 응답에서 Close가 None인 날짜 감지 및 수동 보정 정보 저장 (수동 보정 전)
                                # 최근 10일만 확인하여 웹앱에서 경고 표시용
                                api_missing_close_dates = []
                                api_original_values = {}  # 날짜별 원본 API 값 저장
                                today = self.get_today_date().date()
                                
                                for idx, row in df.iterrows():
                                    date_obj = row['Date'].date() if hasattr(row['Date'], 'date') else pd.Timestamp(row['Date']).date()
                                    # 최근 10일만 확인
                                    if (today - date_obj).days <= 10:
                                        date_str = row['Date'].strftime('%Y-%m-%d')
                                        original_close = row['Close']
                                        
                                        # Close가 None인 경우 또는 수동 보정이 적용될 날짜인 경우 저장
                                        if pd.isna(original_close) or date_str in manual_corrections:
                                            api_missing_close_dates.append(date_str)
                                            # 원본 API 값 저장 (None이면 None, 값이 있으면 그 값)
                                            api_original_values[date_str] = None if pd.isna(original_close) else float(original_close)
                                
                                # 원본 API 응답 정보 저장 (웹앱 경고용)
                                if api_missing_close_dates:
                                    if not hasattr(self, '_api_missing_close_dates'):
                                        self._api_missing_close_dates = {}
                                    if not hasattr(self, '_api_original_values'):
                                        self._api_original_values = {}
                                    
                                    if symbol not in self._api_missing_close_dates:
                                        self._api_missing_close_dates[symbol] = []
                                    if symbol not in self._api_original_values:
                                        self._api_original_values[symbol] = {}
                                    
                                    # 중복 제거 및 추가
                                    existing_dates = set(self._api_missing_close_dates[symbol])
                                    for date_str in api_missing_close_dates:
                                        if date_str not in existing_dates:
                                            self._api_missing_close_dates[symbol].append(date_str)
                                        # 원본 값 저장 (업데이트)
                                        self._api_original_values[symbol][date_str] = api_original_values.get(date_str)
                                
                                # 수동 보정 정보 저장 (웹앱에서 표시용)
                                if manual_corrections:
                                    if not hasattr(self, '_manual_corrections_info'):
                                        self._manual_corrections_info = {}
                                    if symbol not in self._manual_corrections_info:
                                        self._manual_corrections_info[symbol] = {}
                                    # 수동 보정 정보 저장
                                    for date_str, correction_data in manual_corrections.items():
                                        self._manual_corrections_info[symbol][date_str] = {
                                            'original_close': api_original_values.get(date_str),
                                            'corrected_close': correction_data.get('Close')
                                        }
                                
                                # Close가 None인 날짜 감지 및 경고 (수동 보정 적용 전)
                                missing_close_dates = []
                                for idx, row in df.iterrows():
                                    if pd.isna(row['Close']):
                                        date_str = row['Date'].strftime('%Y-%m-%d')
                                        missing_close_dates.append(date_str)
                                
                                # 수동 보정 적용 (해당 symbol에 대한 보정만 적용)
                                for date_str, correction_data in manual_corrections.items():
                                    for idx, row in df.iterrows():
                                        if row['Date'].strftime('%Y-%m-%d') == date_str:
                                            print(f"✅ [수동 보정] {symbol} {date_str} 데이터 적용: Close=${correction_data['Close']:.2f}")
                                            df.loc[idx, 'Open'] = correction_data.get('Open', row['Open'])
                                            df.loc[idx, 'High'] = correction_data.get('High', row['High'])
                                            df.loc[idx, 'Low'] = correction_data.get('Low', row['Low'])
                                            df.loc[idx, 'Close'] = correction_data['Close']
                                            df.loc[idx, 'Volume'] = correction_data.get('Volume', row['Volume'])
                                            # missing_close_dates에서 제거
                                            if date_str in missing_close_dates:
                                                missing_close_dates.remove(date_str)
                                            break
                                
                                # Close가 None인 날짜가 있으면 경고 출력
                                if missing_close_dates:
                                    print(f"⚠️ [경고] 다음 날짜들의 Close 값이 None입니다: {', '.join(missing_close_dates)}")
                                    print(f"   이 날짜들은 dropna()로 제거됩니다. 수동 보정이 필요할 수 있습니다.")
                                    # 경고를 인스턴스 변수에 저장하여 웹앱에서 접근 가능하도록
                                    if not hasattr(self, '_data_warnings'):
                                        self._data_warnings = []
                                    self._data_warnings.extend(missing_close_dates)
                                
                                # NaN 값 제거 (Close가 None인 행은 제거)
                                df = df.dropna(subset=['Close'])  # Close가 있으면 유효한 거래일로 간주
                                df.set_index('Date', inplace=True)
                                df.index = df.index.normalize()
                                
                                # 캐시에 저장
                                self._stock_data_cache[cache_key] = (df, current_time)
                                
                                print(f"[SUCCESS] {symbol} 데이터 가져오기 성공! ({len(df)}일치 데이터)")
                                return df
                            else:
                                print(f"   ❌ 차트 데이터 구조 오류")
                        else:
                            print(f"   ❌ 차트 결과 없음")
                    else:
                        print(f"   ❌ HTTP 오류: {response.status_code}")
                        
                except Exception as e:
                    print(f"   ❌ 요청 오류: {e}")
                    
                # 마지막 시도가 아니면 계속
                if i < len(params_list) - 1:
                    print(f"   다음 파라미터로 재시도...")
            
            print(f"❌ {symbol} 모든 파라미터 시도 실패")
            return None
                
        except Exception as e:
            print(f"❌ {symbol} 데이터 가져오기 오류: {e}")
            return None
    
    def get_intraday_last_price(self, symbol: str) -> Optional[Tuple[datetime, float]]:
        """
        분봉(1m) 기준으로 오늘의 최신 가격을 가져온다.
        Returns:
            Optional[Tuple[datetime, float]]: (마지막 시각, 마지막 가격)
        """
        try:
            url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            params = {'range': '1d', 'interval': '1m'}
            response = requests.get(url, headers=headers, params=params, timeout=10)
            if response.status_code != 200:
                return None
            data = response.json()
            result = data.get('chart', {}).get('result', [])
            if not result:
                return None
            result0 = result[0]
            timestamps = result0.get('timestamp') or []
            indicators = result0.get('indicators', {})
            quotes = indicators.get('quote', [])
            if not timestamps or not quotes:
                return None
            closes = quotes[0].get('close') or []
            # 마지막 유효 값 찾기
            for ts, close_val in reversed(list(zip(timestamps, closes))):
                if close_val is not None:
                    ts_dt = datetime.utcfromtimestamp(ts)
                    return ts_dt, float(close_val)
            return None
        except Exception:
            return None


    def calculate_weekly_rsi_for_dates(self, target_fridays: list, window: int = 14) -> dict:
        """
        특정 금요일 날짜들에 대한 정확한 주간 RSI를 실시간 계산 (5년 데이터 기반)
        참조 데이터에 없을 때 폴백으로 사용
        Args:
            target_fridays: RSI를 계산할 금요일 날짜 리스트 (datetime)
            window: RSI 계산 기간 (기본값: 14)
        Returns:
            dict: {날짜문자열: RSI값} 딕셔너리
        """
        try:
            print(f"📊 RSI 실시간 계산 시작 (5y 데이터 기반, 대상: {len(target_fridays)}개 주차)")
            
            # 5년치 QQQ 데이터 가져오기 (정확한 RSI 계산을 위해 충분한 기간)
            qqq_long = self.get_stock_data("QQQ", "5y")
            if qqq_long is None:
                print("❌ QQQ 5년 데이터를 가져올 수 없습니다.")
                return {}
            
            # 주간 데이터로 변환 (금요일 기준)
            weekly_df = qqq_long.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            
            if len(weekly_df) < window + 1:
                print(f"❌ 주간 데이터 부족 (필요: {window+1}주, 현재: {len(weekly_df)}주)")
                return {}
            
            # RSI 계산
            delta = weekly_df['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            
            # 대상 금요일 날짜들의 RSI 추출
            result = {}
            for friday in target_fridays:
                friday_dt = pd.Timestamp(friday.date()) if isinstance(friday, datetime) else pd.Timestamp(friday)
                # 해당 금요일 이전 또는 같은 날짜의 가장 가까운 주간 데이터 찾기
                earlier_dates = weekly_df.index[weekly_df.index <= friday_dt]
                if len(earlier_dates) > 0:
                    matched_date = earlier_dates[-1]
                    matched_idx = weekly_df.index.get_loc(matched_date)
                    if matched_idx < len(rsi) and not pd.isna(rsi.iloc[matched_idx]):
                        rsi_value = round(rsi.iloc[matched_idx], 2)
                        result[friday_dt.strftime('%Y-%m-%d')] = rsi_value
                        print(f"   ✅ {friday_dt.strftime('%Y-%m-%d')} → RSI: {rsi_value}")
            
            return result
            
        except Exception as e:
            print(f"❌ RSI 실시간 계산 오류: {e}")
            return {}

    def calculate_weekly_rsi(self, df: pd.DataFrame, window: int = 14) -> float:
        """

        주간 RSI 계산 (제공된 함수 방식 적용)
        Args:
            df: 일일 주가 데이터

            window: RSI 계산 기간 (기본값: 14)
        Returns:
            float: 최신 주간 RSI 값
        """
        try:
            # 주간 데이터로 변환 (금요일 기준)
            weekly_df = df.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            

            # 디버깅: 주간 데이터 확인
            print(f"   주간 데이터 변환 결과:")
            print(f"   - 기간: {weekly_df.index[0].strftime('%Y-%m-%d')} ~ {weekly_df.index[-1].strftime('%Y-%m-%d')}")
            print(f"   - 주간 데이터 수: {len(weekly_df)}주")
            print(f"   - 최근 5주 종가: {weekly_df['Close'].tail(5).values}")
            
            if len(weekly_df) < window + 1:
                print(f"❌ 주간 RSI 계산을 위한 데이터 부족 (필요: {window+1}주, 현재: {len(weekly_df)}주)")
                return None
            

            # RSI 계산
            delta = weekly_df['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=window).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=window).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            

            # 디버깅 정보 출력
            latest_rsi = rsi.iloc[-1]
            print(f"📈 QQQ 주간 RSI: {latest_rsi:.2f}")

            print(f"   데이터 기간: {weekly_df.index[0].strftime('%Y-%m-%d')} ~ {weekly_df.index[-1].strftime('%Y-%m-%d')}")
            print(f"   주간 데이터 수: {len(weekly_df)}주")
            print(f"   최근 3개 RSI: {[f'{x:.2f}' if not np.isnan(x) else 'NaN' for x in rsi.tail(3).values]}")
            
            # 상세 계산 과정 출력
            print(f"   최근 3개 계산 과정:")
            for i in range(-3, 0):
                if i + len(weekly_df) >= 0:
                    date_str = weekly_df.index[i].strftime('%Y-%m-%d')
                    delta_val = delta.iloc[i]
                    gain_val = gain.iloc[i]
                    loss_val = loss.iloc[i]
                    rs_val = rs.iloc[i]
                    rsi_val = rsi.iloc[i]
                    print(f"   {date_str}: delta={delta_val:+.4f}, gain={gain_val:.4f}, loss={loss_val:.4f}, RS={rs_val:.4f}, RSI={rsi_val:.2f}")
            
            return latest_rsi
            
        except Exception as e:
            print(f"❌ 주간 RSI 계산 오류: {e}")
            return None
    

    def _is_mode_case_matched(self, current_rsi: float, prev_rsi: float) -> tuple[bool, str]:
        """
        RSI 값으로 안전모드 또는 공세모드 조건에 해당하는지 확인
        Args:
            current_rsi: 1주전 RSI
            prev_rsi: 2주전 RSI
        Returns:
            tuple: (조건에 해당하는지 여부, 모드("SF" 또는 "AG" 또는 None))
        """
        # 안전모드 조건들 (OR로 연결)
        safe_conditions = [
            # RSI > 65 영역에서 하락 (2주전 RSI > 65이고 2주전 > 1주전)
            prev_rsi > 65 and prev_rsi > current_rsi,
            # 40 < RSI < 50에서 하락 (2주전 RSI가 40~50 사이이고 2주전 > 1주전)
            40 < prev_rsi < 50 and prev_rsi > current_rsi,
            # RSI가 50 밑으로 하락 (2주전 >= 50이고 1주전 < 50)
            prev_rsi >= 50 and current_rsi < 50
        ]
        
        # 공세모드 조건들 (OR로 연결)
        aggressive_conditions = [
            # RSI가 50 위로 상승 (2주전 < 50이고 2주전 < 1주전이고 1주전 > 50)
            prev_rsi < 50 and prev_rsi < current_rsi and current_rsi > 50,
            # 50 < RSI < 60에서 상승 (2주전 RSI가 50~60 사이이고 2주전 < 1주전)
            50 < prev_rsi < 60 and prev_rsi < current_rsi,
            # RSI < 35 영역에서 상승 (2주전 < 35이고 2주전 < 1주전)
            prev_rsi < 35 and prev_rsi < current_rsi
        ]
        
        safe_result = any(safe_conditions)
        aggressive_result = any(aggressive_conditions)
        
        if safe_result:
            return (True, "SF")
        if aggressive_result:
            return (True, "AG")
        return (False, None)
    
    def determine_mode(self, current_rsi: float, prev_rsi: float, prev_mode: str) -> str:
        """
        구글스프레드시트 수식 기반 모드 판단
        Args:
            current_rsi: 1주전 RSI (현재 적용할 RSI)
            prev_rsi: 2주전 RSI (이전 RSI)
            prev_mode: 전주 모드
        Returns:
            str: "SF" (안전모드) 또는 "AG" (공세모드)
        """
        # RSI 값이 None인 경우 백테스팅 중단
        if current_rsi is None or prev_rsi is None:
            raise ValueError(f"RSI 데이터가 없습니다. current_rsi: {current_rsi}, prev_rsi: {prev_rsi}")
        
        # _is_mode_case_matched를 사용하여 조건 확인
        is_matched, matched_mode = self._is_mode_case_matched(current_rsi, prev_rsi)
        
        print(f"🔍 determine_mode 호출: 1주전 RSI={current_rsi:.2f}, 2주전 RSI={prev_rsi:.2f}, 전주모드={prev_mode}")
        
        # 조건에 해당하면 해당 모드 반환
        if is_matched:
            print(f"   → 결과: {matched_mode} (조건에 해당)")
            return matched_mode
        
        # 조건에 없으면 전주 모드 유지
        print(f"   → 결과: {prev_mode} (전주 모드 유지 - 조건에 해당하지 않음)")
        return prev_mode
    
    def _calculate_week_mode_recursive_with_reference(self, target_friday: datetime, rsi_ref_data: dict, max_depth: int = 20) -> tuple[str | None, bool]:
        """
        RSI 참조 데이터를 사용하여 특정 주차의 모드를 재귀적으로 계산
        안전모드 3가지 조건 또는 공세모드 3가지 조건에 해당하는 모드가 나올 때까지 이전 주차를 확인
        Args:
            target_friday: 계산할 주차의 금요일
            rsi_ref_data: RSI 참조 데이터 딕셔너리
            max_depth: 최대 재귀 깊이 (무한 루프 방지)
        Returns:
            tuple: (모드("SF" 또는 "AG"), 성공 여부)
        """
        if max_depth <= 0:
            print(f"❌ 모드판정실패: 최대 재귀 깊이 도달 ({target_friday.strftime('%Y-%m-%d')})")
            return None, False
        
        # 1주전, 2주전 금요일 계산
        one_week_ago_friday = target_friday - timedelta(days=7)
        two_weeks_ago_friday = target_friday - timedelta(days=14)
        
        # RSI 찾기
        one_week_ago_rsi = self.get_rsi_from_reference(one_week_ago_friday, rsi_ref_data)
        two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
        
        # RSI 데이터가 없으면 실시간 계산 폴백
        if one_week_ago_rsi is None or two_weeks_ago_rsi is None:
            try:
                missing = []
                if one_week_ago_rsi is None:
                    missing.append(one_week_ago_friday)
                if two_weeks_ago_rsi is None:
                    missing.append(two_weeks_ago_friday)
                fallback = self.calculate_weekly_rsi_for_dates(missing)
                if one_week_ago_rsi is None:
                    one_week_ago_rsi = fallback.get(one_week_ago_friday.strftime('%Y-%m-%d'))
                if two_weeks_ago_rsi is None:
                    two_weeks_ago_rsi = fallback.get(two_weeks_ago_friday.strftime('%Y-%m-%d'))
            except Exception as e:
                print(f"⚠️ RSI 실시간 계산 폴백 실패: {e}")
        
        if one_week_ago_rsi is None or two_weeks_ago_rsi is None:
            failure_reason = []
            if one_week_ago_rsi is None:
                failure_reason.append(f"1주전 RSI 데이터 없음")
            if two_weeks_ago_rsi is None:
                failure_reason.append(f"2주전 RSI 데이터 없음")
            print(f"❌ 모드판정실패: {target_friday.strftime('%Y-%m-%d')} 주차 모드 계산 불가 - {', '.join(failure_reason)}")
            return None, False
        
        # 조건에 해당하는지 확인
        is_matched, matched_mode = self._is_mode_case_matched(one_week_ago_rsi, two_weeks_ago_rsi)
        
        if is_matched:
            # 조건에 해당하는 모드를 찾았음
            print(f"✅ {target_friday.strftime('%Y-%m-%d')} 주차 모드 계산: 1주전 RSI={one_week_ago_rsi:.2f}, 2주전 RSI={two_weeks_ago_rsi:.2f} → {matched_mode} (조건에 해당)")
            return matched_mode, True
        
        # 조건에 해당하지 않으면 이전 주차의 모드를 재귀적으로 확인
        print(f"🔍 {target_friday.strftime('%Y-%m-%d')} 주차: 조건에 해당하지 않음, 이전 주차 확인 중...")
        prev_week_mode, success = self._calculate_week_mode_recursive_with_reference(one_week_ago_friday, rsi_ref_data, max_depth - 1)
        
        if not success:
            print(f"❌ 모드판정실패: {target_friday.strftime('%Y-%m-%d')} 주차의 이전 주차 모드 계산 실패")
            return None, False
        
        # 이전 주차의 모드를 사용하여 현재 주차의 모드 결정
        final_mode = self.determine_mode(one_week_ago_rsi, two_weeks_ago_rsi, prev_week_mode)
        
        # determine_mode가 prev_week_mode를 반환했는지 확인
        if final_mode == prev_week_mode:
            print(f"✅ {target_friday.strftime('%Y-%m-%d')} 주차 모드: {final_mode} (이전 주차 모드 유지)")
        else:
            print(f"✅ {target_friday.strftime('%Y-%m-%d')} 주차 모드: {final_mode} (이전 주차 모드 {prev_week_mode}에서 변경)")
        
        return final_mode, True
    
    def _calculate_week_mode_recursive(self, target_friday: datetime, weekly_df: pd.DataFrame, rsi: pd.Series, max_depth: int = 20) -> tuple[str | None, bool]:
        """
        특정 주차의 모드를 재귀적으로 계산
        안전모드 3가지 조건 또는 공세모드 3가지 조건에 해당하는 모드가 나올 때까지 이전 주차를 확인
        Args:
            target_friday: 계산할 주차의 금요일
            weekly_df: 주간 데이터프레임
            rsi: RSI 시리즈
            max_depth: 최대 재귀 깊이 (무한 루프 방지)
        Returns:
            tuple: (모드("SF" 또는 "AG"), 성공 여부)
        """
        if max_depth <= 0:
            print(f"❌ 모드판정실패: 최대 재귀 깊이 도달 ({target_friday.strftime('%Y-%m-%d')})")
            return None, False
        
        # 1주전, 2주전 금요일 계산
        one_week_ago_friday = target_friday - timedelta(days=7)
        two_weeks_ago_friday = target_friday - timedelta(days=14)
        
        # RSI 찾기
        one_week_ago_friday_dt = pd.Timestamp(one_week_ago_friday.date())
        two_weeks_ago_friday_dt = pd.Timestamp(two_weeks_ago_friday.date())
        
        # 1주전 RSI 찾기
        one_week_ago_rsi = None
        earlier_dates_1w = weekly_df.index[weekly_df.index <= one_week_ago_friday_dt]
        if len(earlier_dates_1w) > 0:
            one_week_rsi_date = earlier_dates_1w[-1]
            one_week_rsi_idx = weekly_df.index.get_loc(one_week_rsi_date)
            if one_week_rsi_idx < len(rsi):
                one_week_ago_rsi = rsi.iloc[one_week_rsi_idx]
                if pd.isna(one_week_ago_rsi):
                    one_week_ago_rsi = None
        
        # 2주전 RSI 찾기
        two_weeks_ago_rsi = None
        earlier_dates_2w = weekly_df.index[weekly_df.index <= two_weeks_ago_friday_dt]
        if len(earlier_dates_2w) > 0:
            two_weeks_rsi_date = earlier_dates_2w[-1]
            two_weeks_rsi_idx = weekly_df.index.get_loc(two_weeks_rsi_date)
            if two_weeks_rsi_idx < len(rsi):
                two_weeks_ago_rsi = rsi.iloc[two_weeks_rsi_idx]
                if pd.isna(two_weeks_ago_rsi):
                    two_weeks_ago_rsi = None
        
        # RSI 데이터가 없으면 실패
        if one_week_ago_rsi is None or two_weeks_ago_rsi is None:
            failure_reason = []
            if one_week_ago_rsi is None:
                failure_reason.append(f"1주전 RSI 데이터 없음")
            if two_weeks_ago_rsi is None:
                failure_reason.append(f"2주전 RSI 데이터 없음")
            print(f"❌ 모드판정실패: {target_friday.strftime('%Y-%m-%d')} 주차 모드 계산 불가 - {', '.join(failure_reason)}")
            return None, False
        
        # 조건에 해당하는지 확인
        is_matched, matched_mode = self._is_mode_case_matched(one_week_ago_rsi, two_weeks_ago_rsi)
        
        if is_matched:
            # 조건에 해당하는 모드를 찾았음
            print(f"✅ {target_friday.strftime('%Y-%m-%d')} 주차 모드 계산: 1주전 RSI={one_week_ago_rsi:.2f}, 2주전 RSI={two_weeks_ago_rsi:.2f} → {matched_mode} (조건에 해당)")
            return matched_mode, True
        
        # 조건에 해당하지 않으면 이전 주차의 모드를 재귀적으로 확인
        print(f"🔍 {target_friday.strftime('%Y-%m-%d')} 주차: 조건에 해당하지 않음, 이전 주차 확인 중...")
        prev_week_mode, success = self._calculate_week_mode_recursive(one_week_ago_friday, weekly_df, rsi, max_depth - 1)
        
        if not success:
            print(f"❌ 모드판정실패: {target_friday.strftime('%Y-%m-%d')} 주차의 이전 주차 모드 계산 실패")
            return None, False
        
        # 이전 주차의 모드를 사용하여 현재 주차의 모드 결정
        # determine_mode를 호출하지만, 이미 prev_week_mode가 case에 해당하는 모드이므로
        # 조건에 해당하지 않으면 prev_week_mode를 반환할 것임
        final_mode = self.determine_mode(one_week_ago_rsi, two_weeks_ago_rsi, prev_week_mode)
        
        # determine_mode가 prev_week_mode를 반환했는지 확인
        if final_mode == prev_week_mode:
            print(f"✅ {target_friday.strftime('%Y-%m-%d')} 주차 모드: {final_mode} (이전 주차 모드 유지)")
        else:
            print(f"✅ {target_friday.strftime('%Y-%m-%d')} 주차 모드: {final_mode} (이전 주차 모드 {prev_week_mode}에서 변경)")
        
        return final_mode, True
    
    def update_mode(self, qqq_data: pd.DataFrame) -> str:
        """
        QQQ 주간 RSI 기반으로 모드 업데이트
        같은 주 내에서는 모드를 변경하지 않음 (주간 RSI는 금요일 기준으로 고정)
        Args:
            qqq_data: QQQ 주가 데이터
        Returns:
            str: 업데이트된 모드
        """
        try:
            # 현재 날짜 기준으로 이번 주 금요일 계산
            today = self.get_today_date()
            days_until_friday = (4 - today.weekday()) % 7  # 금요일(4)까지의 일수
            if days_until_friday == 0 and today.weekday() != 4:  # 금요일이 아닌데 계산이 0이면 다음 주 금요일
                days_until_friday = 7
            this_week_friday = today + timedelta(days=days_until_friday)
            
            # 같은 주 내에서는 모드 변경하지 않음 (날짜만 비교하여 시간 차이 무시)
            if self.current_week_friday is not None:
                # 날짜만 비교 (시간 부분 무시)
                if isinstance(self.current_week_friday, datetime):
                    current_week_friday_date = self.current_week_friday.date()
                else:
                    current_week_friday_date = self.current_week_friday
                
                if isinstance(this_week_friday, datetime):
                    this_week_friday_date = this_week_friday.date()
                else:
                    this_week_friday_date = this_week_friday
                
                if current_week_friday_date == this_week_friday_date:
                    if self.current_mode:
                        print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode}")
                        return self.current_mode
                    # 모드가 없으면 초기화만 진행
            
            # 금요일이면서 장이 종료되지 않았으면 모드 업데이트하지 않음
            # 주간 RSI 및 모드 업데이트는 금요일 장 종료 후에만 계산되어 업데이트되어야 함
            if today.weekday() == 4:  # 금요일
                if not self.is_regular_session_closed_now():
                    # 금요일 장 종료 전에는 모드를 업데이트하지 않음
                    if self.current_mode:
                        print(f"⏳ 금요일 장 종료 전: {today.strftime('%Y-%m-%d')} 장 종료 전이므로 모드 업데이트하지 않음 (현재 모드: {self.current_mode})")
                        return self.current_mode
                    else:
                        print(f"⏳ 금요일 장 종료 전: {today.strftime('%Y-%m-%d')} 장 종료 전이므로 모드 업데이트하지 않음")
                        return None
            
            # 새로운 주차이거나 초기화인 경우 모드 업데이트
            # (금요일 장 종료 후이거나 금요일이 아닌 경우)
            self.current_week_friday = this_week_friday
            
            # 주간 데이터로 변환
            weekly_df = qqq_data.resample('W-FRI').agg({
                'Open': 'first',
                'High': 'max',
                'Low': 'min',
                'Close': 'last',
                'Volume': 'sum'
            }).dropna()
            
            if len(weekly_df) < 15:
                print("⚠️ 주간 데이터 부족, 현재 모드 유지")
                return self.current_mode
            
            # 제공된 함수 방식으로 RSI 계산
            delta = weekly_df['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            rsi = 100 - (100 / (1 + rs))
            
            # get_daily_recommendation과 동일한 방식으로 1주전, 2주전 금요일 계산
            # 오늘 날짜 기준으로 가장 최근 완료된 주차(지난주 금요일) 찾기
            # 월~금은 모드가 변경되지 말아야 함
            # 같은 주 내에서는 모드를 유지하므로, 모드 결정을 위해 지난주 금요일 기준으로 RSI를 확인
            days_until_friday = (4 - today.weekday()) % 7  # 금요일(4)까지의 일수
            if days_until_friday == 0 and today.weekday() != 4:  # 금요일이 아닌데 계산이 0이면 다음 주 금요일
                days_until_friday = 7
            this_week_friday_calc = today + timedelta(days=days_until_friday)
            # 금요일이어도 같은 주 모드를 유지해야 하므로, 항상 지난주 금요일 기준으로 모드 결정
            latest_completed_friday = this_week_friday_calc - timedelta(days=7)
            
            # RSI 참조 데이터 로드
            rsi_ref_data = {}
            try:
                rsi_file_path = str(self._resolve_data_path("weekly_rsi_reference.json"))
                if os.path.exists(rsi_file_path):
                    with open(rsi_file_path, 'r', encoding='utf-8') as f:
                        rsi_ref_data = json.load(f)
            except Exception as e:
                print(f"⚠️ RSI 참조 데이터 로드 실패: {e}")

            # 1주전과 2주전 금요일 계산
            one_week_ago_friday = latest_completed_friday  # 지난주 금요일 (1주전)
            two_weeks_ago_friday = latest_completed_friday - timedelta(days=7)  # 지지난주 금요일 (2주전)
            
            # RSI 값 추출 (참조 데이터 우선 사용)
            one_week_ago_rsi = self.get_rsi_from_reference(one_week_ago_friday, rsi_ref_data)
            two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
            
            # 참조 데이터에 없으면 5년치 데이터로 정확한 RSI 실시간 계산
            if one_week_ago_rsi is None or two_weeks_ago_rsi is None:
                print(f"⚠️ [update_mode] RSI 참조 데이터 부재 → 5년 데이터 기반 실시간 계산 진행")
                target_fridays = []
                if one_week_ago_rsi is None:
                    target_fridays.append(one_week_ago_friday)
                if two_weeks_ago_rsi is None:
                    target_fridays.append(two_weeks_ago_friday)
                
                realtime_rsi = self.calculate_weekly_rsi_for_dates(target_fridays)
                
                if one_week_ago_rsi is None:
                    one_week_key = one_week_ago_friday.strftime('%Y-%m-%d') if isinstance(one_week_ago_friday, datetime) else pd.Timestamp(one_week_ago_friday).strftime('%Y-%m-%d')
                    one_week_ago_rsi = realtime_rsi.get(one_week_key)
                
                if two_weeks_ago_rsi is None:
                    two_weeks_key = two_weeks_ago_friday.strftime('%Y-%m-%d') if isinstance(two_weeks_ago_friday, datetime) else pd.Timestamp(two_weeks_ago_friday).strftime('%Y-%m-%d')
                    two_weeks_ago_rsi = realtime_rsi.get(two_weeks_key)
            
            if one_week_ago_rsi is None or two_weeks_ago_rsi is None:
                print("⚠️ RSI 계산 실패, 현재 모드 유지")
                return self.current_mode
            
            # 전주 모드를 재귀적으로 계산 (참조 데이터 버전 우선)
            prev_week_mode, success = self._calculate_week_mode_recursive_with_reference(one_week_ago_friday, rsi_ref_data)
            if not success:
                prev_week_mode, success = self._calculate_week_mode_recursive(one_week_ago_friday, weekly_df, rsi)
            
            if not success:
                print(f"❌ 모드판정실패: 전주 모드를 계산할 수 없어 현재 주차의 모드를 결정할 수 없습니다.")
                return None
            
            # 현재 주차의 모드 결정
            is_matched, matched_mode = self._is_mode_case_matched(one_week_ago_rsi, two_weeks_ago_rsi)
            
            if is_matched:
                new_mode = matched_mode
                print(f"✅ 현재 주차 모드: {new_mode} (조건에 해당)")
            else:
                new_mode = prev_week_mode
                print(f"✅ 현재 주차 모드: {new_mode} (전주 모드 유지 - 조건에 해당하지 않음)")
            
            if new_mode != self.current_mode:
                self.current_mode = new_mode
            
            return self.current_mode
            
        except Exception as e:
            print(f"❌ 모드 업데이트 오류: {e}")
            return self.current_mode
    
    def get_current_config(self) -> Dict:
        """현재 모드에 따른 설정 반환"""
        return self.sf_config if self.current_mode == "SF" else self.ag_config
    
    def calculate_buy_sell_prices(self, current_price: float) -> Tuple[float, float]:
        """
        매수/매도 가격 계산
        Args:
            current_price: 현재 주가 (전일 종가)
        Returns:
            Tuple[float, float]: (매수가격, 매도가격)
        """
        config = self.get_current_config()
        

        # 매수가: 전일 종가 대비 상승한 가격 (매수가 > 종가)
        buy_price = current_price * (1 + config["buy_threshold"] / 100)

        # 매도가: 전일 종가 대비 상승한 가격 (매도가 < 종가)
        sell_price = current_price * (1 + config["sell_threshold"] / 100)
        
        return buy_price, sell_price
    
    def calculate_position_size(self, round_num: int) -> float:
        """
        회차별 매수 금액 계산
        Args:
            round_num: 매수 회차 (1부터 시작)
        Returns:
            float: 해당 회차 매수 금액
        """
        config = self.get_current_config()
        
        if round_num <= len(config["split_ratios"]):
            ratio = config["split_ratios"][round_num - 1]
            # 투자원금 사용 (10거래일마다 총자산으로 업데이트됨)
            amount = self.current_investment_capital * ratio
            return amount
        else:
            return 0.0
    

    def calculate_stop_loss_date(self, buy_date: datetime, max_hold_days: int) -> str:
        """
        거래일 기준 손절예정일 계산 (주말 + 미국증시 휴장일 제외)
        Args:
            buy_date: 매수일
            max_hold_days: 최대 보유 거래일 수
        Returns:
            str: 손절예정일 (MM.DD.(요일) 형식)
        """
        try:
            # 요일을 한글로 변환
            weekdays_korean = ['월', '화', '수', '목', '금', '토', '일']
            
            # 거래일 기준으로 날짜 계산 (주말 + 휴장일 제외)
            current_date = buy_date
            trading_days_count = 0
            
            while trading_days_count < max_hold_days:
                current_date += timedelta(days=1)
                
                # 거래일인지 확인 (주말이 아니고 휴장일이 아닌 경우)
                if self.is_trading_day(current_date):
                    trading_days_count += 1
            
            weekday_korean = weekdays_korean[current_date.weekday()]
            return current_date.strftime(f"%m.%d.({weekday_korean})")
            
        except Exception as e:
            print(f"⚠️ 손절예정일 계산 오류: {e}")
            # 오류 시 기본값 반환
            fallback_date = buy_date + timedelta(days=max_hold_days)
            weekday_korean = weekdays_korean[fallback_date.weekday()]
            return fallback_date.strftime(f"%m.%d.({weekday_korean})")
    
    def is_trading_day(self, date: datetime) -> bool:
        """
        해당 날짜가 거래일인지 확인 (주말 + 미국증시 휴장일 제외)
        Args:
            date: 확인할 날짜
        Returns:
            bool: 거래일이면 True, 아니면 False
        """
        # 주말 확인 (토요일=5, 일요일=6)
        if date.weekday() >= 5:
            return False
        
        # 미국증시 휴장일 확인
        date_str = date.strftime("%Y-%m-%d")
        if date_str in self.us_holidays:
            return False
        
        return True
    
    def can_buy_next_round(self) -> bool:
        """다음 회차 매수 가능 여부 확인"""
        config = self.get_current_config()
        
        # 최대 분할매수 횟수 확인
        if self.current_round > config["split_count"]:
            return False
        
        # 예수금이 있으면 가능한 만큼이라도 매수 가능
        # (실제 매수 시 예수금 부족 처리는 execute_buy()에서 수행)
        if self.available_cash <= 0:
            return False
        
        return True
    
    def execute_buy(self, target_price: float, actual_price: float, current_date: datetime, mode: Optional[str] = None) -> bool:
        """
        매수 실행 (목표가 기준 수량으로 계산하여 실제가에 매수)
        Args:
            target_price: 목표 매수가 (수량 계산용)
            actual_price: 실제 매수가 (당일 종가)
            current_date: 매수 날짜
            mode: 매수 시점의 모드 (None이면 self.current_mode 사용)
        Returns:
            bool: 매수 성공 여부
        """
        if not self.can_buy_next_round():
            return False
        
        # 매수 시점의 모드 결정 (매개변수가 없으면 self.current_mode 사용)
        buy_mode = mode if mode is not None else self.current_mode

        # 1회시드 금액 계산
        target_amount = self.calculate_position_size(self.current_round)
        
        # 목표가 기준으로 매수할 수량 계산 (round: 반올림으로 실제 체결 수량과 일치율 향상)
        target_shares = round(target_amount / target_price)  # 목표가 기준 수량
        
        if target_shares <= 0:
            return False
        
        # 실제 매수 금액 계산 (목표 수량 × 실제 가격)
        actual_amount = target_shares * actual_price
        
        # 예수금이 부족한 경우 예수금으로 매수 가능한 수량 재계산
        if actual_amount > self.available_cash:
            max_shares = int(self.available_cash / actual_price)  # 예수금 부족 시에는 int(내림)으로 안전하게
            if max_shares <= 0:
                return False
            actual_shares = max_shares
            actual_amount = actual_shares * actual_price
        else:
            actual_shares = target_shares
        
        # 포지션 추가
        position = {
            "round": self.current_round,
            "buy_date": current_date,
            "buy_price": actual_price,  # 실제 매수가
            "shares": actual_shares,    # 실제 매수 수량
            "target_price": target_price,  # 목표가 (참조용)
            "amount": actual_amount,    # 실제 투자금액
            "mode": buy_mode  # 매수 시점의 모드 저장
        }
        
        # 디버깅: 매수 시점의 모드 확인 및 검증
        if mode is not None and mode != buy_mode:
            print(f"⚠️ execute_buy 모드 불일치: 전달된 모드={mode}, buy_mode={buy_mode}, self.current_mode={self.current_mode}")
        print(f"🔍 execute_buy: 매수일 {current_date.strftime('%Y-%m-%d')}, 전달된 모드: {mode}, 저장할 모드: {buy_mode}, 현재 self.current_mode: {self.current_mode}")
        
        # 포지션에 모드가 제대로 저장되었는지 확인
        if position.get('mode') != buy_mode:
            print(f"❌ CRITICAL: 포지션에 모드 저장 실패! 예상: {buy_mode}, 실제: {position.get('mode')}")
            position['mode'] = buy_mode  # 강제로 수정
        
        self.positions.append(position)
        
        # 저장 후 검증
        saved_position = self.positions[-1]
        if saved_position.get('mode') != buy_mode:
            print(f"❌ CRITICAL: 포지션 저장 후 모드 불일치! 예상: {buy_mode}, 실제: {saved_position.get('mode')}")
            self.positions[-1]['mode'] = buy_mode  # 강제로 수정

        self.available_cash -= actual_amount
        self.current_round += 1  # 매수 성공 시에만 회차 증가
        

        print(f"✅ {self.current_round-1}회차 매수 실행: {actual_shares}주 @ ${actual_price:.2f} (목표가: ${target_price:.2f}, 실제투자: ${actual_amount:,.0f})")
        
        return True
    
    def reconcile_positions_with_close_history(self, soxl_data: pd.DataFrame):
        """
        과거 종가가 매도 목표가를 터치한 포지션을 보정하여 자동 매도 처리한다.
        LOC 매도 특성상 종가가 목표가 이상이면 체결되는 것을 반영하기 위함.
        손절예정일이 지난 포지션도 종가 기준으로 LOC 매도 처리한다.
        장중에는 오늘 날짜 데이터를 제외하고 과거 확정된 데이터만 사용한다.
        Args:
            soxl_data (DataFrame): 최근 SOXL 일별 데이터 (Close 필수)
        """
        if not self.positions or soxl_data is None or len(soxl_data) == 0:
            return

        # 장중에는 오늘 날짜 데이터를 제외 (종가가 확정되지 않았으므로)
        today = self.get_today_date()
        today_date = today.date()
        
        # 정규장이 아직 마감되지 않았고, 데이터에 오늘 날짜가 포함되어 있으면 제외
        if not self.is_regular_session_closed_now() and len(soxl_data) > 0:
            if soxl_data.index.max().date() == today_date:
                soxl_data = soxl_data[soxl_data.index.date < today_date]

        sold_rounds = []
        # 리스트 복사본을 사용하여 반복 중 안전하게 제거
        for position in list(self.positions):
            buy_date = position["buy_date"]

            # 매수일 이후(엄격하게 초과) 데이터만 확인
            # buy_date가 Timestamp인 경우와 datetime인 경우 모두 처리
            if isinstance(buy_date, pd.Timestamp):
                buy_date_for_comparison = buy_date
            elif isinstance(buy_date, datetime):
                buy_date_for_comparison = pd.Timestamp(buy_date)
            else:
                buy_date_for_comparison = pd.Timestamp(buy_date)
            
            future_data = soxl_data[soxl_data.index > buy_date_for_comparison]
            if future_data.empty:
                continue

            position_config = self.sf_config if position["mode"] == "SF" else self.ag_config
            target_price = position["buy_price"] * (1 + position_config["sell_threshold"] / 100)

            # 1. 목표가 도달한 경우 매도
            hit_rows = future_data[future_data["Close"] >= target_price]
            if not hit_rows.empty:
                sell_row = hit_rows.iloc[0]
                sell_date = sell_row.name
                sell_close = sell_row["Close"]

                proceeds = position["shares"] * sell_close
                profit = proceeds - position["amount"]
                profit_rate = (profit / position["amount"]) * 100 if position["amount"] else 0.0

                self.positions.remove(position)
                self.available_cash += proceeds
                sold_rounds.append(position["round"])

                print("🧾 과거 종가 매도 보정 실행 (목표가 도달)")
                print(f"   - 회차: {position['round']}회차")
                print(f"   - 매수일: {buy_date.strftime('%Y-%m-%d')} | 매수가: ${position['buy_price']:.2f}")
                print(f"   - 목표가: ${target_price:.2f}")
                print(f"   - sell_date: {sell_date.strftime('%Y-%m-%d')} | 종가: ${sell_close:.2f}")
                print(f"   - 실현손익: ${profit:,.0f} ({profit_rate:+.2f}%)")
                continue  # 이미 매도 처리했으므로 다음 포지션으로

            # 2. 손절예정일이 지난 경우 종가 기준으로 LOC 매도
            # 손절예정일 계산 (거래일 기준)
            stop_loss_date = buy_date
            trading_days_count = 0
            while trading_days_count < position_config["max_hold_days"]:
                stop_loss_date += timedelta(days=1)
                if self.is_trading_day(stop_loss_date):
                    trading_days_count += 1
            
            # 손절예정일이 지났는지 확인 (손절예정일 포함하여 그 이후 날짜)
            # future_data에서 손절예정일 이후 데이터 확인
            stop_loss_rows = future_data[future_data.index >= stop_loss_date]
            if not stop_loss_rows.empty:
                # 손절예정일의 종가로 매도 (손절예정일이 거래일이 아닐 수 있으므로 가장 가까운 거래일 사용)
                sell_row = stop_loss_rows.iloc[0]
                sell_date = sell_row.name
                sell_close = sell_row["Close"]

                proceeds = position["shares"] * sell_close
                profit = proceeds - position["amount"]
                profit_rate = (profit / position["amount"]) * 100 if position["amount"] else 0.0

                self.positions.remove(position)
                self.available_cash += proceeds
                sold_rounds.append(position["round"])

                print("🧾 과거 종가 매도 보정 실행 (손절예정일 경과)")
                print(f"   - 회차: {position['round']}회차")
                print(f"   - 매수일: {buy_date.strftime('%Y-%m-%d')} | 매수가: ${position['buy_price']:.2f}")
                print(f"   - 손절예정일: {stop_loss_date.strftime('%Y-%m-%d')}")
                print(f"   - sell_date: {sell_date.strftime('%Y-%m-%d')} | 종가: ${sell_close:.2f}")
                print(f"   - 실현손익: ${profit:,.0f} ({profit_rate:+.2f}%)")

        if sold_rounds:
            sold_count = len(sold_rounds)
            self.current_round = max(1, self.current_round - sold_count)
            print(f"🔄 보정 후 current_round: {self.current_round} (총 {sold_count}개 회차 보정 매도)")

    def check_sell_conditions(self, row: pd.Series, current_date: datetime, prev_close: float, return_debug_info: bool = False) -> List[Dict]:
        """
        매도 조건 확인
        Args:
            row: 당일 주가 데이터 (Open, High, Low, Close)
            current_date: 현재 날짜
            prev_close: 전일 종가
            return_debug_info: 디버깅 정보 반환 여부
        Returns:
            List[Dict]: 매도할 포지션 리스트 (return_debug_info=True일 경우 튜플로 (리스트, 디버깅정보) 반환)
        """
        sell_positions = []
        debug_info = []  # 디버깅 정보 저장
        
        # 디버깅: 보유 포지션 수 확인
        print(f"🔍 매도 조건 확인: 보유 포지션 {len(self.positions)}개")
        # 모든 포지션 목록 출력 (디버깅용)
        for idx, pos in enumerate(self.positions):
            buy_date_str = pos['buy_date'].strftime('%Y-%m-%d') if isinstance(pos['buy_date'], (datetime, pd.Timestamp)) else str(pos['buy_date'])
            print(f"   포지션 {idx+1}: {pos['round']}회차, 매수일: {buy_date_str}, 모드: {pos.get('mode', 'N/A')}, 매수가: ${pos.get('buy_price', 0):.2f}")
        
        for position in self.positions:
            buy_date = position["buy_date"]

            # 거래일 기준으로 보유기간 계산
            hold_days = 0
            temp_date = buy_date
            while temp_date < current_date:
                temp_date += timedelta(days=1)
                if self.is_trading_day(temp_date):
                    hold_days += 1
            
            # 해당 포지션의 모드 설정 가져오기
            position_config = self.sf_config if position["mode"] == "SF" else self.ag_config
            
            # 손절예정일 계산 (거래일 기준)
            stop_loss_date = buy_date
            trading_days_count = 0
            while trading_days_count < position_config["max_hold_days"]:
                stop_loss_date += timedelta(days=1)
                if self.is_trading_day(stop_loss_date):
                    trading_days_count += 1

            # 해당 포지션의 매수체결가 기준으로 매도가 계산
            position_buy_price = position["buy_price"]
            sell_price = position_buy_price * (1 + position_config["sell_threshold"] / 100)
            
            # 디버깅: 매도 조건 상세 정보
            daily_close = row['Close']
            buy_date_str = buy_date.strftime('%Y-%m-%d') if isinstance(buy_date, (datetime, pd.Timestamp)) else str(buy_date)
            print(f"   📦 {position['round']}회차 (매수일: {buy_date_str}): 매수가 ${position_buy_price:.2f} → 매도목표가 ${sell_price:.2f} (현재가 ${daily_close:.2f})")
            print(f"      보유기간: {hold_days}일 (최대: {position_config['max_hold_days']}일, 손절예정일: {stop_loss_date.strftime('%Y-%m-%d')})")
            
            # 디버깅 정보 수집
            position_debug = {
                "round": position['round'],
                "buy_date": buy_date_str,
                "mode": position.get('mode', 'N/A'),
                "buy_price": position_buy_price,
                "target_sell_price": sell_price,
                "current_close": daily_close,
                "hold_days": hold_days,
                "max_hold_days": position_config['max_hold_days'],
                "stop_loss_date": stop_loss_date.strftime('%Y-%m-%d'),
                "current_date": current_date.strftime('%Y-%m-%d'),
                "meets_target_price": daily_close >= sell_price,
                "meets_stop_loss_date": current_date >= stop_loss_date,
                "will_sell": False,
                "sell_reason": None
            }
            
            # 1. LOC 매도 조건: 종가가 매도목표가에 도달했을 때 (종가 >= 매도목표가)
            if daily_close >= sell_price:
                print(f"      ✅ 매도 조건 1: 목표가 도달 (${daily_close:.2f} >= ${sell_price:.2f})")
                position_debug["will_sell"] = True
                position_debug["sell_reason"] = "목표가 도달"
                sell_positions.append({
                    "position": position,
                    "reason": "목표가 도달",
                    "sell_price": daily_close,  # 종가에 매도
                    "will_sell": True  # 매도 조건 충족
                })
            
            # 2. 손절예정일 경과 시 매도 (당일 종가에 LOC 매도)
            elif current_date >= stop_loss_date:
                print(f"      ✅ 매도 조건 2: 손절예정일 경과 (현재: {current_date.strftime('%Y-%m-%d')} >= 손절예정일: {stop_loss_date.strftime('%Y-%m-%d')})")
                position_debug["will_sell"] = True
                position_debug["sell_reason"] = f"손절예정일 경과 (보유기간: {hold_days}일)"
                sell_positions.append({
                    "position": position,
                    "reason": f"손절예정일 경과 (보유기간: {hold_days}일)",
                    "sell_price": row['Close'],  # 종가에 LOC 매도
                    "will_sell": True  # 매도 조건 충족
                })
            else:
                # 매도 조건을 만족하지 않아도 매도 추천 리스트에 포함 (보유 중 상태로 표시)
                print(f"      ⏳ 매도 조건 미충족: 종가 ${daily_close:.2f} < 목표가 ${sell_price:.2f}, 손절예정일 미경과 ({current_date.strftime('%Y-%m-%d')} < {stop_loss_date.strftime('%Y-%m-%d')})")
                position_debug["will_sell"] = False
                position_debug["sell_reason"] = "보유 중"
                sell_positions.append({
                    "position": position,
                    "reason": "보유 중 (목표가 미도달)",
                    "sell_price": daily_close,  # 현재 종가 (참고용)
                    "will_sell": False  # 매도 조건 미충족
                })
            
            debug_info.append(position_debug)
        
        # 디버깅: 매도 추천 결과
        if sell_positions:
            print(f"✅ 매도 추천 {len(sell_positions)}건 생성됨")
        else:
            print("❌ 매도 추천 없음")
        
        if return_debug_info:
            return sell_positions, debug_info
        return sell_positions
    

    def execute_sell(self, sell_info: Dict) -> tuple:
        """
        매도 실행
        Args:
            sell_info: 매도 정보
        Returns:

            tuple: (매도 수익금, 매도된 회차)
        """
        position = sell_info["position"]
        sell_price = sell_info["sell_price"]

        sold_round = position["round"]
        
        proceeds = position["shares"] * sell_price
        profit = proceeds - position["amount"]
        profit_rate = (profit / position["amount"]) * 100
        
        # 포지션 제거
        self.positions.remove(position)
        self.available_cash += proceeds
        

        print(f"✅ {sold_round}회차 매도 실행: {position['shares']}주 @ ${sell_price:.2f}")
        print(f"   매도 사유: {sell_info['reason']}")
        print(f"   수익: ${profit:,.0f} ({profit_rate:+.2f}%)")
        

        return proceeds, sold_round
    
    def update_position(self, position_index: int, new_shares: int, new_buy_price: Optional[float] = None) -> bool:
        """
        포지션 수정 (수동 수정용)
        Args:
            position_index: 수정할 포지션의 인덱스 (positions 리스트의 인덱스)
            new_shares: 새로운 주식 수량
            new_buy_price: 새로운 매수가 (None이면 기존 매수가 유지)
        Returns:
            bool: 수정 성공 여부
        """
        # 인덱스 범위 확인
        if position_index < 0 or position_index >= len(self.positions):
            return False
        
        # 해당 인덱스의 포지션 가져오기
        position = self.positions[position_index]
        
        # 기존 값 저장
        old_shares = position["shares"]
        old_amount = position["amount"]
        old_buy_price = position["buy_price"]
        
        # 새로운 매수가 설정 (없으면 기존 가격 유지)
        if new_buy_price is None:
            new_buy_price = old_buy_price
        
        # 새로운 투자금액 계산
        new_amount = new_shares * new_buy_price
        
        # 예수금 조정 (차액 반영)
        # 수량이 증가하면 예수금 감소, 수량이 감소하면 예수금 증가
        cash_adjustment = old_amount - new_amount
        self.available_cash += cash_adjustment
        
        # 예수금이 음수가 되면 수정 불가
        if self.available_cash < 0:
            self.available_cash -= cash_adjustment  # 원복
            return False
        
        # 포지션 정보 업데이트
        position["shares"] = new_shares
        position["buy_price"] = new_buy_price
        position["amount"] = new_amount
        
        print(f"✅ {position['round']}회차 포지션 수정 완료 (인덱스: {position_index})")
        print(f"   기존: {old_shares}주 @ ${old_buy_price:.2f} (${old_amount:,.0f})")
        print(f"   수정: {new_shares}주 @ ${new_buy_price:.2f} (${new_amount:,.0f})")
        print(f"   예수금 조정: ${cash_adjustment:+,.0f}")
        
        return True
    
    def get_daily_recommendation(self) -> Dict:
        """
        일일 매매 추천 생성
        Returns:
            Dict: 매매 추천 정보
        """
        print("=" * 60)
        print("🚀 SOXL 퀀트투자 일일 매매 추천")
        print("=" * 60)
        
        # 현재 상태를 최신으로 업데이트 (시작일부터 현재까지 시뮬레이션)
        if self.session_start_date:
            print("🔄 트레이더 상태를 최신으로 업데이트 중...")
            sim_result = self.simulate_from_start_to_today(self.session_start_date, quiet=True)
            if "error" in sim_result:
                return {"error": f"상태 업데이트 실패: {sim_result['error']}"}

        # 시장 휴장일 확인 (테스트 날짜 오버라이드 고려)
        today = self.get_today_date()
        is_market_closed = self.is_market_closed(today)
        
        if is_market_closed:
            latest_trading_day = self.get_latest_trading_day()
            if today.weekday() >= 5:
                print(f"📅 주말입니다. 최신 거래일({latest_trading_day.strftime('%Y-%m-%d')}) 데이터를 사용합니다.")
            else:
                print(f"📅 휴장일입니다. 최신 거래일({latest_trading_day.strftime('%Y-%m-%d')}) 데이터를 사용합니다.")
        
        # 1. SOXL 데이터 가져오기
        soxl_data = self.get_stock_data("SOXL", "1mo")
        if soxl_data is None:
            return {"error": "SOXL 데이터를 가져올 수 없습니다."}
        
        # 오늘 날짜의 미처리 시드증액 적용 (장중엔 시뮬레이션이 어제까지만 실행되므로 오늘 시드가 누락됨)
        today_str = today.strftime('%Y-%m-%d')
        unprocessed_today_seeds = [si for si in self.seed_increases
            if si["date"] == today_str and si["date"] not in self.processed_seed_dates]
        if unprocessed_today_seeds:
            total_seed = sum(si["amount"] for si in unprocessed_today_seeds)
            current_price = soxl_data.iloc[-1]['Close'] if len(soxl_data) > 0 else 0
            total_shares = sum(pos["shares"] for pos in self.positions)
            current_total_assets = self.available_cash + (total_shares * current_price)
            self.available_cash += total_seed
            self.current_investment_capital = current_total_assets + total_seed
            for si in unprocessed_today_seeds:
                self.processed_seed_dates.add(si["date"])
            seed_dates_str = ", ".join(si["date"] for si in unprocessed_today_seeds)
            print(f"💰 오늘 시드증액 반영: {seed_dates_str} - ${total_seed:,.0f} 추가 (매수추천에 반영)")
        
        # 장중에는 오늘 날짜 데이터 제외 (종가가 확정되지 않았으므로)
        today_date = today.date()
        if not self.is_regular_session_closed_now() and len(soxl_data) > 0:
            if soxl_data.index.max().date() == today_date:
                soxl_data = soxl_data[soxl_data.index.date < today_date]
        
        # 2. QQQ 데이터 가져오기 (주간 RSI 계산용)
        qqq_data = self.get_stock_data("QQQ", "6mo")  # 충분한 데이터 확보
        if qqq_data is None:
            return {"error": "QQQ 데이터를 가져올 수 없습니다."}

        # 3. 기존 포지션의 모드 백업 및 재검증 (매수일 기준으로 재계산)
        # RSI 참조 데이터 로드 (수동 입력값 포함)
        rsi_ref_data = {}
        try:
            rsi_file_path = str(self._resolve_data_path("weekly_rsi_reference.json"))
            if os.path.exists(rsi_file_path):
                with open(rsi_file_path, 'r', encoding='utf-8') as f:
                    rsi_ref_data = json.load(f)
        except Exception as e:
            print(f"⚠️ RSI 참조 데이터 로드 실패: {e}")

        # 시뮬레이션 후 포지션 모드를 백업하여 이후 모드 재계산 시 보존
        position_mode_backup = {}
        for pos in self.positions:
            buy_date = pos.get('buy_date')
            if isinstance(buy_date, pd.Timestamp):
                buy_date_dt = buy_date.to_pydatetime()
            elif isinstance(buy_date, datetime):
                buy_date_dt = buy_date
            else:
                continue
            
            # 포지션 키 생성 (회차_매수일)
            pos_key = f"{pos['round']}_{buy_date_dt.strftime('%Y-%m-%d')}"
            stored_mode = pos.get('mode')
            if stored_mode:
                position_mode_backup[pos_key] = stored_mode
                print(f"🔍 포지션 모드 백업: {pos_key} = {stored_mode}")
        
        # 3-0. 포지션 모드 재검증 및 수정 (매수일 기준으로 재계산, 수량/금액도 재계산)
        # QQQ 데이터로 주간 RSI 계산
        weekly_df_for_positions = qqq_data.resample('W-FRI').agg({
            'Open': 'first',
            'High': 'max',
            'Low': 'min',
            'Close': 'last',
            'Volume': 'sum'
        }).dropna()
        
        if len(weekly_df_for_positions) >= 15:
            # RSI 계산
            delta = weekly_df_for_positions['Close'].diff()
            gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
            loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
            rs = gain / loss
            rsi_for_positions = 100 - (100 / (1 + rs))
            
            for pos in self.positions:
                buy_date = pos.get('buy_date')
                if isinstance(buy_date, pd.Timestamp):
                    buy_date_dt = buy_date.to_pydatetime()
                elif isinstance(buy_date, datetime):
                    buy_date_dt = buy_date
                else:
                    continue
                
                # 포지션 키 생성 (회차_매수일)
                pos_key = f"{pos['round']}_{buy_date_dt.strftime('%Y-%m-%d')}"
                current_stored_mode = pos.get('mode')
                
                # 매수일이 속한 주의 금요일 계산
                buy_date_weekday = buy_date_dt.weekday()
                days_until_friday = (4 - buy_date_weekday) % 7
                if days_until_friday == 0 and buy_date_weekday != 4:
                    days_until_friday = 7
                buy_week_friday = buy_date_dt + timedelta(days=days_until_friday)
                
                # 매수일의 1주전, 2주전 금요일 계산
                one_week_ago_friday = buy_week_friday - timedelta(days=7)
                two_weeks_ago_friday = buy_week_friday - timedelta(days=14)
                
                # RSI 값 추출 (참조 데이터 우선 사용)
                one_week_ago_rsi = self.get_rsi_from_reference(one_week_ago_friday, rsi_ref_data)
                two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
                
                # 참조 데이터에 없으면 실시간 계산 데이터 사용
                if one_week_ago_rsi is None:
                    one_week_ago_friday_dt = pd.Timestamp(one_week_ago_friday.date())
                    earlier_dates_1w = weekly_df_for_positions.index[weekly_df_for_positions.index <= one_week_ago_friday_dt]
                    if len(earlier_dates_1w) > 0:
                        one_week_rsi_date = earlier_dates_1w[-1]
                        one_week_rsi_idx = weekly_df_for_positions.index.get_loc(one_week_rsi_date)
                        if one_week_rsi_idx < len(rsi_for_positions):
                            one_week_ago_rsi = rsi_for_positions.iloc[one_week_rsi_idx]

                if two_weeks_ago_rsi is None:
                    two_weeks_ago_friday_dt = pd.Timestamp(two_weeks_ago_friday.date())
                    earlier_dates_2w = weekly_df_for_positions.index[weekly_df_for_positions.index <= two_weeks_ago_friday_dt]
                    if len(earlier_dates_2w) > 0:
                        two_weeks_rsi_date = earlier_dates_2w[-1]
                        two_weeks_rsi_idx = weekly_df_for_positions.index.get_loc(two_weeks_rsi_date)
                        if two_weeks_rsi_idx < len(rsi_for_positions):
                            two_weeks_ago_rsi = rsi_for_positions.iloc[two_weeks_rsi_idx]
                
                # RSI 값으로 매수일의 모드 재계산
                if one_week_ago_rsi is not None and two_weeks_ago_rsi is not None:
                    # 전주 모드를 재귀적으로 계산 (참조 데이터 사용 버전 우선)
                    prev_week_mode, success = self._calculate_week_mode_recursive_with_reference(one_week_ago_friday, rsi_ref_data)
                    
                    if not success:
                        # 참조 데이터 실패 시 실시간 계산 버전으로 시도
                        prev_week_mode, success = self._calculate_week_mode_recursive(one_week_ago_friday, weekly_df_for_positions, rsi_for_positions)
                    
                    if success:
                        # 매수일의 모드 결정
                        is_matched, matched_mode = self._is_mode_case_matched(one_week_ago_rsi, two_weeks_ago_rsi)
                        if is_matched:
                            correct_mode = matched_mode
                        else:
                            correct_mode = prev_week_mode
                        
                        # 저장된 모드와 비교
                        if current_stored_mode != correct_mode:
                            print(f"⚠️ 포지션 모드 불일치 감지: {pos_key}")
                            print(f"   매수일: {buy_date_dt.strftime('%Y-%m-%d')}, 저장된 모드: {current_stored_mode}, 올바른 모드: {correct_mode}")
                            print(f"   RSI 값: 1주전={one_week_ago_rsi:.2f}, 2주전={two_weeks_ago_rsi:.2f}")
                            print(f"🔧 포지션 모드 수정: {pos_key} = {current_stored_mode} → {correct_mode}")
                            
                            # 모드 변경 시 수량과 금액도 재계산
                            old_mode = current_stored_mode
                            old_config = self.sf_config if old_mode == "SF" else self.ag_config
                            new_config = self.sf_config if correct_mode == "SF" else self.ag_config
                            
                            # 매수일 시점의 투자원금 추정 (기존 금액과 split_ratios로 역산)
                            old_round = pos['round']
                            if old_round <= len(old_config["split_ratios"]):
                                old_ratio = old_config["split_ratios"][old_round - 1]
                                if old_ratio > 0:
                                    estimated_investment_capital = pos['amount'] / old_ratio
                                    
                                    # 새로운 모드의 split_ratios로 재계산
                                    if old_round <= len(new_config["split_ratios"]):
                                        new_ratio = new_config["split_ratios"][old_round - 1]
                                        new_target_amount = estimated_investment_capital * new_ratio
                                        
                                        # 수량과 금액 재계산
                                        buy_price = pos['buy_price']
                                        new_shares = int(new_target_amount / buy_price)
                                        if new_shares > 0:
                                            new_amount = new_shares * buy_price
                                            
                                            print(f"   수량/금액 재계산: {pos['shares']}주 @ ${pos['amount']:,.0f} → {new_shares}주 @ ${new_amount:,.0f}")
                                            
                                            # 예수금 차이 계산 (기존 금액에서 새 금액 차감)
                                            amount_diff = pos['amount'] - new_amount
                                            self.available_cash += amount_diff
                                            
                                            # 포지션 업데이트
                                            pos['mode'] = correct_mode
                                            pos['shares'] = new_shares
                                            pos['amount'] = new_amount
                                            
                                            print(f"   예수금 조정: ${amount_diff:+,.0f} (총 예수금: ${self.available_cash:,.0f})")
                                        else:
                                            # 수량이 0이면 모드만 변경
                                            pos['mode'] = correct_mode
                                            print(f"   ⚠️ 수량이 0이므로 모드만 변경")
                                    else:
                                        # 새로운 모드의 회차가 범위를 벗어나면 모드만 변경
                                        pos['mode'] = correct_mode
                                        print(f"   ⚠️ 새로운 모드의 회차 범위를 벗어나므로 모드만 변경")
                                else:
                                    # 비율이 0이면 모드만 변경
                                    pos['mode'] = correct_mode
                                    print(f"   ⚠️ 기존 비율이 0이므로 모드만 변경")
                            else:
                                # 기존 회차가 범위를 벗어나면 모드만 변경
                                pos['mode'] = correct_mode
                                print(f"   ⚠️ 기존 회차가 범위를 벗어나므로 모드만 변경")
        
        # 3-1. 12/29일 매수 포지션 보정 (안전모드로 강제 변경 및 매수 금액 재계산)
        target_date = datetime(2025, 12, 29)
        for pos in self.positions:
            buy_date = pos.get('buy_date')
            if isinstance(buy_date, pd.Timestamp):
                buy_date_dt = buy_date.to_pydatetime()
            elif isinstance(buy_date, datetime):
                buy_date_dt = buy_date
            else:
                continue
            
            # 12/29일 매수 포지션인지 확인
            if buy_date_dt.date() == target_date.date():
                # 모드를 안전모드로 강제 변경
                pos['mode'] = 'SF'
                
                # 12/29일 이전 보유중인 안전모드 포지션 확인
                prev_positions = []
                for p in self.positions:
                    p_buy_date = p.get('buy_date')
                    if not p_buy_date:
                        continue
                    if isinstance(p_buy_date, pd.Timestamp):
                        p_buy_date_dt = p_buy_date.to_pydatetime()
                    elif isinstance(p_buy_date, datetime):
                        p_buy_date_dt = p_buy_date
                    else:
                        continue
                    
                    if p_buy_date_dt.date() < target_date.date() and p.get('mode') == 'SF':
                        prev_positions.append(p)
                
                # 12/29일 포지션의 회차 결정 (이전 안전모드 포지션 수 + 1)
                new_round = len(prev_positions) + 1
                
                # 안전모드의 회차별 시드금액 계산
                if new_round <= len(self.sf_config["split_ratios"]):
                    ratio = self.sf_config["split_ratios"][new_round - 1]
                    
                    # 12/29일 시점의 투자원금 추정
                    # 이전 안전모드 포지션이 있으면 그 포지션의 매수 금액과 split_ratios를 사용하여 역산
                    investment_capital = self.current_investment_capital  # 기본값
                    if prev_positions:
                        # 첫 번째 안전모드 포지션의 매수 금액과 split_ratios로 투자원금 역산
                        first_sf_pos = prev_positions[0]
                        first_round = first_sf_pos.get('round', 1)
                        if first_round <= len(self.sf_config["split_ratios"]):
                            first_ratio = self.sf_config["split_ratios"][first_round - 1]
                            first_amount = first_sf_pos.get('amount', 0)
                            if first_ratio > 0:
                                investment_capital = first_amount / first_ratio
                    
                    target_amount = investment_capital * ratio
                    
                    # 매수 가격은 그대로 유지하고, 수량과 금액만 재계산
                    buy_price = pos['buy_price']
                    new_shares = int(target_amount / buy_price)
                    if new_shares > 0:
                        new_amount = new_shares * buy_price
                        pos['shares'] = new_shares
                        pos['amount'] = new_amount
                        pos['round'] = new_round
                
                break  # 12/29일 포지션은 하나만 있을 것으로 예상
        
        # 4. 과거 종가 기반 포지션 보정 (LOC 매도)
        self.reconcile_positions_with_close_history(soxl_data)

        # 5. QQQ 주간 RSI 기반 모드 자동 전환
        # get_daily_recommendation()에서는 항상 오늘 날짜 기준으로 실시간 QQQ 데이터를 사용하여 모드를 계산함
        # 투자시작일과 무관하게 오늘 날짜 기준으로 모드를 판정해야 함
        # simulate_from_start_to_today()가 시작일 기준으로 모드를 설정했을 수 있으므로,
        # 항상 오늘 날짜 기준으로 모드를 재계산해야 함
        
        today = self.get_today_date()
        days_until_friday = (4 - today.weekday()) % 7
        if days_until_friday == 0 and today.weekday() != 4:
            days_until_friday = 7
        this_week_friday_calc = today + timedelta(days=days_until_friday)
        
        # simulate_from_start_to_today()에서 설정된 현재 주 금요일 확인
        old_week_friday_raw = self.current_week_friday
        
        # 투자시작일과 무관하게 항상 오늘 날짜 기준으로 모드 계산
        # 시작일이 이번 주 내에 있으면, simulate_from_start_to_today()가 시작일 기준으로 모드를 설정했을 수 있으므로
        # 강제로 오늘 날짜 기준 모드를 재계산해야 함
        force_recalculate = False
        if self.session_start_date:
            try:
                session_start_dt = datetime.strptime(self.session_start_date, "%Y-%m-%d")
                start_days_until_friday = (4 - session_start_dt.weekday()) % 7
                if start_days_until_friday == 0 and session_start_dt.weekday() != 4:
                    start_days_until_friday = 7
                start_week_friday = session_start_dt + timedelta(days=start_days_until_friday)
                
                # 시작일이 이번 주 내에 있으면 강제로 재계산
                # (시작일의 모드가 아닌 오늘 날짜 기준 모드를 사용해야 함)
                if start_week_friday.date() == this_week_friday_calc.date():
                    force_recalculate = True
                    print(f"🔄 시작일({self.session_start_date})이 이번 주 내에 있음. 오늘 날짜 기준 모드 강제 재계산")
            except Exception as e:
                print(f"⚠️ 시작일 확인 중 오류: {e}")
        
        # 같은 주 내에서는 모드를 재계산하지 않음 (월요일에 정해진 모드는 그 주 내내 유지)
        # 월~금은 모드가 변경되지 말아야 함 (금요일 장 종료 전에도 같은 주 모드 유지)
        # 날짜만 비교하여 시간 차이 무시
        old_week_friday_date = None
        if old_week_friday_raw is not None:
            if isinstance(old_week_friday_raw, datetime):
                old_week_friday_date = old_week_friday_raw.date()
            else:
                old_week_friday_date = old_week_friday_raw
        
        this_week_friday_date = this_week_friday_calc.date()
        
        # 월요일인지 확인 (월요일 = 0)
        is_monday = today.weekday() == 0
        
        # 같은 주 내이고 모드가 이미 설정되어 있으면 모드 유지
        # 월~금은 모드가 변경되지 말아야 함 (금요일 장 종료 전에도 같은 주 모드 유지)
        is_same_week = (old_week_friday_date is not None and old_week_friday_date == this_week_friday_date)
        
        if is_same_week and self.current_mode is not None and not is_monday:
            # 같은 주 내에서는 모드를 재계산하지 않음 (월요일에 정해진 모드는 그 주 내내 유지)
            # 단, 월요일이 아닌 경우에만 유지
            # 하지만 모드가 잘못 설정되었을 수 있으므로 검증 필요
            # RSI 값을 확인하여 모드가 올바른지 검증
            try:
                # 주간 데이터로 변환하여 RSI 계산
                weekly_df_temp = qqq_data.resample('W-FRI').agg({
                    'Open': 'first',
                    'High': 'max',
                    'Low': 'min',
                    'Close': 'last',
                    'Volume': 'sum'
                }).dropna()
                
                if len(weekly_df_temp) >= 15:
                    # RSI 계산
                    delta = weekly_df_temp['Close'].diff()
                    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
                    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
                    rs = gain / loss
                    rsi = 100 - (100 / (1 + rs))
                    
                    # 1주전, 2주전 금요일 계산
                    days_until_friday = (4 - today.weekday()) % 7
                    if days_until_friday == 0 and today.weekday() != 4:
                        days_until_friday = 7
                    this_week_friday_temp = today + timedelta(days=days_until_friday)
                    # 월~금은 모드가 변경되지 말아야 함
                    # 같은 주 내에서는 모드를 유지하므로, 검증을 위해 지난주 금요일 기준으로 RSI를 확인
                    # 금요일이어도 같은 주 모드를 유지해야 하므로, 항상 지난주 금요일 기준으로 검증
                    latest_completed_friday = this_week_friday_temp - timedelta(days=7)
                    
                    one_week_ago_friday = latest_completed_friday
                    two_weeks_ago_friday = latest_completed_friday - timedelta(days=7)
                    
                    # RSI 값 추출 (참조 데이터 우선 사용)
                    one_week_ago_rsi = self.get_rsi_from_reference(one_week_ago_friday, rsi_ref_data)
                    two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
                    
                    # 참조 데이터에 없으면 실시간 계산 데이터 사용
                    if one_week_ago_rsi is None:
                        one_week_ago_friday_dt = pd.Timestamp(one_week_ago_friday.date())
                        earlier_dates_1w = weekly_df_temp.index[weekly_df_temp.index <= one_week_ago_friday_dt]
                        if len(earlier_dates_1w) > 0:
                            one_week_rsi_date = earlier_dates_1w[-1]
                            one_week_rsi_idx = weekly_df_temp.index.get_loc(one_week_rsi_date)
                            if one_week_rsi_idx < len(rsi):
                                one_week_ago_rsi = rsi.iloc[one_week_rsi_idx]

                    if two_weeks_ago_rsi is None:
                        two_weeks_ago_friday_dt = pd.Timestamp(two_weeks_ago_friday.date())
                        earlier_dates_2w = weekly_df_temp.index[weekly_df_temp.index <= two_weeks_ago_friday_dt]
                        if len(earlier_dates_2w) > 0:
                            two_weeks_rsi_date = earlier_dates_2w[-1]
                            two_weeks_rsi_idx = weekly_df_temp.index.get_loc(two_weeks_rsi_date)
                            if two_weeks_rsi_idx < len(rsi):
                                two_weeks_ago_rsi = rsi.iloc[two_weeks_rsi_idx]
                    
                    # RSI 값으로 모드 검증
                    if one_week_ago_rsi is not None and two_weeks_ago_rsi is not None:
                        is_matched, expected_mode = self._is_mode_case_matched(one_week_ago_rsi, two_weeks_ago_rsi)
                        if is_matched:
                            # 조건에 해당하는 모드가 있으면 그 모드를 사용
                            if expected_mode != self.current_mode:
                                print(f"⚠️ 모드 불일치 감지: 현재 모드={self.current_mode}, 예상 모드={expected_mode}")
                                print(f"   RSI 값: 1주전={one_week_ago_rsi:.2f}, 2주전={two_weeks_ago_rsi:.2f}")
                                print(f"🔄 모드 재계산 필요 (잘못된 모드 감지)")
                                # 모드 재계산
                                temp_current_mode = self.current_mode
                                self.current_week_friday = None
                                self.current_mode = None
                                new_mode = self.update_mode(qqq_data)
                                if new_mode is None:
                                    return {"error": "모드 판정 실패: 전주 모드를 계산할 수 없어 현재 주차의 모드를 결정할 수 없습니다."}
                                print(f"✅ 모드 재계산 완료: {this_week_friday_date} 주차 모드 = {new_mode}")
                            else:
                                print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode} (검증 완료)")
                                new_mode = self.current_mode
                        else:
                            # 조건에 해당하지 않으면 전주 모드 사용해야 함
                            # 전주 모드를 계산하여 검증 (참조 데이터 우선)
                            prev_week_mode, success = self._calculate_week_mode_recursive_with_reference(one_week_ago_friday, rsi_ref_data)
                            if not success:
                                # 참조 데이터 실패 시 실시간 계산 버전으로 시도
                                prev_week_mode, success = self._calculate_week_mode_recursive(one_week_ago_friday, weekly_df_temp, rsi)
                            
                            if success and prev_week_mode != self.current_mode:
                                print(f"⚠️ 모드 불일치 감지: 현재 모드={self.current_mode}, 전주 모드={prev_week_mode}")
                                print(f"   RSI 값: 1주전={one_week_ago_rsi:.2f}, 2주전={two_weeks_ago_rsi:.2f}")
                                print(f"🔄 모드 재계산 필요 (전주 모드와 불일치)")
                                # 모드 재계산
                                temp_current_mode = self.current_mode
                                self.current_week_friday = None
                                self.current_mode = None
                                new_mode = self.update_mode(qqq_data)
                                if new_mode is None:
                                    return {"error": "모드 판정 실패: 전주 모드를 계산할 수 없어 현재 주차의 모드를 결정할 수 없습니다."}
                                print(f"✅ 모드 재계산 완료: {this_week_friday_date} 주차 모드 = {new_mode}")
                            else:
                                print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode} (검증 완료)")
                                new_mode = self.current_mode
                    else:
                        # RSI 값을 가져올 수 없으면 모드 유지
                        print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode} (RSI 검증 불가)")
                        new_mode = self.current_mode
                else:
                    # 주간 데이터가 부족하면 모드 유지
                    print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode} (데이터 부족)")
                    new_mode = self.current_mode
            except Exception as e:
                # 검증 중 오류 발생 시 모드 유지
                print(f"⚠️ 모드 검증 중 오류 발생: {e}, 모드 유지")
                print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode}")
                new_mode = self.current_mode
        elif is_monday:
            # 월요일인 경우 항상 모드를 재계산 (이번 주 모드를 올바르게 설정)
            print(f"🔄 월요일 모드 재계산: {this_week_friday_date} 주차 (월요일이므로 항상 재계산)")
            # 같은 주 체크를 우회하기 위해 current_week_friday를 임시로 None으로 설정
            temp_current_mode = self.current_mode
            self.current_week_friday = None
            self.current_mode = None  # 전주 모드를 올바르게 계산하도록 None으로 설정
            new_mode = self.update_mode(qqq_data)
            if new_mode is None:
                return {"error": "모드 판정 실패: 전주 모드를 계산할 수 없어 현재 주차의 모드를 결정할 수 없습니다."}
            print(f"✅ 월요일 모드 재계산 완료: {this_week_friday_date} 주차 모드 = {new_mode}")
        elif force_recalculate:
            # 강제 재계산이 필요한 경우 (시작일이 이번 주 내에 있는 경우)
            # 같은 주 체크를 우회하기 위해 current_week_friday를 임시로 None으로 설정
            # 이렇게 하면 update_mode()가 같은 주 체크를 건너뛰고 모드를 재계산함
            # 또한 시작일 기준으로 설정된 모드가 전주 모드로 사용되지 않도록,
            # current_mode를 임시로 저장하고 None으로 설정하여 update_mode()가 전주 모드를 올바르게 계산하도록 함
            temp_current_mode = self.current_mode
            self.current_week_friday = None
            self.current_mode = None  # 전주 모드를 올바르게 계산하도록 None으로 설정
            print(f"🔄 모드 재계산 필요 (오늘 날짜 기준, 실시간 QQQ 데이터 사용)")
            new_mode = self.update_mode(qqq_data)
            # update_mode()가 모드 판정 실패 시 None 반환
            if new_mode is None:
                return {"error": "모드 판정 실패: 전주 모드를 계산할 수 없어 현재 주차의 모드를 결정할 수 없습니다."}
            print(f"✅ 오늘 날짜 기준 모드 재계산 완료: {this_week_friday_calc.strftime('%Y-%m-%d')} 주차 모드 = {new_mode}")
        elif old_week_friday_date is None or old_week_friday_date != this_week_friday_date:
            # 새로운 주차인 경우 모드 재계산
            # 단, 금요일이면서 장이 종료되지 않았으면 모드를 업데이트하지 않음
            # 주간 RSI 및 모드 업데이트는 금요일 장 종료 후에만 계산되어 업데이트되어야 함
            if today.weekday() == 4 and not self.is_regular_session_closed_now():
                # 금요일 장 종료 전에는 모드를 업데이트하지 않음
                if self.current_mode:
                    print(f"⏳ 금요일 장 종료 전: {today.strftime('%Y-%m-%d')} 장 종료 전이므로 모드 업데이트하지 않음 (현재 모드: {self.current_mode})")
                    new_mode = self.current_mode
                else:
                    print(f"⏳ 금요일 장 종료 전: {today.strftime('%Y-%m-%d')} 장 종료 전이므로 모드 업데이트하지 않음")
                    return {"error": "금요일 장 종료 전에는 모드를 결정할 수 없습니다."}
            else:
                # 금요일 장 종료 후이거나 금요일이 아닌 경우 모드 재계산
                print(f"🔄 새로운 주차 모드 계산: {this_week_friday_date} 주차")
                new_mode = self.update_mode(qqq_data)
                if new_mode is None:
                    return {"error": "모드 판정 실패: 전주 모드를 계산할 수 없어 현재 주차의 모드를 결정할 수 없습니다."}
        else:
            # 같은 주 내이고 시작일이 다른 주에 있으면 모드 유지
            print(f"✅ 같은 주 내 모드 유지: {this_week_friday_date} 주차 모드 = {self.current_mode}")
            new_mode = self.current_mode
        
        today = self.get_today_date()
        
        # 모드 판단에 사용되는 RSI 계산 (1주전과 2주전)
        # RSI 참조 데이터를 우선 사용하고, 없으면 실시간 계산
        # RSI 참조 데이터 로드
        rsi_ref_data = {}
        try:
            rsi_file_path = str(self._resolve_data_path("weekly_rsi_reference.json"))
            if os.path.exists(rsi_file_path):
                with open(rsi_file_path, 'r', encoding='utf-8') as f:
                    rsi_ref_data = json.load(f)
        except Exception as e:
            print(f"⚠️ RSI 참조 데이터 로드 실패: {e}")
        
        # 오늘 날짜 기준으로 가장 최근 완료된 주차(지난주 금요일) 찾기
        today_date = today.date()
        days_until_friday = (4 - today.weekday()) % 7  # 금요일(4)까지의 일수
        if days_until_friday == 0 and today.weekday() != 4:  # 금요일이 아닌데 계산이 0이면 다음 주 금요일
            days_until_friday = 7
        
        # 이번 주 금요일 계산
        this_week_friday = today + timedelta(days=days_until_friday)
        # 월~금은 모드가 변경되지 말아야 함
        # 같은 주 내에서는 모드를 유지하므로, 표시를 위해 지난주 금요일 기준으로 RSI를 확인
        # 금요일이어도 같은 주 모드를 유지해야 하므로, 항상 지난주 금요일 기준으로 표시
        latest_completed_friday = this_week_friday - timedelta(days=7)
        
        # 1주전과 2주전 금요일 계산
        # 1주전 RSI = 지난주 금요일의 RSI (latest_completed_friday)
        # 2주전 RSI = 지지난주 금요일의 RSI (latest_completed_friday - 7일)
        one_week_ago_friday = latest_completed_friday  # 지난주 금요일 (1주전)
        two_weeks_ago_friday = latest_completed_friday - timedelta(days=7)  # 지지난주 금요일 (2주전)
        
        # RSI 값 추출 (참조 데이터 우선 사용)
        one_week_ago_rsi = self.get_rsi_from_reference(one_week_ago_friday, rsi_ref_data)
        two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
        
        # 참조 데이터에 없으면 5년치 데이터로 정확한 RSI 실시간 계산
        if one_week_ago_rsi is None or two_weeks_ago_rsi is None:
            print(f"⚠️ RSI 참조 데이터 부재 → 5년 데이터 기반 실시간 계산 진행")
            target_fridays = []
            if one_week_ago_rsi is None:
                target_fridays.append(one_week_ago_friday)
            if two_weeks_ago_rsi is None:
                target_fridays.append(two_weeks_ago_friday)
            
            realtime_rsi = self.calculate_weekly_rsi_for_dates(target_fridays)
            
            if one_week_ago_rsi is None:
                one_week_key = one_week_ago_friday.strftime('%Y-%m-%d') if isinstance(one_week_ago_friday, datetime) else pd.Timestamp(one_week_ago_friday).strftime('%Y-%m-%d')
                one_week_ago_rsi = realtime_rsi.get(one_week_key)
            
            if two_weeks_ago_rsi is None:
                two_weeks_key = two_weeks_ago_friday.strftime('%Y-%m-%d') if isinstance(two_weeks_ago_friday, datetime) else pd.Timestamp(two_weeks_ago_friday).strftime('%Y-%m-%d')
                two_weeks_ago_rsi = realtime_rsi.get(two_weeks_key)
            
        if one_week_ago_rsi is None:
            return {"error": "QQQ 주간 RSI를 계산할 수 없습니다."}

        # 5. 최신 SOXL 가격 정보 (최소 2일 데이터 필요)
        if len(soxl_data) < 2:
            return {"error": "데이터가 부족합니다. 최소 2일의 데이터가 필요합니다."}
        
        latest_soxl = soxl_data.iloc[-1]
        current_price = latest_soxl['Close']
        latest_data_date = soxl_data.index[-1]

        # 전일 종가 계산
        # - 데이터의 마지막 날짜가 오늘보다 이전(주말/휴장/장전)인 경우: 마지막 종가를 기준으로 매수가 계산
        # - 데이터의 마지막 날짜가 오늘(이미 오늘 종가가 존재)인 경우: 그 전날 종가를 기준으로 계산
        last_data_date = latest_data_date.date()
        today_date = today.date()  # 테스트 날짜 오버라이드 고려
        if last_data_date < today_date:
            # 최신 거래일 종가를 전일 종가로 간주
            prev_close = soxl_data.iloc[-1]['Close']
            prev_close_basis_date = soxl_data.index[-1].strftime("%Y-%m-%d")
            display_date = today_date.strftime("%Y-%m-%d")  # 장중/휴장일에도 화면 날짜는 오늘로 표시
            # 매도 조건 확인은 어제(최신 거래일) 종가 기준으로 확인
            check_sell_date = latest_data_date
            check_sell_row = latest_soxl
        else:
            # 오늘 종가가 포함되어 있으므로 바로 전날 종가 사용
            prev_close = soxl_data.iloc[-2]['Close']
            prev_close_basis_date = soxl_data.index[-2].strftime("%Y-%m-%d")
            display_date = latest_data_date.strftime("%Y-%m-%d")
            # 매도 조건 확인은 전날 종가 기준으로 확인
            check_sell_date = soxl_data.index[-2]
            check_sell_row = soxl_data.iloc[-2]
        
        # 6. 매수/매도 가격 계산

        buy_price, sell_price = self.calculate_buy_sell_prices(prev_close)
        
        # 7. 매도 조건 확인
        # 오늘 기준으로 매도 조건 확인 (어제 종가 기준)
        # current_date는 오늘 날짜로 설정하여 손절예정일 체크가 올바르게 작동하도록 함
        today_datetime = datetime.combine(today_date, datetime.min.time())
        all_sell_recommendations, sell_debug_info = self.check_sell_conditions(check_sell_row, today_datetime, prev_close, return_debug_info=True)
        
        # 매도 추천 리스트에는 매도 조건이 충족되지 않은 포지션만 포함 (사용자 요청)
        # 실제 매도 조건이 충족된 것(will_sell=True)은 리스트에서 제외
        sell_recommendations = [s for s in all_sell_recommendations if not s.get('will_sell', False)]
        
        # 8. 매수 조건 확인
        can_buy = self.can_buy_next_round()
        next_buy_amount = self.calculate_position_size(self.current_round) if can_buy else 0
        
        # 9. 포트폴리오 현황
        total_position_value = sum([pos["shares"] * current_price for pos in self.positions])
        total_invested = sum([pos["amount"] for pos in self.positions])
        unrealized_pnl = total_position_value - total_invested
        
        market_closed = self.is_regular_session_closed_now()
        recommendation = {
            "date": display_date,  # 화면 표시용 날짜 (가능하면 오늘)
            "basis_date": prev_close_basis_date,  # 매수가 계산에 사용된 기준 종가의 날짜
            "data_last_date": latest_data_date.strftime("%Y-%m-%d"),  # 확정종가 마지막 날짜
            "market_closed": market_closed,  # 미국 정규장 마감 여부
            "mode": self.current_mode,
            "qqq_one_week_ago_rsi": one_week_ago_rsi,  # 1주전 RSI (모드 판단에 사용)
            "qqq_two_weeks_ago_rsi": two_weeks_ago_rsi,  # 2주전 RSI (모드 판단에 사용)
            "soxl_current_price": current_price,
            "buy_price": buy_price,
            "sell_price": sell_price,
            "can_buy": can_buy,
            "next_buy_round": self.current_round if can_buy else None,
            "next_buy_amount": next_buy_amount,
            "sell_recommendations": sell_recommendations,
            "sell_debug_info": sell_debug_info,  # 매도 조건 확인 디버깅 정보
            "portfolio": {
                "positions_count": len(self.positions),
                "total_invested": total_invested,
                "total_position_value": total_position_value,
                "unrealized_pnl": unrealized_pnl,
                "available_cash": self.available_cash,
                "total_portfolio_value": self.available_cash + total_position_value
            }
        }
        
        return recommendation
    
    def print_recommendation(self, rec: Dict):
        """매매 추천 출력"""
        if "error" in rec:
            print(f"❌ 오류: {rec['error']}")
            return
        
        print(f"📅 날짜: {rec['date']}")

        mode_name = "안전모드" if rec['mode'] == "SF" else "공세모드"
        print(f"🎯 모드: {rec['mode']} ({mode_name})")
        
        # RSI 정보 출력 (모드 판단에 사용되는 1주전과 2주전 RSI)
        one_week_ago_rsi = rec.get('qqq_one_week_ago_rsi')
        two_weeks_ago_rsi = rec.get('qqq_two_weeks_ago_rsi')
        if one_week_ago_rsi is not None:
            if two_weeks_ago_rsi is not None:
                print(f"📊 QQQ 주간 RSI: 1주전 {one_week_ago_rsi:.2f} | 2주전 {two_weeks_ago_rsi:.2f}")
            else:
                print(f"📊 QQQ 주간 RSI: 1주전 {one_week_ago_rsi:.2f} | 2주전 (데이터 없음)")
        else:
            print(f"📊 QQQ 주간 RSI: (계산 불가)")
        
        print(f"💰 SOXL 현재가: ${rec['soxl_current_price']:.2f}")
        print()
        
        print("📋 오늘의 매매 추천:")
        print("-" * 40)
        
        # 매수 추천
        if rec['can_buy']:
            print(f"🟢 매수 추천: {rec['next_buy_round']}회차")
            print(f"   매수가: ${rec['buy_price']:.2f} (LOC 주문)")
            print(f"   매수금액: ${rec['next_buy_amount']:,.0f}")
            shares = int(rec['next_buy_amount'] / rec['buy_price'])
            print(f"   매수주식수: {shares}주")
        else:
            if self.current_round > self.get_current_config()["split_count"]:
                print("🔴 매수 불가: 모든 분할매수 완료")
            else:
                print("🔴 매수 불가: 시드 부족")
        
        print()
        
        # 매도 추천
        if rec['sell_recommendations']:
            print(f"🔴 매도 추천: {len(rec['sell_recommendations'])}건")
            for sell_info in rec['sell_recommendations']:
                pos = sell_info['position']
                print(f"   {pos['round']}회차 매도: {pos['shares']}주 @ ${sell_info['sell_price']:.2f}")
                print(f"   매도 사유: {sell_info['reason']}")
        else:
            # 보유 포지션이 있으면 매도 목표가 안내
            if self.positions:
                print("📋 보유 포지션 LOC 매도 목표가 안내:")
                for pos in self.positions:
                    config = self.sf_config if pos['mode'] == "SF" else self.ag_config
                    target_sell_price = pos['buy_price'] * (1 + config['sell_threshold'] / 100)
                    current_price = rec['soxl_current_price']
                    price_diff = target_sell_price - current_price
                    price_diff_pct = (price_diff / current_price) * 100
                    
                    # 매수체결일 포맷팅
                    buy_date = pos.get('buy_date')
                    if isinstance(buy_date, pd.Timestamp):
                        buy_date_str = buy_date.strftime('%Y-%m-%d')
                    elif isinstance(buy_date, datetime):
                        buy_date_str = buy_date.strftime('%Y-%m-%d')
                    else:
                        buy_date_str = str(buy_date)
                    
                    # 모드 정보
                    mode = pos.get('mode', 'SF')
                    mode_name = "안전모드(SF)" if mode == "SF" else "공세모드(AG)"
                    
                    # 매도 목표일자 계산 (최대 보유기간)
                    buy_date_dt = pos.get('buy_date')
                    if isinstance(buy_date_dt, pd.Timestamp):
                        buy_date_dt = buy_date_dt.to_pydatetime()
                    elif not isinstance(buy_date_dt, datetime):
                        try:
                            buy_date_dt = datetime.strptime(buy_date_str, '%Y-%m-%d')
                        except:
                            buy_date_dt = datetime.now()
                    
                    target_sell_date = self.calculate_stop_loss_date(buy_date_dt, config['max_hold_days'])
                    
                    print(f"   📦 {pos['round']}회차: 목표가 ${target_sell_price:.2f}")
                    print(f"      매수체결일: {buy_date_str}")
                    print(f"      매수체결가격: ${pos['buy_price']:.2f}")
                    print(f"      모드: {mode_name}")
                    print(f"      매도목표일자: {target_sell_date}")
                    print(f"      보유: {pos['shares']}주")
            else:
                print("🟡 매도 추천 없음")
        
        print()
        print("💼 포트폴리오 현황:")
        print("-" * 40)
        portfolio = rec['portfolio']
        print(f"보유 포지션: {portfolio['positions_count']}개")
        print(f"투자원금: ${portfolio['total_invested']:,.0f}")
        print(f"평가금액: ${portfolio['total_position_value']:,.0f}")
        print(f"평가손익: ${portfolio['unrealized_pnl']:,.0f} ({(portfolio['unrealized_pnl']/portfolio['total_invested']*100) if portfolio['total_invested'] > 0 else 0:+.2f}%)")
        print(f"현금잔고: ${portfolio['available_cash']:,.0f}")
        print(f"총 자산: ${portfolio['total_portfolio_value']:,.0f}")
        
        print()
        print("📊 보유 포지션 상세:")
        print("-" * 40)
        if self.positions:
            for pos in self.positions:
                hold_days = (datetime.now() - pos['buy_date']).days
                current_value = pos['shares'] * rec['soxl_current_price']
                pnl = current_value - pos['amount']
                pnl_rate = (pnl / pos['amount']) * 100
                
                # 매수체결일 포맷팅
                buy_date = pos.get('buy_date')
                if isinstance(buy_date, pd.Timestamp):
                    buy_date_str = buy_date.strftime('%Y-%m-%d')
                elif isinstance(buy_date, datetime):
                    buy_date_str = buy_date.strftime('%Y-%m-%d')
                else:
                    buy_date_str = str(buy_date)
                
                # 모드 정보
                mode = pos.get('mode', 'SF')
                mode_name = "안전모드(SF)" if mode == "SF" else "공세모드(AG)"
                
                # 매도 목표가 계산
                config = self.sf_config if mode == "SF" else self.ag_config
                target_sell_price = pos['buy_price'] * (1 + config['sell_threshold'] / 100)
                
                print(f"{pos['round']}회차: {pos['shares']}주 @ ${pos['buy_price']:.2f} ({hold_days}일 보유)")
                print(f"        매수체결일: {buy_date_str}")
                print(f"        모드: {mode_name}")
                print(f"        매도목표가: ${target_sell_price:.2f}")
                print(f"        평가: ${current_value:,.0f} | 손익: ${pnl:,.0f} ({pnl_rate:+.2f}%)")
        else:
            print("보유 포지션 없음")
    
    def reset_portfolio(self):
        """포트폴리오 초기화 (백테스팅용)"""
        self.positions = []
        self.current_round = 1
        self.available_cash = self.initial_capital

        
        # 투자원금 관리 초기화
        self.current_investment_capital = self.initial_capital
        self.trading_days_count = 0
        
        # 처리된 시드증액 날짜 초기화
        self.processed_seed_dates = set()
        
        # 주차 추적 초기화
        self.current_week_friday = None
    
    def clear_cache(self):
        """캐시 초기화 (설정 변경 시 호출)"""
        self._stock_data_cache.clear()
        self._simulation_cache.clear()
        print("🧹 캐시 초기화 완료")
    
    def check_backtest_starting_state(self, start_date: str, rsi_ref_data: dict) -> dict:
        """
        백테스팅 시작 시점의 상태 확인
        Args:
            start_date: 백테스팅 시작일
            rsi_ref_data: RSI 참조 데이터
        Returns:
            dict: 시작 시점 상태 정보
        """
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            
            # 시작일의 주차와 RSI 확인
            days_until_friday = (4 - start_dt.weekday()) % 7
            if days_until_friday == 0 and start_dt.weekday() != 4:
                days_until_friday = 7
            start_week_friday = start_dt + timedelta(days=days_until_friday)
            
            # 시작 주차의 RSI와 모드 확인
            start_week_rsi = self.get_rsi_from_reference(start_week_friday, rsi_ref_data)
            
            # 1주전, 2주전 RSI 확인
            prev_week_friday = start_week_friday - timedelta(days=7)
            two_weeks_ago_friday = start_week_friday - timedelta(days=14)
            
            prev_week_rsi = self.get_rsi_from_reference(prev_week_friday, rsi_ref_data)
            two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
            
            # RSI 데이터가 없는 경우 실시간 계산 폴백
            missing_fridays = []
            if start_week_rsi is None:
                missing_fridays.append(start_week_friday)
            if prev_week_rsi is None:
                missing_fridays.append(prev_week_friday)
            if two_weeks_ago_rsi is None:
                missing_fridays.append(two_weeks_ago_friday)
            
            if missing_fridays:
                try:
                    fallback_rsi = self.calculate_weekly_rsi_for_dates(missing_fridays)
                    if start_week_rsi is None:
                        start_week_rsi = fallback_rsi.get(start_week_friday.strftime('%Y-%m-%d'))
                    if prev_week_rsi is None:
                        prev_week_rsi = fallback_rsi.get(prev_week_friday.strftime('%Y-%m-%d'))
                    if two_weeks_ago_rsi is None:
                        two_weeks_ago_rsi = fallback_rsi.get(two_weeks_ago_friday.strftime('%Y-%m-%d'))
                except Exception as e:
                    print(f"⚠️ RSI 실시간 계산 폴백 실패: {e}")
            
            # 시작 모드 결정
            # 시작일 이전 주차의 모드를 계산하여 시작 모드 결정
            if prev_week_rsi is not None and two_weeks_ago_rsi is not None:
                # 시작일 이전 주차의 모드를 계산하기 위해 1주전 금요일의 1주전/2주전 RSI 사용
                prev_prev_week_friday = prev_week_friday - timedelta(days=7)
                prev_prev_week_rsi = self.get_rsi_from_reference(prev_prev_week_friday, rsi_ref_data)
                prev_prev_two_weeks_rsi = self.get_rsi_from_reference(prev_prev_week_friday - timedelta(days=7), rsi_ref_data)
                
                # 시작일 이전 주차의 모드를 재귀적으로 계산 (case에 해당하는 모드가 나올 때까지)
                prev_week_mode, success = self._calculate_week_mode_recursive_with_reference(prev_week_friday, rsi_ref_data)
                
                if not success:
                    print(f"❌ 모드판정실패: 백테스팅 시작일 이전 주차 모드 계산 실패")
                    return {
                        "error": f"백테스팅 시작일 이전 주차 모드 계산 실패",
                        "start_mode": None,
                        "start_round": 1,
                        "start_week_rsi": start_week_rsi,
                        "prev_week_rsi": prev_week_rsi,
                        "two_weeks_ago_rsi": two_weeks_ago_rsi
                    }
                
                print(f"🔍 백테스팅 시작 모드 계산:")
                print(f"   시작일: {start_date}")
                print(f"   시작 주차 금요일: {start_week_friday.strftime('%Y-%m-%d')}")
                print(f"   1주전 RSI: {prev_week_rsi:.2f}, 2주전 RSI: {two_weeks_ago_rsi:.2f}")
                print(f"   이전 주차 모드: {prev_week_mode}")
                
                # 시작 모드 결정 (이전 주차의 모드를 사용)
                start_mode = self.determine_mode(prev_week_rsi, two_weeks_ago_rsi, prev_week_mode)
                print(f"   결정된 시작 모드: {start_mode}")
            else:
                print(f"[ERROR] 백테스팅 시작 시점의 RSI 데이터가 없습니다.")
                print(f"   시작 주차 RSI: {start_week_rsi}")
                print(f"   1주전 RSI: {prev_week_rsi}")
                print(f"   2주전 RSI: {two_weeks_ago_rsi}")
                return {
                    "error": f"백테스팅 시작 시점의 RSI 데이터가 없습니다. 1주전: {prev_week_rsi}, 2주전: {two_weeks_ago_rsi}",
                    "start_mode": "SF",
                    "start_round": 1,
                    "start_week_rsi": None,
                    "prev_week_rsi": None,
                    "two_weeks_ago_rsi": None
                }
            
            # 해당 모드에서 몇 회차까지 매수했는지 추정
            # (실제로는 과거 매수 기록이 있어야 정확하지만, 여기서는 간단히 추정)
            estimated_round = 1  # 기본값
            
            print(f"[INFO] 백테스팅 시작 상태:")
            print(f"   - 시작일: {start_date}")
            print(f"   - 시작 주차 RSI: {start_week_rsi:.2f}")
            print(f"   - 1주전 RSI: {prev_week_rsi:.2f}")
            print(f"   - 2주전 RSI: {two_weeks_ago_rsi:.2f}")
            print(f"   - 시작 모드: {start_mode}")
            print(f"   - 시작 회차: {estimated_round}회차")
            
            return {
                "start_mode": start_mode,
                "start_round": estimated_round,
                "start_week_rsi": start_week_rsi,
                "prev_week_rsi": prev_week_rsi,
                "two_weeks_ago_rsi": two_weeks_ago_rsi
            }
            
        except Exception as e:
            print(f"[ERROR] 백테스팅 시작 상태 확인 오류: {e}")
            return {
                "start_mode": "SF",
                "start_round": 1,
                "start_week_rsi": None,
                "prev_week_rsi": None,
                "two_weeks_ago_rsi": None
            }
    
    def run_backtest(self, start_date: str, end_date: str = None) -> Dict:
        """
        백테스팅 실행
        Args:
            start_date: 시작 날짜 (YYYY-MM-DD 형식)
            end_date: 종료 날짜 (None이면 오늘까지)
        Returns:
            Dict: 백테스팅 결과
        """
        print(f"🔄 백테스팅 시작: {start_date} ~ {end_date or '오늘'}")
        
        # 로그 저장용 리스트 초기화
        self.backtest_logs = []

        
        # RSI 참조 데이터 로드
        rsi_ref_data = self.load_rsi_reference_data()
        
        # 포트폴리오 초기화
        self.reset_portfolio()

        
        # 백테스팅 시작 상태 확인
        starting_state = self.check_backtest_starting_state(start_date, rsi_ref_data)
        
        # RSI 데이터가 없는 경우 백테스팅 중단
        if "error" in starting_state:
            return {"error": starting_state["error"]}
        
        # 시작 모드와 회차 설정
        self.current_mode = starting_state["start_mode"]
        self.current_round = starting_state["start_round"]
        
        print(f"🎯 백테스팅 시작 설정:")
        print(f"   - 모드: {self.current_mode}")
        print(f"   - 회차: {self.current_round}")
        print(f"   - 1회시드 예상: ${self.initial_capital * self.get_current_config()['split_ratios'][self.current_round-1]:,.0f}")
        
        # 날짜 파싱 (종료일은 해당 날짜의 23:59:59로 설정하여 당일 데이터 포함)
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            if end_date:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                end_dt = end_dt.replace(hour=23, minute=59, second=59)
            else:
                end_dt = datetime.now()
        except ValueError:
            return {"error": "날짜 형식이 올바르지 않습니다. YYYY-MM-DD 형식으로 입력해주세요."}
        
        # 장 마감 전에는 종료일을 확정된 최신 거래일로 강제 보정
        try:
            if not self.is_regular_session_closed_now():
                latest_trading_day = self.get_latest_trading_day().date()
                effective_end_date = min(end_dt.date(), latest_trading_day)
                end_dt = datetime(effective_end_date.year, effective_end_date.month, effective_end_date.day, 23, 59, 59)
        except Exception:
            pass
        

        # 충분한 기간의 데이터 가져오기
        data_start = start_dt - timedelta(days=180)
        

        # SOXL 데이터 가져오기 (2011년부터 데이터 확보)
        period_days = (datetime.now() - data_start).days
        if period_days <= 365:
            period = "1y"
        elif period_days <= 730:
            period = "2y"

        elif period_days <= 1825:  # 5년
            period = "5y"

        elif period_days <= 3650:  # 10년
            period = "10y"
        else:
            period = "15y"  # 15년 (SOXL은 2010년 출시)
            
        soxl_data = self.get_stock_data("SOXL", period)
        if soxl_data is None:
            return {"error": "SOXL 데이터를 가져올 수 없습니다."}
        
        # QQQ 데이터 가져오기
        qqq_data = self.get_stock_data("QQQ", period)
        if qqq_data is None:
            return {"error": "QQQ 데이터를 가져올 수 없습니다."}
        
        # 정규장 미마감이고, 마지막 인덱스 날짜가 오늘이면 무조건 제외 (공급사 조기 생성 일봉 방지)
        try:
            today_date = datetime.now().date()
            # 오늘이 거래일이고 정규장이 아직 마감되지 않았다면 오늘 데이터 제외
            if self.is_trading_day(datetime.now()) and not self.is_regular_session_closed_now():
                if len(soxl_data) > 0 and soxl_data.index.max().date() == today_date:
                    soxl_data = soxl_data[soxl_data.index.date < today_date]
                if len(qqq_data) > 0 and qqq_data.index.max().date() == today_date:
                    qqq_data = qqq_data[qqq_data.index.date < today_date]
        except Exception:
            pass

        # 종료일이 데이터의 마지막 날짜와 같고, 정규장이 아직 마감되지 않았다면 마지막 행 제외
        # 단, 백테스팅 종료일이 과거 날짜라면 이미 확정된 데이터이므로 포함
        try:
            if end_date:
                end_d = datetime.strptime(end_date, "%Y-%m-%d").date()
                last_date = soxl_data.index.max().date() if len(soxl_data) > 0 else None
                today_date = datetime.now().date()
                
                # 종료일이 오늘이 아니거나, 오늘이더라도 정규장이 마감되었다면 포함
                if last_date and end_d == last_date:
                    if end_d < today_date or (end_d == today_date and self.is_regular_session_closed_now()):
                        # 과거 날짜이거나 오늘 정규장이 마감되었다면 포함
                        pass
                    else:
                        # 오늘이고 정규장이 아직 마감되지 않았다면 제외
                        soxl_data = soxl_data[soxl_data.index.date < last_date]
                        qqq_data = qqq_data[qqq_data.index.date < last_date]
        except Exception:
            pass
        
        # 백테스팅 기간 데이터 필터링 (기존 방식: 타임스탬프 비교)
        soxl_backtest = soxl_data[soxl_data.index >= start_dt]
        soxl_backtest = soxl_backtest[soxl_backtest.index <= end_dt]
        
        if len(soxl_backtest) == 0:
            return {"error": "해당 기간에 대한 데이터가 없습니다."}
        

        # 매매 기록 저장용 (실제 양식에 맞게)
        daily_records = []  # 일별 기록
        current_week_rsi = starting_state["start_week_rsi"]  # 시작 주차 RSI
        current_mode = starting_state["start_mode"]  # 시작 모드
        current_week = 0  # 현재 주차 (첫 번째 주차 처리 후 1이 됨)
        total_realized_pnl = 0  # 누적 실현손익
        total_invested = 0  # 총 투자금
        cash_balance = self.initial_capital  # 현금 잔고
        
        print(f"📊 총 {len(soxl_backtest)}일 백테스팅 진행...")
        

        # 백테스팅 시작일의 전일 종가 설정
        prev_close = None

        if len(soxl_backtest) > 0:
            # 시작일 전날의 종가를 찾기 위해 전체 데이터에서 검색
            start_date_prev = start_dt - timedelta(days=1)
            prev_data = soxl_data[soxl_data.index <= start_date_prev]
            if len(prev_data) > 0:
                prev_close = prev_data.iloc[-1]['Close']
                print(f"📅 백테스팅 시작 전일 종가: {prev_close:.2f} (날짜: {prev_data.index[-1].strftime('%Y-%m-%d')})")
            else:
                print("⚠️ 백테스팅 시작 전일 데이터를 찾을 수 없습니다.")
        
        current_week_friday = None  # 현재 주차의 금요일 (로컬 변수)
        previous_day_sold_rounds = 0  # 전날 매도된 회차 수 추적
        
        # 주차별 모드 저장 (금요일 날짜를 키로 사용)
        week_modes = {}  # {금요일 날짜 문자열: 모드}
        
        # 시작 주차의 모드 저장
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        days_until_friday = (4 - start_dt.weekday()) % 7
        if days_until_friday == 0 and start_dt.weekday() != 4:
            days_until_friday = 7
        start_week_friday = start_dt + timedelta(days=days_until_friday)
        start_week_friday_str = start_week_friday.strftime('%Y-%m-%d')
        week_modes[start_week_friday_str] = current_mode
        print(f"🔍 시작 주차 모드 저장: {start_week_friday_str} = {current_mode}")
        
        for i, (current_date, row) in enumerate(soxl_backtest.iterrows()):
            current_price = row['Close']
            
            # 매도 후 current_round를 보유 중인 회차 수 + 1로 재계산
            if previous_day_sold_rounds > 0:
                # 현재 보유 중인 회차 수 계산
                holding_rounds = len(self.positions)
                # 다음 매수 회차 = 보유 회차 수 + 1
                self.current_round = holding_rounds + 1
                print(f"🔄 전날 매도 완료: {previous_day_sold_rounds}개 회차 매도 → 보유: {holding_rounds}개 → 다음 매수: {self.current_round}회차")
                previous_day_sold_rounds = 0  # 반영 후 초기화
            

            # 거래일 카운터 증가 (거래일인 경우에만)
            if self.is_trading_day(current_date):
                self.trading_days_count += 1
                
                # 시드증액 반영 (해당 날짜 또는 그 이전 날짜에 시드증액이 있는 경우)
                current_date_str = current_date.strftime('%Y-%m-%d')
                current_date_obj = current_date.date()
                
                # 현재 날짜 이하의 모든 미반영 시드증액 찾기
                unprocessed_seeds = []
                for seed in self.seed_increases:
                    seed_date_str = seed["date"]
                    seed_date_obj = datetime.strptime(seed_date_str, "%Y-%m-%d").date()
                    # 시드증액 날짜가 현재 날짜 이하이고, 아직 처리되지 않은 경우
                    if seed_date_obj <= current_date_obj and seed_date_str not in self.processed_seed_dates:
                        unprocessed_seeds.append(seed)
                
                if unprocessed_seeds:
                    # 현재 총자산 계산 (현금 + 보유주식 평가금액)
                    total_shares = sum([pos["shares"] for pos in self.positions])
                    current_total_assets = self.available_cash + (total_shares * current_price)
                    
                    # 시드증액 총합 계산
                    total_seed_increase = sum([si["amount"] for si in unprocessed_seeds])
                    
                    # 시드증액을 현금잔고에 추가
                    self.available_cash += total_seed_increase
                    
                    # 투자원금을 현재 총자산 + 시드증액으로 갱신
                    new_investment_capital = current_total_assets + total_seed_increase
                    old_capital = self.current_investment_capital
                    self.current_investment_capital = new_investment_capital
                    
                    # 처리된 시드증액 날짜 기록
                    for seed in unprocessed_seeds:
                        self.processed_seed_dates.add(seed["date"])
                    
                    seed_dates_str = ", ".join([si["date"] for si in unprocessed_seeds])
                    print(f"💰 시드증액 반영: {current_date_str} (시드증액 날짜: {seed_dates_str}) - ${total_seed_increase:,.0f} 추가")
                    print(f"   현재 총자산: ${current_total_assets:,.0f} + 시드증액: ${total_seed_increase:,.0f} = ${new_investment_capital:,.0f}")
                    print(f"   투자원금 갱신: ${old_capital:,.0f} → ${new_investment_capital:,.0f}")
                
                # 10거래일마다 투자원금 업데이트는 매매 처리 후로 이동 (아래 참조)
            
            # 현재 날짜가 속하는 주차의 금요일 계산
            days_until_friday = (4 - current_date.weekday()) % 7  # 금요일(4)까지의 일수
            if days_until_friday == 0 and current_date.weekday() != 4:  # 금요일이 아닌데 계산이 0이면 다음 주 금요일
                days_until_friday = 7
            this_week_friday = current_date + timedelta(days=days_until_friday)
            
            # 새로운 주차인지 확인 (금요일이 바뀌었는지 또는 첫 번째 날짜인 경우)
            if current_week_friday is None or current_week_friday != this_week_friday:
                current_week_friday = this_week_friday
                # self.current_week_friday도 업데이트 (get_daily_recommendation에서 사용)
                self.current_week_friday = this_week_friday
                
                # 새로운 주차의 RSI 값 가져오기 (해당 주차의 금요일 기준)
                current_week_rsi = self.get_rsi_from_reference(this_week_friday, rsi_ref_data)
                
                # 모드 업데이트 (2주전 RSI와 1주전 RSI 비교)
                # 2주전과 1주전 RSI 계산
                prev_week_friday = this_week_friday - timedelta(days=7)  # 1주전
                two_weeks_ago_friday = this_week_friday - timedelta(days=14)  # 2주전
                
                prev_week_rsi = self.get_rsi_from_reference(prev_week_friday, rsi_ref_data)  # 1주전 RSI
                two_weeks_ago_rsi = self.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)  # 2주전 RSI
                
                # RSI 데이터가 없는 경우 실시간 계산 폴백
                missing_fridays = []
                if current_week_rsi is None:
                    missing_fridays.append(this_week_friday)
                if prev_week_rsi is None:
                    missing_fridays.append(prev_week_friday)
                if two_weeks_ago_rsi is None:
                    missing_fridays.append(two_weeks_ago_friday)
                
                if missing_fridays:
                    try:
                        fallback_rsi = self.calculate_weekly_rsi_for_dates(missing_fridays)
                        if current_week_rsi is None:
                            current_week_rsi = fallback_rsi.get(this_week_friday.strftime('%Y-%m-%d'))
                        if prev_week_rsi is None:
                            prev_week_rsi = fallback_rsi.get(prev_week_friday.strftime('%Y-%m-%d'))
                        if two_weeks_ago_rsi is None:
                            two_weeks_ago_rsi = fallback_rsi.get(two_weeks_ago_friday.strftime('%Y-%m-%d'))
                    except Exception as e:
                        print(f"⚠️ RSI 실시간 계산 폴백 실패: {e}")
                
                if current_week_rsi is None:
                    return {"error": f"RSI 데이터가 없습니다. 주차: {this_week_friday.strftime('%Y-%m-%d')}"}
                if prev_week_rsi is None or two_weeks_ago_rsi is None:
                    return {"error": f"RSI 데이터가 없습니다. 1주전 RSI: {prev_week_rsi}, 2주전 RSI: {two_weeks_ago_rsi}"}
                
                # 모드 결정 (2주전 vs 1주전 비교)
                # 중요: 이전 주차의 모드를 사용하여 현재 주차의 모드를 결정
                # 하지만 RSI가 같으면 이전 모드를 유지하므로, 이전 주차의 모드가 중요함
                # 이전 주차의 모드를 정확히 계산하기 위해 이전 주차의 1주전/2주전 RSI를 사용
                prev_week_prev_rsi = self.get_rsi_from_reference(prev_week_friday - timedelta(days=7), rsi_ref_data)  # 이전 주차의 1주전 RSI
                prev_week_two_weeks_rsi = self.get_rsi_from_reference(prev_week_friday - timedelta(days=14), rsi_ref_data)  # 이전 주차의 2주전 RSI
                
                # 이전 주차의 모드를 정확히 계산하기 위해 순차적으로 이전 주차들의 모드를 계산
                # 전전주, 전전전주의 모드를 확인하여 전 주의 모드를 정확히 결정
                prev_week_friday_str = prev_week_friday.strftime('%Y-%m-%d')
                
                # 이전 주차의 모드가 이미 계산되어 있으면 사용 (우선순위 1)
                if prev_week_friday_str in week_modes:
                    actual_prev_week_mode = week_modes[prev_week_friday_str]
                    print(f"🔍 이전 주차 모드 (저장된 값 사용): {prev_week_friday_str} = {actual_prev_week_mode}")
                    # current_mode도 동기화
                    if current_mode != actual_prev_week_mode:
                        print(f"⚠️ current_mode 동기화: {current_mode} → {actual_prev_week_mode}")
                        current_mode = actual_prev_week_mode
                elif prev_week_prev_rsi is not None and prev_week_two_weeks_rsi is not None:
                    # 이전 주차의 모드를 계산
                    actual_prev_week_mode = current_mode  # 기본값은 현재 모드
                    
                    # 이전 주차의 모드를 계산하기 위해 전전주, 전전전주의 모드를 확인
                    prev_prev_week_friday = prev_week_friday - timedelta(days=7)
                    prev_prev_prev_week_friday = prev_week_friday - timedelta(days=14)
                    
                    prev_prev_week_friday_str = prev_prev_week_friday.strftime('%Y-%m-%d')
                    prev_prev_prev_week_friday_str = prev_prev_prev_week_friday.strftime('%Y-%m-%d')
                    
                    # 전전주의 모드 확인
                    prev_prev_week_mode = "SF"  # 기본값
                    if prev_prev_week_friday_str in week_modes:
                        prev_prev_week_mode = week_modes[prev_prev_week_friday_str]
                        print(f"🔍 전전주 모드 (저장된 값 사용): {prev_prev_week_friday_str} = {prev_prev_week_mode}")
                    else:
                        # 전전주의 모드를 계산하기 위해 전전전주의 모드 확인
                        prev_prev_prev_week_mode = "SF"  # 기본값
                        if prev_prev_prev_week_friday_str in week_modes:
                            prev_prev_prev_week_mode = week_modes[prev_prev_prev_week_friday_str]
                            print(f"🔍 전전전주 모드 (저장된 값 사용): {prev_prev_prev_week_friday_str} = {prev_prev_prev_week_mode}")
                        
                        # 전전주의 RSI 확인
                        prev_prev_prev_rsi = self.get_rsi_from_reference(prev_prev_prev_week_friday, rsi_ref_data)
                        prev_prev_prev_two_weeks_rsi = self.get_rsi_from_reference(prev_prev_prev_week_friday - timedelta(days=7), rsi_ref_data)
                        
                        if prev_prev_prev_rsi is not None and prev_prev_prev_two_weeks_rsi is not None:
                            prev_prev_week_mode = self.determine_mode(prev_prev_prev_rsi, prev_prev_prev_two_weeks_rsi, prev_prev_prev_week_mode)
                            print(f"🔍 전전주 모드 계산: 1주전 RSI={prev_prev_prev_rsi:.2f}, 2주전 RSI={prev_prev_prev_two_weeks_rsi:.2f}, 이전 모드={prev_prev_prev_week_mode} → {prev_prev_week_mode}")
                            # 전전주 모드 저장
                            week_modes[prev_prev_week_friday_str] = prev_prev_week_mode
                    
                    # 이전 주차의 모드 계산 (전전주 모드 사용)
                    calculated_prev_week_mode = self.determine_mode(prev_week_prev_rsi, prev_week_two_weeks_rsi, prev_prev_week_mode)
                    print(f"🔍 이전 주차 모드 재계산: 1주전 RSI={prev_week_prev_rsi:.2f}, 2주전 RSI={prev_week_two_weeks_rsi:.2f}, 전전주 모드={prev_prev_week_mode} → {calculated_prev_week_mode}")
                    
                    # 이전 주차 모드 저장
                    week_modes[prev_week_friday_str] = calculated_prev_week_mode
                    
                    # 재계산된 모드를 사용 (더 정확함)
                    actual_prev_week_mode = calculated_prev_week_mode
                    # current_mode도 업데이트하여 동기화
                    if current_mode != calculated_prev_week_mode:
                        print(f"⚠️ 이전 주차 모드 불일치: current_mode={current_mode}, 재계산={calculated_prev_week_mode}")
                        current_mode = calculated_prev_week_mode
                    else:
                        print(f"✅ 이전 주차 모드 일치 확인: {current_mode}")
                else:
                    # RSI 데이터가 없으면 current_mode 사용
                    actual_prev_week_mode = current_mode
                    print(f"⚠️ 이전 주차 RSI 데이터 없음, current_mode 사용: {current_mode}")
                
                # 이번 주 모드 결정 (이전 주차 모드 사용)
                new_mode = self.determine_mode(prev_week_rsi, two_weeks_ago_rsi, actual_prev_week_mode)
                
                # 12/29~1/2 주차 특별 검증
                if this_week_friday.date() >= datetime(2025, 12, 29).date() and this_week_friday.date() <= datetime(2026, 1, 2).date():
                    print(f"🔍 [12/29~1/2 주차 검증]")
                    print(f"   이번주 금요일: {this_week_friday.strftime('%Y-%m-%d')}")
                    print(f"   1주전 RSI: {prev_week_rsi:.2f}, 2주전 RSI: {two_weeks_ago_rsi:.2f}")
                    print(f"   이전 주차 모드: {actual_prev_week_mode}")
                    print(f"   결정된 모드: {new_mode}")
                    # 안전모드 조건 확인
                    safe_cond1 = two_weeks_ago_rsi > 65 and two_weeks_ago_rsi > prev_week_rsi
                    safe_cond2 = 40 < two_weeks_ago_rsi < 50 and two_weeks_ago_rsi > prev_week_rsi
                    safe_cond3 = two_weeks_ago_rsi >= 50 and prev_week_rsi < 50
                    ag_cond1 = two_weeks_ago_rsi < 50 and two_weeks_ago_rsi < prev_week_rsi and prev_week_rsi > 50
                    ag_cond2 = 50 < two_weeks_ago_rsi < 60 and two_weeks_ago_rsi < prev_week_rsi
                    ag_cond3 = two_weeks_ago_rsi < 35 and two_weeks_ago_rsi < prev_week_rsi
                    print(f"   안전모드 조건: cond1={safe_cond1}, cond2={safe_cond2}, cond3={safe_cond3}")
                    print(f"   공세모드 조건: cond1={ag_cond1}, cond2={ag_cond2}, cond3={ag_cond3}")
                    if not (safe_cond1 or safe_cond2 or safe_cond3) and not (ag_cond1 or ag_cond2 or ag_cond3):
                        print(f"   ⚠️ 조건 충족 없음 → 이전 주차 모드 유지: {actual_prev_week_mode}")
                        # 이전 주차가 안전모드였다면 이번 주도 안전모드여야 함
                        if actual_prev_week_mode == "SF" and new_mode != "SF":
                            print(f"   ❌ CRITICAL: 이전 주차가 안전모드인데 이번 주가 공세모드로 결정됨! 강제로 안전모드로 수정")
                            new_mode = "SF"
                
                # 디버깅: 모드 결정 과정 로그
                prev_rsi_display = f"{prev_week_rsi:.2f}" if prev_week_rsi is not None else "None"
                two_weeks_rsi_display = f"{two_weeks_ago_rsi:.2f}" if two_weeks_ago_rsi is not None else "None"
                print(f"🔍 주차 모드 결정: 날짜={current_date.strftime('%Y-%m-%d')}, 이번주 금요일={this_week_friday.strftime('%Y-%m-%d')}")
                print(f"   1주전 금요일: {prev_week_friday.strftime('%Y-%m-%d')}, RSI: {prev_rsi_display}")
                print(f"   2주전 금요일: {two_weeks_ago_friday.strftime('%Y-%m-%d')}, RSI: {two_weeks_rsi_display}")
                print(f"   이전 주차 모드: {current_mode} → 결정된 모드: {new_mode}")
                
                # 12/29~1/2 주차 특별 디버깅
                if this_week_friday.date() >= datetime(2025, 12, 29).date() and this_week_friday.date() <= datetime(2026, 1, 2).date():
                    print(f"⚠️ [12/29~1/2 주차 디버깅]")
                    print(f"   이번주 금요일: {this_week_friday.strftime('%Y-%m-%d')}")
                    print(f"   1주전 RSI: {prev_week_rsi:.2f}, 2주전 RSI: {two_weeks_ago_rsi:.2f}")
                    print(f"   이전 주차 모드: {current_mode}")
                    print(f"   결정된 모드: {new_mode}")
                    # determine_mode 조건 확인
                    safe_cond1 = prev_week_rsi > 65 and prev_week_rsi > two_weeks_ago_rsi
                    safe_cond2 = 40 < prev_week_rsi < 50 and prev_week_rsi > two_weeks_ago_rsi
                    safe_cond3 = prev_week_rsi >= 50 and two_weeks_ago_rsi < 50
                    ag_cond1 = prev_week_rsi < 50 and prev_week_rsi < two_weeks_ago_rsi and two_weeks_ago_rsi > 50
                    ag_cond2 = 50 < prev_week_rsi < 60 and prev_week_rsi < two_weeks_ago_rsi
                    ag_cond3 = prev_week_rsi < 35 and prev_week_rsi < two_weeks_ago_rsi
                    print(f"   안전모드 조건: cond1={safe_cond1}, cond2={safe_cond2}, cond3={safe_cond3}")
                    print(f"   공세모드 조건: cond1={ag_cond1}, cond2={ag_cond2}, cond3={ag_cond3}")
                    if not (safe_cond1 or safe_cond2 or safe_cond3) and not (ag_cond1 or ag_cond2 or ag_cond3):
                        print(f"   ⚠️ 조건 충족 없음 → 이전 모드 유지: {current_mode}")
                
                if new_mode != current_mode:
                    print(f"🔄 백테스팅 모드 전환: {current_mode} → {new_mode} (1주전 RSI: {prev_rsi_display}, 2주전 RSI: {two_weeks_rsi_display})")
                    print(f"   현재 회차: {self.current_round} → 최대 회차: {7 if new_mode == 'SF' else 8}")
                else:
                    # 모드가 변경되지 않았어도 주차 시작 시점임을 명확히 표시
                    print(f"📅 주차 시작: {current_date.strftime('%Y-%m-%d')} (모드: {current_mode} 유지)")
                
                # 모드 업데이트 전 검증
                if current_mode != new_mode:
                    print(f"⚠️ 모드 변경: {current_mode} → {new_mode}")
                else:
                    print(f"✅ 모드 유지: {current_mode}")
                
                # 현재 주차의 모드 저장 (먼저 저장)
                this_week_friday_str = this_week_friday.strftime('%Y-%m-%d')
                week_modes[this_week_friday_str] = new_mode
                
                # current_mode 업데이트 (week_modes와 동기화)
                current_mode = new_mode
                self.current_mode = new_mode  # 클래스 변수도 업데이트 (모드가 변경되지 않았어도 주차 시작 시점에 명확히 설정)
                
                # 12/29~1/2 주차 강제 검증 및 수정
                if this_week_friday.date() >= datetime(2025, 12, 29).date() and this_week_friday.date() <= datetime(2026, 1, 2).date():
                    print(f"🔍 [12/29~1/2 주차 강제 검증]")
                    print(f"   이전 주차 모드: {actual_prev_week_mode}")
                    print(f"   결정된 모드: {new_mode}")
                    if actual_prev_week_mode == "SF":
                        if new_mode != "SF":
                            print(f"❌ CRITICAL: 이전 주차가 안전모드인데 이번 주가 공세모드로 결정됨! 강제로 안전모드로 수정")
                            new_mode = "SF"
                            current_mode = "SF"
                            self.current_mode = "SF"
                            week_modes[this_week_friday_str] = "SF"
                        else:
                            print(f"✅ 이전 주차가 안전모드이고 이번 주도 안전모드로 올바르게 결정됨")
                    else:
                        print(f"⚠️ 이전 주차가 공세모드: {actual_prev_week_mode}")
                
                # 모드 변경 시 current_round 유지 (최대 회차만 변경)
                
                current_week += 1  # 주차 번호 증가 (0 → 1, 1 → 2, ...)
                current_rsi_display = f"{current_week_rsi:.2f}" if current_week_rsi is not None else "None"
                print(f"📅 주차 {current_week}: ~{this_week_friday.strftime('%m-%d')} | RSI: {current_rsi_display} | 모드: {current_mode} | self.current_mode: {self.current_mode}")
            
            # 매매 실행 (전일 종가가 있는 경우만)
            if prev_close is not None:

                # 현재 모드 설정 가져오기 (주차 단위로 결정된 모드 사용)
                config = self.sf_config if current_mode == "SF" else self.ag_config
                
                # 디버깅: 매매 실행 시점의 모드 확인
                debug_mode_msg = f"🔍 {current_date.strftime('%Y-%m-%d')} 매매 실행 - 현재 주차 모드: {current_mode}"
                print(debug_mode_msg)
                self.backtest_logs.append(debug_mode_msg)
                

                # 매수/매도 가격 계산 (전일 종가 기준)
                buy_price = prev_close * (1 + config["buy_threshold"] / 100)  # 매수가
                sell_price = prev_close * (1 + config["sell_threshold"] / 100)  # 매도가 (임시, 매수 체결 시 재계산됨)
                
                # 매도 조건 확인 및 실행
                # 1월 13일 특별 디버깅
                if current_date.strftime('%Y-%m-%d') == '2025-01-13':
                    print(f"⚠️ 1월 13일 특별 디버깅 (매도 조건 확인 전):")
                    print(f"   - 현재 포지션 수: {len(self.positions)}")
                    if self.positions:
                        print(f"   - 보유 포지션 목록:")
                        for pos in self.positions:
                            buy_date_str = pos['buy_date'].strftime('%Y-%m-%d') if isinstance(pos['buy_date'], (datetime, pd.Timestamp)) else str(pos['buy_date'])
                            pos_config = self.sf_config if pos["mode"] == "SF" else self.ag_config
                            target_price = pos["buy_price"] * (1 + pos_config["sell_threshold"] / 100)
                            print(f"      {pos['round']}회차: 매수일 {buy_date_str}, 모드 {pos.get('mode', 'N/A')}, 매수가 ${pos.get('buy_price', 0):.2f}, 목표가 ${target_price:.2f}")
                            print(f"         당일 종가: ${row['Close']:.2f}, 매도 조건: {row['Close']:.2f} >= {target_price:.2f} = {row['Close'] >= target_price}")
                
                # ── 매수 회차를 매도 처리 전에 미리 결정 (보유 포지션 수 + 1) ──
                # LOC 주문 특성상 매수/매도가 동시에 장 마감 시 체결되므로,
                # 매수 회차는 매도 전 보유 포지션 수 기준으로 결정해야 함
                buy_round_for_today = len(self.positions) + 1
                
                sell_recommendations = self.check_sell_conditions(row, current_date, prev_close)
                
                # 1월 13일 특별 디버깅 (매도 조건 확인 후)
                if current_date.strftime('%Y-%m-%d') == '2025-01-13':
                    print(f"⚠️ 1월 13일 특별 디버깅 (매도 조건 확인 후):")
                    print(f"   - 매도 추천 수: {len(sell_recommendations)}")
                    if sell_recommendations:
                        for sell_info in sell_recommendations:
                            pos = sell_info["position"]
                            buy_date_str = pos['buy_date'].strftime('%Y-%m-%d') if isinstance(pos['buy_date'], (datetime, pd.Timestamp)) else str(pos['buy_date'])
                            print(f"      매도 추천: {pos['round']}회차 (매수일: {buy_date_str}), 사유: {sell_info['reason']}, 매도가: ${sell_info['sell_price']:.2f}")

                daily_realized = 0
                sell_date = ""
                sell_executed_price = 0
                
                sold_rounds = []  # 매도된 회차들 추적
                sold_positions = []  # 매도된 포지션들 (매수 행에 기록용)
                
                for sell_info in sell_recommendations:
                    # will_sell이 True인 경우에만 실제 매도 실행
                    will_sell = sell_info.get("will_sell", True)  # 기본값은 True (기존 호환성)
                    
                    if not will_sell:
                        # 매도 조건 미충족 - 매도 추천 리스트에는 있지만 실제 매도는 하지 않음
                        continue
                    
                    position = sell_info["position"]
                    proceeds, sold_round = self.execute_sell(sell_info)
                    realized_pnl = proceeds - position["amount"]
                    daily_realized += realized_pnl
                    total_realized_pnl += realized_pnl
                    cash_balance += proceeds
                    sold_rounds.append(sold_round)
                    
                    # 매도 정보를 매수 행에 기록하기 위해 저장
                    # 요일을 한글로 변환
                    weekdays_korean = ['월', '화', '수', '목', '금', '토', '일']
                    weekday_korean = weekdays_korean[current_date.weekday()]
                    sold_positions.append({
                        "round": sold_round,
                        "sell_date": current_date.strftime(f"%m.%d.({weekday_korean})"),
                        "sell_price": sell_info["sell_price"],

                        "realized_pnl": realized_pnl
                    })
                
                # 매도 후 current_round를 매도 전 미리 결정한 값으로 설정
                # (매수와 매도는 LOC로 동시에 체결되므로, 매도 전 보유 수 기준)
                if sold_rounds:
                    self.current_round = buy_round_for_today
                    print(f"🔄 매도 발생: {len(sold_rounds)}건 매도 → 매수 회차는 매도 전 기준 유지: {self.current_round}회차")
                
                # 매수 조건 확인 및 실행
                buy_executed = False
                buy_price_executed = 0
                buy_quantity = 0
                buy_amount = 0
                current_round_before_buy = self.current_round  # 매수 전 회차 저장
                
                if self.can_buy_next_round():
                    # LOC 매수 조건: 매수가가 종가보다 유리할 때 (매수가 > 종가)
                    daily_close = row['Close']
                    
                    # 디버깅: 매수 조건 확인
                    log_msg = f"🔍 {current_date.strftime('%Y-%m-%d')} 매수 조건 확인:\n"
                    log_msg += f"   전일 종가(prev_close): ${prev_close:.2f}\n"
                    log_msg += f"   당일 종가(daily_close): ${daily_close:.2f}\n"
                    log_msg += f"   매수가(buy_price): ${buy_price:.2f} = prev_close * {1 + config['buy_threshold'] / 100}\n"
                    log_msg += f"   매수 조건: {buy_price:.2f} > {daily_close:.2f} = {buy_price > daily_close}\n"
                    log_msg += f"   현재 회차: {self.current_round}, 현금잔고: ${self.available_cash:,.0f}"
                    
                    print(log_msg)
                    self.backtest_logs.append(log_msg)
                    
                    if buy_price > daily_close:
                        success_msg = f"✅ 매수 조건 충족! 매수 실행 시도..."
                        print(success_msg)
                        self.backtest_logs.append(success_msg)
                        
                        # 매수 실행 전 모드 확인 및 검증
                        mode_before_buy = current_mode
                        if current_mode != self.current_mode:
                            print(f"⚠️ 매수 전 모드 불일치 감지! current_mode={current_mode}, self.current_mode={self.current_mode}")
                            print(f"   → current_mode를 self.current_mode로 동기화")
                            # 강제로 동기화 (self.current_mode가 더 최신일 수 있음)
                            current_mode = self.current_mode
                        
                        # 매수 시점의 모드 로그 (디버깅)
                        print(f"🔍 매수 실행 전: 날짜={current_date.strftime('%Y-%m-%d')}, 주차 모드={current_mode}, self.current_mode={self.current_mode}")
                        
                        if self.execute_buy(buy_price, daily_close, current_date, current_mode):  # 목표가 기준 수량으로 계산하여 종가에 매수, 매수 시점의 모드 전달
                            exec_msg = f"✅ 매수 체결 성공! (모드: {current_mode})"
                            print(exec_msg)
                            self.backtest_logs.append(exec_msg)
                            
                            buy_executed = True
                            position = self.positions[-1]
                            buy_price_executed = position["buy_price"]
                            buy_quantity = position["shares"]
                            buy_amount = position["amount"]
                            
                            # 디버깅: 저장된 모드 확인 및 검증
                            stored_mode = position.get("mode", "N/A")
                            if stored_mode != current_mode:
                                error_msg = f"❌ CRITICAL: 모드 불일치 감지! 매수일={current_date.strftime('%Y-%m-%d')}, 전달된 모드={current_mode}, 저장된 모드={stored_mode}"
                                print(error_msg)
                                self.backtest_logs.append(error_msg)
                                # 강제로 올바른 모드로 수정
                                position["mode"] = current_mode
                                self.positions[-1]["mode"] = current_mode
                                print(f"🔧 모드 수정 완료: {stored_mode} → {current_mode}")
                            else:
                                print(f"✅ 모드 일치 확인: 매수일={current_date.strftime('%Y-%m-%d')}, 모드={stored_mode}")
                            total_invested += buy_amount
                            cash_balance -= buy_amount
                            
                            # 매수 체결 시 매도목표가 재계산 (매수체결된 날의 종가 기준)
                            sell_price = daily_close * (1 + config["sell_threshold"] / 100)
                            
                            # 매수 행에서 매도 정보 초기화 (나중에 매도되면 업데이트됨)
                            sell_date = ""
                            sell_executed_price = 0
                        else:
                            fail_msg = f"❌ 매수 실행 실패 (execute_buy returned False)"
                            print(fail_msg)
                            self.backtest_logs.append(fail_msg)
                    else:
                        nocond_msg = f"❌ 매수 조건 불충족: {buy_price:.2f} <= {daily_close:.2f}"
                        print(nocond_msg)
                        self.backtest_logs.append(nocond_msg)
                else:
                    nobuy_msg = f"❌ 매수 불가능: can_buy_next_round() = False"
                    print(nobuy_msg)
                    self.backtest_logs.append(nobuy_msg)
                
                # 일일 처리 완료 후 다음 날을 위한 current_round 재계산
                # (매수/매도 모두 완료된 후의 보유 포지션 수 기준)
                self.current_round = len(self.positions) + 1
                if sold_rounds:
                    print(f"🔄 일일 처리 완료 (매도 {len(sold_rounds)}건): 보유 {len(self.positions)}개 → 다음 날 매수 회차: {self.current_round}")
                
                # 10거래일마다 투자원금 업데이트 (매매 처리 완료 후 실행)
                # 매수추천 시점의 투자원금과 백테스트 실제 매수 시 투자원금이 동일하도록
                # 매매 후에 갱신하여, 다음 거래일부터 적용되게 함
                if self.is_trading_day(current_date) and self.trading_days_count % 10 == 0 and self.trading_days_count > 0:
                    # 현재 총자산 계산 (현금 + 보유주식 평가금액)
                    total_shares_for_update = sum([pos["shares"] for pos in self.positions])
                    total_assets_for_update = self.available_cash + (total_shares_for_update * current_price)
                    
                    # 투자원금 업데이트
                    old_capital = self.current_investment_capital
                    self.current_investment_capital = total_assets_for_update
                    
                    print(f"💰 투자원금 업데이트: {self.trading_days_count}거래일째 - ${old_capital:,.0f} → ${total_assets_for_update:,.0f}")
                
                # 현재 보유 주식수와 평가손익 계산
                total_shares = sum([pos["shares"] for pos in self.positions])
                position_value = total_shares * current_price
                
                # 보유 주식의 매수 원가 계산
                total_buy_cost = sum([pos["amount"] for pos in self.positions])
                
                
                # 일별 기록 생성
                # 요일을 한글로 변환
                weekdays_korean = ['월', '화', '수', '목', '금', '토', '일']
                weekday_korean = weekdays_korean[current_date.weekday()]
                
                # 매도 정보 초기화 (현재 날짜의 매수 행에는 매도 정보 없음)
                sell_date_final = ""
                sell_executed_price_final = 0
                realized_pnl_final = 0
                
                daily_record = {
                    "date": current_date.strftime("%Y-%m-%d"),  # 표준 ISO 형식으로 변경
                    "week": current_week,
                    "rsi": current_week_rsi if current_week_rsi is not None else 50.0,  # None일 때만 기본값 사용
                    "mode": current_mode,
                    "current_round": min(current_round_before_buy, 7 if current_mode == "SF" else 8),  # 매수 전 회차 사용 (최대값 제한)
                    "seed_amount": self.calculate_position_size(current_round_before_buy) if buy_executed else 0,
                    "buy_order_price": buy_price,
                    "close_price": current_price,
                    "sell_target_price": sell_price,
                    "stop_loss_date": self.calculate_stop_loss_date(current_date, config["max_hold_days"]),
                    "d": 0,  # D 컬럼 (의미 불명)
                    "trading_days": i + 1,
                    "buy_executed_price": buy_price_executed,
                    "buy_quantity": buy_quantity,
                    "buy_amount": buy_amount,
                    "buy_round": current_round_before_buy if buy_executed else 0,  # 매수 회차 저장
                    "commission": 0.0,
                    "sell_date": sell_date_final,
                    "sell_executed_price": sell_executed_price_final,
                    "holding_days": 0,  # 보유기간 (거래일 기준)
                    "holdings": total_shares,
                    "realized_pnl": realized_pnl_final,
                    "cumulative_realized": total_realized_pnl,
                    "daily_realized": daily_realized,
                    "update": False,
                    "investment_update": self.initial_capital,
                    "withdrawal": False,
                    "withdrawal_amount": 0,
                    "seed_increase": 0,
                    "position_value": position_value,
                    "cash_balance": cash_balance,
                    "total_assets": cash_balance + position_value
                }
                
                daily_records.append(daily_record)
                
                # 오늘 매도된 포지션들의 정보를 과거 매수 행에 기록 (daily_record 생성 후)
                if sold_positions:
                    for sold_pos in sold_positions:
                        
                        # 해당 회차의 매수 행을 찾아서 매도 정보 업데이트
                        found = False
                        for record in daily_records:
                            if (record.get('buy_executed_price', 0) > 0 and 
                                record.get('buy_quantity', 0) > 0 and
                                record.get('sell_date', '') == ''):  # 아직 매도되지 않은 행
                                
                                # 해당 회차인지 확인 (buy_round로 정확한 매칭)
                                if record.get('buy_round', 0) == sold_pos["round"]:
                                    # 보유기간 계산 (거래일 기준)
                                    try:
                                        buy_date_str = record['date']
                                        sell_date_str = sold_pos["sell_date"]
                                        
                                        # 날짜 파싱 (예: "25.01.02.(목)" -> "2025-01-02")
                                        buy_date_str_clean = buy_date_str.split('(')[0].strip().rstrip('.')
                                        sell_date_str_clean = sell_date_str.split('(')[0].strip().rstrip('.')
                                        
                                        buy_date = datetime.strptime(buy_date_str_clean, "%y.%m.%d")
                                        sell_date = datetime.strptime(sell_date_str_clean, "%m.%d")
                                        
                                        # 연도 보정 (매도일에는 연도가 없으므로)
                                        if sell_date.month < buy_date.month or (sell_date.month == buy_date.month and sell_date.day < buy_date.day):
                                            sell_date = sell_date.replace(year=buy_date.year + 1)
                                        else:

                                            sell_date = sell_date.replace(year=buy_date.year)
                                        
                                        # 거래일 계산 (주말 + 휴장일 제외)
                                        holding_days = 0
                                        temp_date = buy_date
                                        while temp_date <= sell_date:
                                            if self.is_trading_day(temp_date):
                                                holding_days += 1
                                            temp_date += timedelta(days=1)
                                        
                                        record['holding_days'] = holding_days
                                        
                                    except Exception as e:
                                        print(f"⚠️ 보유기간 계산 오류: {e}")
                                        record['holding_days'] = 0
                                    
                                    record['sell_date'] = sold_pos["sell_date"]
                                    record['sell_executed_price'] = sold_pos["sell_price"]
                                    record['realized_pnl'] = sold_pos["realized_pnl"]
                                    found = True
                                    break
                        
            
            # 진행상황 출력
            if (i + 1) % 10 == 0:
                print(f"진행: {i+1}/{len(soxl_backtest)}일 ({(i+1)/len(soxl_backtest)*100:.1f}%)")

            
            prev_close = current_price
        
        # 최종 결과 계산
        
        # 백테스팅 완료 후 current_round를 올바르게 설정
        # 보유 중인 회차 수 + 1 = 다음 매수 회차
        holding_rounds = len(self.positions)
        self.current_round = holding_rounds + 1
        print(f"🔄 백테스팅 완료 후 current_round 설정: 보유 {holding_rounds}개 → 다음 매수 {self.current_round}회차")

        final_value = daily_records[-1]["total_assets"] if daily_records else self.initial_capital
        total_return = ((final_value - self.initial_capital) / self.initial_capital) * 100
        
        summary = {
            "start_date": start_date,
            "end_date": end_date or datetime.now().strftime("%Y-%m-%d"),

            "trading_days": len(soxl_backtest),
            "initial_capital": self.initial_capital,
            "final_value": final_value,
            "total_return": total_return,
            "final_positions": len(self.positions),

            "daily_records": daily_records,
            "logs": self.backtest_logs if hasattr(self, 'backtest_logs') else []
        }

        
        # MDD 계산 및 출력
        mdd_info = self.calculate_mdd(daily_records)
        
        print("✅ 백테스팅 완료!")

        print(f"\n📊 백테스팅 결과 요약:")
        print(f"   📅 기간: {start_date} ~ {end_date or datetime.now().strftime('%Y-%m-%d')}")
        print(f"   💰 초기자본: ${self.initial_capital:,.0f}")
        print(f"   💰 최종자산: ${final_value:,.0f}")
        print(f"   📈 총수익률: {total_return:+.2f}%")
        print(f"   📦 최종보유포지션: {len(self.positions)}개")
        print(f"\n⚠️ 리스크 지표:")
        print(f"   📉 MDD (최대낙폭): {mdd_info.get('mdd_percent', 0.0):.2f}%")
        print(f"   📅 MDD 발생일: {mdd_info.get('mdd_date', '')}")
        print(f"   💰 최저자산: ${mdd_info.get('mdd_value', 0.0):,.0f}")
        print(f"   📅 MDD 발생 최고자산일: {mdd_info.get('mdd_peak_date', '')}")
        print(f"   📅 최고자산일: {mdd_info.get('overall_peak_date', '')}")
        print(f"   💰 최고자산: ${mdd_info.get('overall_peak_value', 0.0):,.0f}")
        
        return summary
    

    
    def get_week_number(self, date: datetime) -> int:
        """날짜로부터 주차 계산"""
        year = date.year
        week_num = date.isocalendar()[1]
        return f"{year}W{week_num:02d}"
    
    def calculate_mdd(self, daily_records: List[Dict]) -> Dict:
        """
        MDD (Maximum Drawdown) 계산
        Args:
            daily_records: 일별 백테스팅 기록
        Returns:
            Dict: MDD 정보
        """
        if not daily_records:
            return {
                "mdd_percent": 0.0, 
                "mdd_date": "", 
                "mdd_value": 0.0, 
                "mdd_peak_date": "",  # MDD 계산 시점의 최고자산일
                "overall_peak_date": "",  # 전체 기간 최고자산일
                "overall_peak_value": 0.0  # 전체 기간 최고자산
            }
        
        max_assets = 0.0
        max_drawdown = 0.0
        mdd_peak_date = ""  # MDD 계산 시점의 최고자산일
        mdd_date = ""
        mdd_value = 0.0
        
        # 전체 기간 최고자산 추적
        overall_max_assets = 0.0
        overall_peak_date = ""
        
        # MDD 계산용 변수들
        current_peak_assets = 0.0
        current_peak_date = ""
        
        for record in daily_records:
            current_assets = record.get('total_assets', 0.0)
            
            # 전체 기간 최고자산 갱신
            if current_assets > overall_max_assets:
                overall_max_assets = current_assets
                overall_peak_date = record.get('date', '')
            
            # 새로운 최고자산 갱신 (MDD 계산용)
            if current_assets > current_peak_assets:
                current_peak_assets = current_assets
                current_peak_date = record.get('date', '')
            
            # 현재 자산이 현재 최고자산보다 낮으면 낙폭 계산
            if current_peak_assets > 0:
                drawdown = (current_peak_assets - current_assets) / current_peak_assets * 100
                if drawdown > max_drawdown:
                    max_drawdown = drawdown
                    mdd_date = record.get('date', '')
                    mdd_value = current_assets
                    mdd_peak_date = current_peak_date  # MDD 발생 시점의 기준 최고자산일
        
        return {
            "mdd_percent": max_drawdown,
            "mdd_date": mdd_date,
            "mdd_value": mdd_value,
            "mdd_peak_date": mdd_peak_date,  # MDD 계산 시점의 최고자산일
            "overall_peak_date": overall_peak_date,  # 전체 기간 최고자산일
            "overall_peak_value": overall_max_assets  # 전체 기간 최고자산
        }
    
    def export_backtest_to_excel(self, backtest_result: Dict, filename: str = None):
        """
        백테스팅 결과를 엑셀 파일로 내보내기
        Args:
            backtest_result: 백테스팅 결과
            filename: 파일명 (None이면 자동 생성)
        """
        if "error" in backtest_result:
            print(f"❌ 엑셀 내보내기 실패: {backtest_result['error']}")
            return
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"SOXL_백테스팅_{backtest_result['start_date']}_{timestamp}.xlsx"
        
        # 엑셀 워크북 생성
        wb = openpyxl.Workbook()

        
        # 가운데 정렬 설정
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "백테스팅 요약"

        
        # 첫 번째 행 고정 (헤더 고정)
        ws_summary.freeze_panes = "A2"
        
        # MDD 계산
        mdd_info = self.calculate_mdd(backtest_result['daily_records'])
        
        # 요약 데이터 작성
        summary_data = [
            ["SOXL 퀀트투자 백테스팅 결과", ""],
            ["", ""],
            ["시작일", backtest_result['start_date']],
            ["종료일", backtest_result['end_date']],
            ["거래일수", f"{backtest_result['trading_days']}일"],
            ["", ""],
            ["초기자본", f"${backtest_result['initial_capital']:,.0f}"],
            ["최종자산", f"${backtest_result['final_value']:,.0f}"],
            ["총수익률", f"{backtest_result['total_return']:+.2f}%"],

            ["최종보유포지션", f"{backtest_result['final_positions']}개"],
            ["", ""],

            ["=== 리스크 지표 ===", ""],
            ["MDD (최대낙폭)", f"{mdd_info.get('mdd_percent', 0.0):.2f}%"],
            ["MDD 발생일", mdd_info.get('mdd_date', '')],
            ["최저자산", f"${mdd_info.get('mdd_value', 0.0):,.0f}"],
            ["MDD 발생 최고자산일", mdd_info.get('mdd_peak_date', '')],
            ["최고자산일", mdd_info.get('overall_peak_date', '')],
            ["최고자산", f"${mdd_info.get('overall_peak_value', 0.0):,.0f}"]
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 1):

            cell1 = ws_summary.cell(row=row_idx, column=1, value=label)
            cell2 = ws_summary.cell(row=row_idx, column=2, value=value)
            cell1.alignment = center_alignment
            cell2.alignment = center_alignment
        
        # 스타일 적용
        title_font = Font(size=16, bold=True)

        title_cell = ws_summary.cell(row=1, column=1)
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        
        # 상세 거래 내역 시트

        ws_detail = wb.create_sheet("매매 상세내역")
        

        # 첫 번째 행 고정 (헤더 고정)
        ws_detail.freeze_panes = "A2"
        
        # 헤더 작성 (실제 양식에 맞게)
        headers = [

            "날짜", "주차", "RSI", "모드", "현재회차", "1회시드", 
            "매수주문가", "종가", "매도목표가", "손절예정일", "거래일수", 
            "매수체결", "수량", "매수대금", "매도일", "매도체결", "보유기간",
            "보유", "실현손익", "누적실현", "당일실현",
            "예수금", "총자산"
        ]
        
        header_font = Font(size=11, bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            cell.alignment = center_alignment
        
        # 데이터 작성

        prev_close_price = None  # 전일 종가 추적용
        
        for row_idx, record in enumerate(backtest_result['daily_records'], 2):
            # 날짜 (첫 데이터와 매주 월요일은 볼드체)
            cell = ws_detail.cell(row=row_idx, column=1, value=record['date'])
            cell.alignment = center_alignment
            
            # 첫 데이터 또는 월요일 체크
            if row_idx == 2:  # 첫 데이터
                cell.font = Font(bold=True)
            else:
                # 날짜에서 요일 추출 (예: "25.01.02.(목)" -> "월")
                date_str = record['date']
                if '(월)' in date_str:
                    cell.font = Font(bold=True)
            
            # 주차
            cell = ws_detail.cell(row=row_idx, column=2, value=record['week'])
            cell.alignment = center_alignment
            
            # RSI
            rsi_value = record.get('rsi', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=3, value=f"{rsi_value:.2f}")
            cell.alignment = center_alignment
            
            # 모드 (SF: 초록색 글자, AG: 주황색 글자)
            cell = ws_detail.cell(row=row_idx, column=4, value=record['mode'])
            cell.alignment = center_alignment
            
            if record['mode'] == 'SF':
                cell.font = Font(color="008000")  # 초록색 글자
            elif record['mode'] == 'AG':
                cell.font = Font(color="FF8C00")  # 주황색 글자
            
            # 현재회차
            cell = ws_detail.cell(row=row_idx, column=5, value=record['current_round'])
            cell.alignment = center_alignment
            
            # 1회시드
            seed_amount = record.get('seed_amount', 0.0) or 0.0
            if seed_amount > 0:
                cell = ws_detail.cell(row=row_idx, column=6, value=f"${seed_amount:,.0f}")
            else:
                cell = ws_detail.cell(row=row_idx, column=6, value="")
            cell.alignment = center_alignment
            
            # 매수주문가
            buy_order_price = record.get('buy_order_price', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=7, value=f"${buy_order_price:.2f}")
            cell.alignment = center_alignment
            
            # 종가 (어제 대비 상승: 빨간색, 하락: 파란색)
            close_price = record.get('close_price', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=8, value=f"{close_price:.2f}")
            cell.alignment = center_alignment
            
            # 전일 대비 상승/하락 색상 적용
            if prev_close_price is not None:
                if close_price > prev_close_price:
                    cell.font = Font(color="FF0000")  # 빨간색
                elif close_price < prev_close_price:
                    cell.font = Font(color="0000FF")  # 파란색
            
            prev_close_price = close_price  # 다음 행을 위해 저장
            
            # 매도목표가
            sell_target_price = record.get('sell_target_price', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=9, value=f"${sell_target_price:.2f}")
            cell.alignment = center_alignment
            
            # 손절예정일
            cell = ws_detail.cell(row=row_idx, column=10, value=record['stop_loss_date'])
            cell.alignment = center_alignment
            
            # 거래일수
            cell = ws_detail.cell(row=row_idx, column=11, value=record['trading_days'])
            cell.alignment = center_alignment
            
            # 매수체결 (빨간색)
            buy_executed_price = record.get('buy_executed_price', 0.0) or 0.0
            if buy_executed_price > 0:
                cell = ws_detail.cell(row=row_idx, column=12, value=f"${buy_executed_price:.2f}")
                cell.font = Font(color="FF0000")  # 빨간색
            else:
                cell = ws_detail.cell(row=row_idx, column=12, value="")
            cell.alignment = center_alignment
            
            # 수량 (매수체결 시 빨간색)
            buy_quantity = record.get('buy_quantity', 0) or 0
            if buy_quantity > 0:
                cell = ws_detail.cell(row=row_idx, column=13, value=buy_quantity)
                cell.font = Font(color="FF0000")  # 빨간색
            else:
                cell = ws_detail.cell(row=row_idx, column=13, value="")
            cell.alignment = center_alignment
            
            # 매수대금 (매수체결 시 빨간색)
            buy_amount = record.get('buy_amount', 0.0) or 0.0
            if buy_amount > 0:
                cell = ws_detail.cell(row=row_idx, column=14, value=f"${buy_amount:,.0f}")
                cell.font = Font(color="FF0000")  # 빨간색
            else:
                cell = ws_detail.cell(row=row_idx, column=14, value="")
            cell.alignment = center_alignment
            
            # 매도일 (파란색 글씨)
            cell = ws_detail.cell(row=row_idx, column=15, value=record['sell_date'])
            cell.alignment = center_alignment
            if record['sell_date']:  # 매도일이 있는 경우에만 파란색 적용
                cell.font = Font(color="0000FF")  # 파란색 글씨
            
            # 매도체결 (파란색 글씨)
            sell_executed_price = record.get('sell_executed_price', 0.0) or 0.0
            if sell_executed_price > 0:
                cell = ws_detail.cell(row=row_idx, column=16, value=f"${sell_executed_price:.2f}")
                cell.font = Font(color="0000FF")  # 파란색 글씨
            else:
                cell = ws_detail.cell(row=row_idx, column=16, value="")
            cell.alignment = center_alignment
            
            # 보유기간
            holding_days = record.get('holding_days', 0) or 0
            if holding_days > 0:
                cell = ws_detail.cell(row=row_idx, column=17, value=f"{holding_days}일")
            else:
                cell = ws_detail.cell(row=row_idx, column=17, value="")
            cell.alignment = center_alignment
            
            # 보유
            cell = ws_detail.cell(row=row_idx, column=18, value=record['holdings'])
            cell.alignment = center_alignment
            
            # 실현손익
            realized_pnl = record.get('realized_pnl', 0.0) or 0.0
            if realized_pnl != 0:
                cell = ws_detail.cell(row=row_idx, column=19, value=f"${realized_pnl:,.0f}")
            else:
                cell = ws_detail.cell(row=row_idx, column=19, value="")
            cell.alignment = center_alignment
            
            # 누적실현
            cumulative_realized = record.get('cumulative_realized', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=20, value=f"${cumulative_realized:,.0f}")
            cell.alignment = center_alignment
            cell.font = Font(color="FF0000")  # 빨간색
            
            # 당일실현
            daily_realized = record.get('daily_realized', 0.0) or 0.0
            if daily_realized != 0:
                cell = ws_detail.cell(row=row_idx, column=21, value=f"${daily_realized:,.0f}")
            else:
                cell = ws_detail.cell(row=row_idx, column=21, value="")
            cell.alignment = center_alignment
            
            # 예수금
            cash_balance = record.get('cash_balance', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=22, value=f"${cash_balance:,.0f}")
            cell.alignment = center_alignment
            
            # 총자산 (숫자 형식으로 저장)
            total_assets = record.get('total_assets', 0.0) or 0.0
            cell = ws_detail.cell(row=row_idx, column=23, value=total_assets)
            cell.alignment = center_alignment
            # 숫자 형식으로 표시 (천 단위 구분자 포함)
            cell.number_format = '#,##0'
        
        # 열 너비 자동 조정
        for ws in [ws_summary, ws_detail]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 25)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        try:
            wb.save(filename)
            print(f"✅ 백테스팅 결과가 엑셀 파일로 저장되었습니다: {filename}")
            return filename
        except Exception as e:
            print(f"❌ 엑셀 파일 저장 실패: {e}")
            return None

def main():
    """메인 실행 함수"""
    print("🚀 SOXL 퀀트투자 시스템")
    print("=" * 50)
    

    # 투자원금 사용자 입력
    while True:
        try:
            initial_capital_input = input("💰 초기 투자금을 입력하세요 (달러): ").strip()
            if not initial_capital_input:
                initial_capital = 40000  # 기본값
                print(f"💰 투자원금: ${initial_capital:,.0f} (기본값)")
                break
            
            initial_capital = float(initial_capital_input)
            if initial_capital <= 0:
                print("❌ 투자금은 0보다 큰 값이어야 합니다.")
                continue
                
            print(f"💰 투자원금: ${initial_capital:,.0f}")
            break
            
        except ValueError:
            print("❌ 올바른 숫자를 입력해주세요.")
            continue
    
    # 트레이더 초기화
    trader = SOXLQuantTrader(initial_capital)
    
    # 시작일 입력(엔터 시 1년 전)
    start_date_input = input("📅 투자 시작일을 입력하세요 (YYYY-MM-DD, 엔터시 1년 전): ").strip()
    if not start_date_input:
        start_date_input = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
    trader.session_start_date = start_date_input
    
    # 시드증액 추가 (선택사항)
    print("\n💰 시드증액 설정 (선택사항)")
    print("- 시드증액을 추가하지 않으려면 엔터를 누르세요")
    while True:
        seed_date_input = input("시드증액 날짜 (YYYY-MM-DD, 엔터시 종료): ").strip()
        if not seed_date_input:
            break
        
        try:
            seed_amount_input = input("시드증액 금액 (달러, 양수: 증액, 음수: 인출): ").strip()
            if not seed_amount_input:
                break
            
            seed_amount = float(seed_amount_input)
            trader.add_seed_increase(seed_date_input, seed_amount, f"시드증액 {seed_date_input}")
            print(f"✅ 시드증액 추가됨: {seed_date_input} - ${seed_amount:,.0f}")
        except ValueError:
            print("❌ 올바른 숫자를 입력해주세요.")
        except Exception as e:
            print(f"❌ 오류: {e}")
    
    while True:
        print("\n" + "=" * 50)
        print("메뉴를 선택하세요:")
        print("1. 오늘의 매매 추천 보기")
        print("2. 포트폴리오 현황 보기")
        print("3. 백테스팅 실행")
        print("4. 매수 실행 (테스트)")
        print("5. 매도 실행 (테스트)")
        print("T. 테스트 날짜(오늘) 설정/해제")
        print("6. 종료")
        
        choice = input("\n선택 (1-6): ").strip()
        
        if choice == '1':
            # 저장된 시작일부터 오늘까지 시뮬레이션으로 현재 상태 산출
            start_date = trader.session_start_date or (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
            sim_result = trader.simulate_from_start_to_today(start_date, quiet=True)
            if "error" in sim_result:
                print(f"❌ 시뮬레이션 실패: {sim_result['error']}")
            
            # 현재 상태 기반 오늘의 추천 출력
            recommendation = trader.get_daily_recommendation()
            trader.print_recommendation(recommendation)
            
        elif choice == '2':
            # 저장된 시작일부터 오늘까지 시뮬레이션으로 현황 재계산
            start_date = trader.session_start_date or (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')
            sim_result = trader.simulate_from_start_to_today(start_date, quiet=True)
            if "error" in sim_result:
                print(f"❌ 시뮬레이션 실패: {sim_result['error']}")
            
            # 기존 형식 유지하여 현황 출력
            if trader.positions:
                print("\n💼 현재 포트폴리오:")
                print("-" * 30)
                for pos in trader.positions:
                    hold_days = (datetime.now() - pos['buy_date']).days
                    print(f"{pos['round']}회차: {pos['shares']}주 @ ${pos['buy_price']:.2f} ({hold_days}일)")
                print(f"\n현금잔고: ${trader.available_cash:,.0f}")
            else:
                print("\n보유 포지션이 없습니다.")
                print(f"현금잔고: ${trader.available_cash:,.0f}")
        
        elif choice == '3':
            # 백테스팅 실행
            print("\n📊 백테스팅 실행")
            print("-" * 30)
            
            start_date = input("시작 날짜를 입력하세요 (YYYY-MM-DD): ").strip()
            if not start_date:
                print("날짜를 입력해주세요.")
                continue
            
            end_date = input("종료 날짜를 입력하세요 (YYYY-MM-DD, 엔터시 오늘까지): ").strip()
            if not end_date:
                end_date = None
            
            print("\n백테스팅을 시작합니다...")
            backtest_result = trader.run_backtest(start_date, end_date)
            
            if "error" in backtest_result:
                print(f"❌ 백테스팅 실패: {backtest_result['error']}")
                continue
            

            # MDD 계산
            mdd_info = trader.calculate_mdd(backtest_result['daily_records'])
            
            # 결과 출력
            print("\n" + "=" * 60)
            print("📊 백테스팅 결과 요약")
            print("=" * 60)
            print(f"기간: {backtest_result['start_date']} ~ {backtest_result['end_date']}")
            print(f"거래일수: {backtest_result['trading_days']}일")
            print(f"초기자본: ${backtest_result['initial_capital']:,.0f}")
            print(f"최종자산: ${backtest_result['final_value']:,.0f}")
            print(f"총수익률: {backtest_result['total_return']:+.2f}%")

            print(f"최대 MDD: {mdd_info.get('mdd_percent', 0.0):.2f}%")
            print(f"최종보유포지션: {backtest_result['final_positions']}개")

            print(f"총 거래일수: {len(backtest_result['daily_records'])}일")
            
            # 엑셀 내보내기 여부 확인
            export_choice = input("\n엑셀 파일로 내보내시겠습니까? (y/n): ").strip().lower()
            if export_choice == 'y':
                filename = trader.export_backtest_to_excel(backtest_result)
                if filename:
                    print(f"📁 파일 위치: {os.path.abspath(filename)}")
            
        elif choice == '4':
            print("\n🔧 매수 테스트 기능 (개발 중)")
            
        elif choice == '5':
            print("\n🔧 매도 테스트 기능 (개발 중)")
            
        elif choice.lower() == 't':
            print("\n🧪 테스트 날짜 설정")
            print("- 비우고 엔터하면 해제됩니다")
            test_date = input("테스트 오늘 날짜 (YYYY-MM-DD): ").strip()
            trader.set_test_today(test_date if test_date else None)
            
        elif choice == '6':
            print("프로그램을 종료합니다.")
            break
            
        else:
            print("올바른 선택지를 입력하세요.")

if __name__ == "__main__":
    main()

