"""
íŒŒë¼ë¯¸í„°.xlsx íŒŒì¼ì—ì„œ íŒŒë¼ë¯¸í„°ë¥¼ ì½ì–´ì„œ ë°±í…ŒìŠ¤íŒ…ì„ ì‹¤í–‰í•˜ëŠ” ìŠ¤í¬ë¦½íŠ¸
ì´ˆê¸°ìì‚° 4000ë‹¬ëŸ¬, ê¸°ê°„ 2011-01-01 ~ 2025-12-07
"""
import openpyxl
from datetime import datetime, timedelta
from soxl_quant_system import SOXLQuantTrader


def safe_float(value, default=None):
    """ì•ˆì „í•˜ê²Œ floatë¡œ ë³€í™˜ (None, ë¹ˆ ë¬¸ìì—´, '-' ë“± ì²˜ë¦¬)"""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    value_str = str(value).strip()
    if value_str == '' or value_str == '-' or value_str.lower() == 'none':
        return default
    try:
        return float(value_str)
    except (ValueError, TypeError):
        return default

def safe_int(value, default=None):
    """ì•ˆì „í•˜ê²Œ intë¡œ ë³€í™˜ (None, ë¹ˆ ë¬¸ìì—´, '-' ë“± ì²˜ë¦¬)"""
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    value_str = str(value).strip()
    if value_str == '' or value_str == '-' or value_str.lower() == 'none':
        return default
    try:
        return int(float(value_str))
    except (ValueError, TypeError):
        return default


def load_parameters_from_excel(excel_file: str = "íŒŒë¼ë¯¸í„°.xlsx"):
    """
    ì—‘ì…€ íŒŒì¼ì—ì„œ íŒŒë¼ë¯¸í„° ì½ê¸°
    Args:
        excel_file: ì—‘ì…€ íŒŒì¼ ê²½ë¡œ
    Returns:
        tuple: (ag_config, sf_config)
    """
    import os
    try:
        # íŒŒì¼ ì¡´ì¬ í™•ì¸
        if not os.path.exists(excel_file):
            raise FileNotFoundError(f"ì—‘ì…€ íŒŒì¼ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {excel_file}")
        
        print(f"ğŸ“‚ ì—‘ì…€ íŒŒì¼ ì½ê¸°: {excel_file}")
        file_size = os.path.getsize(excel_file)
        file_mtime = datetime.fromtimestamp(os.path.getmtime(excel_file))
        print(f"   íŒŒì¼ í¬ê¸°: {file_size} bytes")
        print(f"   íŒŒì¼ ìˆ˜ì • ì‹œê°„: {file_mtime.strftime('%Y-%m-%d %H:%M:%S')}")
        
        # ì—‘ì…€ íŒŒì¼ ì—´ê¸°
        # data_only=False: íŒŒì¼ì´ ì—´ë ¤ìˆì–´ë„ ìµœì‹  ê°’ì„ ì½ì„ ìˆ˜ ìˆìŒ
        # read_only=True: ì½ê¸° ì „ìš© ëª¨ë“œ (ë” ë¹ ë¥´ê³  ë©”ëª¨ë¦¬ íš¨ìœ¨ì )
        # íŒŒì¼ì´ ì—´ë ¤ìˆìœ¼ë©´ read_only=Falseë¡œ ìë™ ì „í™˜
        try:
            # ë¨¼ì € read_only=Trueë¡œ ì‹œë„
            wb = openpyxl.load_workbook(excel_file, data_only=False, read_only=True)
            print("   âœ… ì½ê¸° ì „ìš© ëª¨ë“œë¡œ ì—´ê¸° ì„±ê³µ")
        except (PermissionError, Exception) as e:
            # íŒŒì¼ì´ ì—´ë ¤ìˆê±°ë‚˜ ì ê²¨ìˆëŠ” ê²½ìš° read_only=Falseë¡œ ì¬ì‹œë„
            print(f"   âš ï¸ ì½ê¸° ì „ìš© ëª¨ë“œ ì‹¤íŒ¨ ({str(e)[:50]}), ì¼ë°˜ ëª¨ë“œë¡œ ì¬ì‹œë„...")
            try:
                wb = openpyxl.load_workbook(excel_file, data_only=False, read_only=False)
                print("   âœ… ì¼ë°˜ ëª¨ë“œë¡œ ì—´ê¸° ì„±ê³µ (íŒŒì¼ì´ ì—´ë ¤ìˆì„ ìˆ˜ ìˆìŒ)")
            except Exception as e2:
                raise PermissionError(f"ì—‘ì…€ íŒŒì¼ì„ ì½ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤. íŒŒì¼ì„ ë‹«ê³  ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”: {excel_file}\nì˜¤ë¥˜: {e2}")
        
        ws = wb.active
        print(f"ğŸ“‹ í™œì„± ì‹œíŠ¸: {ws.title}")
        
        # ê³µì„¸ëª¨ë“œ(AG) íŒŒë¼ë¯¸í„° ì½ê¸°
        # ì—‘ì…€ì— ì†Œìˆ˜ì (0.035)ìœ¼ë¡œ ì €ì¥ë˜ì–´ ìˆìœ¼ë©´ í¼ì„¼íŠ¸(3.5)ë¡œ ë³€í™˜
        ag_buy_threshold_raw = safe_float(ws['B11'].value)  # ê³µì„¸ëª¨ë“œ ë§¤ìˆ˜ì„ê³„ê°’
        ag_sell_threshold_raw = safe_float(ws['B12'].value)  # ê³µì„¸ëª¨ë“œ ë§¤ë„ì„ê³„ê°’
        print(f"ğŸ” B11 ì›ë³¸ê°’: {ws['B11'].value} â†’ ë³€í™˜: {ag_buy_threshold_raw}")
        print(f"ğŸ” B12 ì›ë³¸ê°’: {ws['B12'].value} â†’ ë³€í™˜: {ag_sell_threshold_raw}")
        
        # ê°’ì´ 1ë³´ë‹¤ ì‘ìœ¼ë©´ ì†Œìˆ˜ì  ê°’ì´ë¯€ë¡œ 100ì„ ê³±í•´ì„œ í¼ì„¼íŠ¸ë¡œ ë³€í™˜
        ag_buy_threshold = ag_buy_threshold_raw * 100 if ag_buy_threshold_raw and ag_buy_threshold_raw < 1 else ag_buy_threshold_raw
        ag_sell_threshold = ag_sell_threshold_raw * 100 if ag_sell_threshold_raw and ag_sell_threshold_raw < 1 else ag_sell_threshold_raw
        ag_max_hold_days = safe_int(ws['B13'].value)  # ê³µì„¸ëª¨ë“œ ìµœëŒ€ë³´ìœ ì¼
        ag_split_count = safe_int(ws['B14'].value)  # ê³µì„¸ëª¨ë“œ ë¶„í• íšŸìˆ˜
        
        # ì•ˆì „ëª¨ë“œ(SF) íŒŒë¼ë¯¸í„° ì½ê¸°
        # ì—‘ì…€ì— ì†Œìˆ˜ì (0.035)ìœ¼ë¡œ ì €ì¥ë˜ì–´ ìˆìœ¼ë©´ í¼ì„¼íŠ¸(3.5)ë¡œ ë³€í™˜
        sf_buy_threshold_raw = safe_float(ws['B15'].value)  # ì•ˆì „ëª¨ë“œ ë§¤ìˆ˜ì„ê³„ê°’
        sf_sell_threshold_raw = safe_float(ws['B16'].value)  # ì•ˆì „ëª¨ë“œ ë§¤ë„ì„ê³„ê°’
        print(f"ğŸ” B15 ì›ë³¸ê°’: {ws['B15'].value} â†’ ë³€í™˜: {sf_buy_threshold_raw}")
        print(f"ğŸ” B16 ì›ë³¸ê°’: {ws['B16'].value} â†’ ë³€í™˜: {sf_sell_threshold_raw}")
        
        # ê°’ì´ 1ë³´ë‹¤ ì‘ìœ¼ë©´ ì†Œìˆ˜ì  ê°’ì´ë¯€ë¡œ 100ì„ ê³±í•´ì„œ í¼ì„¼íŠ¸ë¡œ ë³€í™˜
        sf_buy_threshold = sf_buy_threshold_raw * 100 if sf_buy_threshold_raw and sf_buy_threshold_raw < 1 else sf_buy_threshold_raw
        sf_sell_threshold = sf_sell_threshold_raw * 100 if sf_sell_threshold_raw and sf_sell_threshold_raw < 1 else sf_sell_threshold_raw
        sf_max_hold_days = safe_int(ws['B17'].value)  # ì•ˆì „ëª¨ë“œ ìµœëŒ€ë³´ìœ ì¼
        sf_split_count = safe_int(ws['B18'].value)  # ì•ˆì „ëª¨ë“œ ë¶„í• íšŸìˆ˜
        
        # í•„ìˆ˜ íŒŒë¼ë¯¸í„° ê²€ì¦
        if ag_buy_threshold is None or ag_sell_threshold is None:
            raise ValueError("ê³µì„¸ëª¨ë“œ ë§¤ìˆ˜/ë§¤ë„ ì„ê³„ê°’ì´ í•„ìš”í•©ë‹ˆë‹¤ (B11, B12)")
        if sf_buy_threshold is None or sf_sell_threshold is None:
            raise ValueError("ì•ˆì „ëª¨ë“œ ë§¤ìˆ˜/ë§¤ë„ ì„ê³„ê°’ì´ í•„ìš”í•©ë‹ˆë‹¤ (B15, B16)")
        
        # ê³µì„¸ëª¨ë“œ íšŒì°¨ë³„ ë¹„ì¤‘ ì½ê¸° (B21~B28)
        ag_split_ratios = []
        print(f"ğŸ“Š ê³µì„¸ëª¨ë“œ íšŒì°¨ë³„ ë¹„ì¤‘ ì½ê¸° (B21~B28):")
        for row in range(21, 29):  # B21 ~ B28
            cell_value = ws[f'B{row}'].value
            ratio = safe_float(cell_value)
            if ratio is not None:
                ag_split_ratios.append(ratio)
                print(f"   B{row}: {cell_value} â†’ {ratio}")
            else:
                print(f"   B{row}: ë¹ˆì¹¸")
        
        # ì•ˆì „ëª¨ë“œ íšŒì°¨ë³„ ë¹„ì¤‘ ì½ê¸° (B29~B36, ë¹ˆì¹¸ì´ ìˆìœ¼ë©´ ë§¤ìˆ˜í•˜ì§€ ì•ŠìŒ)
        sf_split_ratios = []
        print(f"ğŸ“Š ì•ˆì „ëª¨ë“œ íšŒì°¨ë³„ ë¹„ì¤‘ ì½ê¸° (B29~B36):")
        for row in range(29, 37):  # B29 ~ B36
            cell_value = ws[f'B{row}'].value
            ratio = safe_float(cell_value)
            if ratio is not None:
                sf_split_ratios.append(ratio)
                print(f"   B{row}: {cell_value} â†’ {ratio}")
            else:
                print(f"   B{row}: ë¹ˆì¹¸ (ì´í›„ ë¬´ì‹œ)")
                # ë¹ˆì¹¸ì´ ìˆìœ¼ë©´ ê·¸ ì´í›„ëŠ” ë¬´ì‹œ (ë§¤ìˆ˜í•˜ì§€ ì•ŠìŒ)
                break
        
        # ê³µì„¸ëª¨ë“œ ì„¤ì •
        ag_config = {
            "buy_threshold": ag_buy_threshold,
            "sell_threshold": ag_sell_threshold,
            "max_hold_days": ag_max_hold_days,
            "split_count": len(ag_split_ratios) if ag_split_count is None else ag_split_count,
            "split_ratios": ag_split_ratios
        }
        
        # ì•ˆì „ëª¨ë“œ ì„¤ì •
        sf_config = {
            "buy_threshold": sf_buy_threshold,
            "sell_threshold": sf_sell_threshold,
            "max_hold_days": sf_max_hold_days,
            "split_count": len(sf_split_ratios) if sf_split_count is None else sf_split_count,
            "split_ratios": sf_split_ratios
        }
        
        print("\nâœ… íŒŒë¼ë¯¸í„° ë¡œë“œ ì™„ë£Œ")
        print(f"   ê³µì„¸ëª¨ë“œ: ë§¤ìˆ˜ {ag_config['buy_threshold']}%, ë§¤ë„ {ag_config['sell_threshold']}%, ë³´ìœ ì¼ {ag_config['max_hold_days']}ì¼, ë¶„í•  {ag_config['split_count']}íšŒ")
        print(f"   ì•ˆì „ëª¨ë“œ: ë§¤ìˆ˜ {sf_config['buy_threshold']}%, ë§¤ë„ {sf_config['sell_threshold']}%, ë³´ìœ ì¼ {sf_config['max_hold_days']}ì¼, ë¶„í•  {sf_config['split_count']}íšŒ")
        print(f"   ê³µì„¸ëª¨ë“œ ë¹„ì¤‘ í•©: {sum(ag_split_ratios):.4f}")
        print(f"   ì•ˆì „ëª¨ë“œ ë¹„ì¤‘ í•©: {sum(sf_split_ratios):.4f}")
        
        wb.close()  # íŒŒì¼ ë‹«ê¸°
        
        return ag_config, sf_config
        
    except Exception as e:
        print(f"âŒ ì—‘ì…€ íŒŒì¼ ì½ê¸° ì˜¤ë¥˜: {e}")
        import traceback
        traceback.print_exc()
        print("\nğŸ’¡ íŒ:")
        print("   1. ì—‘ì…€ íŒŒì¼ì´ ì—´ë ¤ìˆìœ¼ë©´ ë‹«ê³  ë‹¤ì‹œ ì‹œë„í•´ì£¼ì„¸ìš”.")
        print("   2. ì—‘ì…€ íŒŒì¼ì„ ì €ì¥í–ˆëŠ”ì§€ í™•ì¸í•´ì£¼ì„¸ìš”.")
        print("   3. íŒŒì¼ ê²½ë¡œê°€ ì˜¬ë°”ë¥¸ì§€ í™•ì¸í•´ì£¼ì„¸ìš”.")
        raise


def calculate_mdd(daily_records):
    """
    ìµœëŒ€ ë‚™í­(MDD) ê³„ì‚°
    Args:
        daily_records: ì¼ë³„ ê¸°ë¡ ë¦¬ìŠ¤íŠ¸
    Returns:
        dict: MDD ì •ë³´
    """
    if not daily_records:
        return {"mdd_percent": 0.0, "mdd_date": "", "mdd_value": 0.0}
    
    peak = daily_records[0]['total_assets']
    max_drawdown = 0.0
    mdd_date = ""
    mdd_value = 0.0
    
    for record in daily_records:
        total_assets = record['total_assets']
        if total_assets > peak:
            peak = total_assets
        
        drawdown = (peak - total_assets) / peak * 100
        if drawdown > max_drawdown:
            max_drawdown = drawdown
            mdd_date = record['date']
            mdd_value = total_assets
    
    return {
        "mdd_percent": max_drawdown,
        "mdd_date": mdd_date,
        "mdd_value": mdd_value
    }


def main():
    """ë©”ì¸ ì‹¤í–‰ í•¨ìˆ˜"""
    print("=" * 60)
    print("ğŸš€ íŒŒë¼ë¯¸í„° ê¸°ë°˜ ë°±í…ŒìŠ¤íŒ…")
    print("=" * 60)
    
    # ì—‘ì…€ íŒŒì¼ì—ì„œ íŒŒë¼ë¯¸í„° ì½ê¸°
    try:
        ag_config, sf_config = load_parameters_from_excel("íŒŒë¼ë¯¸í„°.xlsx")
    except Exception as e:
        print(f"âŒ íŒŒë¼ë¯¸í„° ë¡œë“œ ì‹¤íŒ¨: {e}")
        return
    
    # ê¸°ë³¸ ì„¤ì •ê°’
    initial_capital = 40000  # íˆ¬ìì›ê¸ˆ 4ë§Œ ë‹¬ëŸ¬
    start_date = "2011-01-01"  # íˆ¬ìì‹œì‘ì¼
    end_date = "2025-12-07"  # íˆ¬ìì¢…ë£Œì¼
    
    print(f"\nğŸ’° íˆ¬ìì›ê¸ˆ: ${initial_capital:,.0f}")
    print(f"ğŸ“… íˆ¬ìê¸°ê°„: {start_date} ~ {end_date}")
    
    # íŠ¸ë ˆì´ë” ì´ˆê¸°í™” (íŒŒë¼ë¯¸í„° ì ìš©)
    print("\nğŸ”„ íŠ¸ë ˆì´ë” ì´ˆê¸°í™” ì¤‘...")
    print(f"   ì ìš©ë  ê³µì„¸ëª¨ë“œ ì„¤ì •: {ag_config}")
    print(f"   ì ìš©ë  ì•ˆì „ëª¨ë“œ ì„¤ì •: {sf_config}")
    trader = SOXLQuantTrader(
        initial_capital=initial_capital,
        sf_config=sf_config,
        ag_config=ag_config
    )
    
    # íŠ¸ë ˆì´ë”ì— ì‹¤ì œë¡œ ì ìš©ëœ ì„¤ì • í™•ì¸
    print(f"\nâœ… íŠ¸ë ˆì´ë” ì´ˆê¸°í™” ì™„ë£Œ")
    print(f"   ì‹¤ì œ ê³µì„¸ëª¨ë“œ ì„¤ì •: {trader.ag_config}")
    print(f"   ì‹¤ì œ ì•ˆì „ëª¨ë“œ ì„¤ì •: {trader.sf_config}")
    
    # RSI ì°¸ì¡° ë°ì´í„° í™•ì¸
    print("\nğŸ” ì£¼ê°„ RSI ë°ì´í„° í™•ì¸ ì¤‘...")
    try:
        rsi_ref_data = trader.load_rsi_reference_data()
        if not rsi_ref_data or len(rsi_ref_data) == 0:
            print("âŒ ì£¼ê°„ RSI ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤!")
            print("   ë°±í…ŒìŠ¤íŒ…ì„ ì‹¤í–‰í•˜ê¸° ì „ì— ì£¼ê°„ RSI ë°ì´í„°ë¥¼ ì—…ë°ì´íŠ¸í•´ì£¼ì„¸ìš”.")
            print("   ì›¹ì•±ì„ ì‹¤í–‰í•˜ê±°ë‚˜ update_rsi_data.pyë¥¼ ì‹¤í–‰í•˜ì—¬ RSI ë°ì´í„°ë¥¼ ìƒì„±í•˜ì„¸ìš”.")
            return
        
        # ì‹œì‘ ë‚ ì§œì˜ ì£¼ì°¨ RSI í™•ì¸
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        # ì‹œì‘ì¼ì´ ì†í•œ ì£¼ì˜ ê¸ˆìš”ì¼ ê³„ì‚° (soxl_quant_system.pyì™€ ë™ì¼í•œ ë¡œì§)
        days_until_friday = (4 - start_dt.weekday()) % 7  # ê¸ˆìš”ì¼(4)ê¹Œì§€ì˜ ì¼ìˆ˜
        if days_until_friday == 0 and start_dt.weekday() != 4:  # ê¸ˆìš”ì¼ì´ ì•„ë‹Œë° ê³„ì‚°ì´ 0ì´ë©´ ë‹¤ìŒ ì£¼ ê¸ˆìš”ì¼
            days_until_friday = 7
        start_week_friday = start_dt + timedelta(days=days_until_friday)
        
        # ì‹œì‘ ì£¼ì°¨ì˜ RSI í™•ì¸
        start_rsi = trader.get_rsi_from_reference(start_week_friday, rsi_ref_data)
        if start_rsi is None:
            print(f"âŒ ë°±í…ŒìŠ¤íŒ… ì‹œì‘ì¼({start_date})ì˜ ì£¼ê°„ RSI ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤!")
            print(f"   í•„ìš”í•œ ì£¼ì°¨: {start_week_friday.strftime('%Y-%m-%d')} (ê¸ˆìš”ì¼)")
            print("   ë°±í…ŒìŠ¤íŒ…ì„ ì‹¤í–‰í•˜ê¸° ì „ì— ì£¼ê°„ RSI ë°ì´í„°ë¥¼ ì—…ë°ì´íŠ¸í•´ì£¼ì„¸ìš”.")
            print("   ì›¹ì•±ì„ ì‹¤í–‰í•˜ê±°ë‚˜ update_rsi_data.pyë¥¼ ì‹¤í–‰í•˜ì—¬ RSI ë°ì´í„°ë¥¼ ìƒì„±í•˜ì„¸ìš”.")
            return
        
        # 1ì£¼ì „, 2ì£¼ì „ RSIë„ í™•ì¸ (ëª¨ë“œ ê²°ì •ì— í•„ìš”)
        prev_week_friday = start_week_friday - timedelta(days=7)
        two_weeks_ago_friday = start_week_friday - timedelta(days=14)
        
        prev_rsi = trader.get_rsi_from_reference(prev_week_friday, rsi_ref_data)
        two_weeks_ago_rsi = trader.get_rsi_from_reference(two_weeks_ago_friday, rsi_ref_data)
        
        if prev_rsi is None or two_weeks_ago_rsi is None:
            print(f"âŒ ë°±í…ŒìŠ¤íŒ… ì‹œì‘ì¼({start_date})ì˜ 1ì£¼ì „ ë˜ëŠ” 2ì£¼ì „ RSI ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤!")
            if prev_rsi is None:
                print(f"   ëˆ„ë½ëœ ì£¼ì°¨: {prev_week_friday.strftime('%Y-%m-%d')} (1ì£¼ì „ ê¸ˆìš”ì¼)")
            if two_weeks_ago_rsi is None:
                print(f"   ëˆ„ë½ëœ ì£¼ì°¨: {two_weeks_ago_friday.strftime('%Y-%m-%d')} (2ì£¼ì „ ê¸ˆìš”ì¼)")
            print("   ë°±í…ŒìŠ¤íŒ…ì„ ì‹¤í–‰í•˜ê¸° ì „ì— ì£¼ê°„ RSI ë°ì´í„°ë¥¼ ì—…ë°ì´íŠ¸í•´ì£¼ì„¸ìš”.")
            print("   ì›¹ì•±ì„ ì‹¤í–‰í•˜ê±°ë‚˜ update_rsi_data.pyë¥¼ ì‹¤í–‰í•˜ì—¬ RSI ë°ì´í„°ë¥¼ ìƒì„±í•˜ì„¸ìš”.")
            return
        
        print(f"âœ… ì£¼ê°„ RSI ë°ì´í„° í™•ì¸ ì™„ë£Œ")
        print(f"   ì‹œì‘ ì£¼ì°¨ ({start_week_friday.strftime('%Y-%m-%d')}): RSI {start_rsi:.2f}")
        print(f"   1ì£¼ì „ ({prev_week_friday.strftime('%Y-%m-%d')}): RSI {prev_rsi:.2f}")
        print(f"   2ì£¼ì „ ({two_weeks_ago_friday.strftime('%Y-%m-%d')}): RSI {two_weeks_ago_rsi:.2f}")
        
    except Exception as e:
        print(f"âŒ RSI ë°ì´í„° í™•ì¸ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {e}")
        print("   ë°±í…ŒìŠ¤íŒ…ì„ ì‹¤í–‰í•˜ê¸° ì „ì— ì£¼ê°„ RSI ë°ì´í„°ë¥¼ ì—…ë°ì´íŠ¸í•´ì£¼ì„¸ìš”.")
        print("   ì›¹ì•±ì„ ì‹¤í–‰í•˜ê±°ë‚˜ update_rsi_data.pyë¥¼ ì‹¤í–‰í•˜ì—¬ RSI ë°ì´í„°ë¥¼ ìƒì„±í•˜ì„¸ìš”.")
        return
    
    # ë°±í…ŒìŠ¤íŒ… ì‹¤í–‰
    print("\nğŸ“Š ë°±í…ŒìŠ¤íŒ… ì‹¤í–‰ ì¤‘...")
    backtest_result = trader.run_backtest(start_date, end_date)
    
    if "error" in backtest_result:
        print(f"\nâŒ ë°±í…ŒìŠ¤íŒ… ì‹¤íŒ¨: {backtest_result['error']}")
        print("\nğŸ’¡ í•´ê²° ë°©ë²•:")
        print("   1. ì›¹ì•±ì„ ì‹¤í–‰í•˜ì—¬ RSI ë°ì´í„°ë¥¼ ìë™ìœ¼ë¡œ ì—…ë°ì´íŠ¸í•˜ì„¸ìš”.")
        print("   2. ë˜ëŠ” update_rsi_data.pyë¥¼ ì‹¤í–‰í•˜ì—¬ RSI ë°ì´í„°ë¥¼ ìˆ˜ë™ìœ¼ë¡œ ì—…ë°ì´íŠ¸í•˜ì„¸ìš”.")
        return
    
    # MDD ê³„ì‚°
    mdd_info = calculate_mdd(backtest_result['daily_records'])
    
    # ê²°ê³¼ ì¶œë ¥
    print("\n" + "=" * 60)
    print("ğŸ“Š ë°±í…ŒìŠ¤íŒ… ê²°ê³¼")
    print("=" * 60)
    print(f"ê¸°ê°„: {backtest_result['start_date']} ~ {backtest_result['end_date']}")
    print(f"ê±°ë˜ì¼ìˆ˜: {backtest_result['trading_days']}ì¼")
    print(f"ì´ˆê¸°ìë³¸: ${backtest_result['initial_capital']:,.0f}")
    print(f"ìµœì¢…ìì‚°: ${backtest_result['final_value']:,.0f}")
    print(f"ì´ìˆ˜ìµë¥ : {backtest_result['total_return']:+.2f}%")
    print(f"ìµœëŒ€ MDD: {mdd_info.get('mdd_percent', 0.0):.2f}%")
    print(f"MDD ë°œìƒì¼: {mdd_info.get('mdd_date', 'N/A')}")
    print(f"ìµœì¢…ë³´ìœ í¬ì§€ì…˜: {backtest_result['final_positions']}ê°œ")
    print(f"ì´ ê±°ë˜ì¼ìˆ˜: {len(backtest_result['daily_records'])}ì¼")
    
    # ì—°í‰ê·  ìˆ˜ìµë¥  ê³„ì‚°
    if backtest_result['trading_days'] > 0:
        years = backtest_result['trading_days'] / 252  # ì—°ê°„ ê±°ë˜ì¼ ì•½ 252ì¼
        if years > 0:
            annual_return = ((backtest_result['final_value'] / backtest_result['initial_capital']) ** (1 / years) - 1) * 100
            print(f"ì—°í‰ê·  ìˆ˜ìµë¥ : {annual_return:+.2f}%")
    
    print("=" * 60)


if __name__ == "__main__":
    main()
